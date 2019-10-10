from django.db import models
from django.db.models import Avg, Case, Count, F, Max, Min, Prefetch, Q, Sum, When
from django.db.models.signals import post_init
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ValidationError
from django.conf import settings
from .enums import Where_Identified_CHOICES

import logging
logger = logging.getLogger(__name__)

import mimetypes
from functools import lru_cache, partial
from decimal import Decimal
import decimal
import re

from mptt.models import MPTTModel, TreeForeignKey
import openpyxl

from . import grabbag

def make_random_filename(instance, filename):
    mt = mimetypes.guess_type(filename)
    file_ext = mimetypes.guess_extension(mt[0])

    return grabbag.make_random_code(code_length=16) + file_ext

def ou_dict_from_path(*ou_path, start_level=1):
    if start_level < 0:
        start_level = 0
    return dict(zip(OrgUnit.level_fields()[start_level:], ou_path))

def ou_path_from_dict(v_dict):
    return tuple((v_dict[f] for f in OrgUnit.level_fields() if f in v_dict))

fs = FileSystemStorage(location=settings.SOURCE_DOC_DIR)

class SourceDocument(models.Model):
    orig_filename = models.CharField(max_length=128, blank=True, null=True)
    file1 = models.FileField(upload_to=make_random_filename, storage=fs)
    uploaded_at = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        # store the original filename away for later
        self.orig_filename = self.file1.name
        super(SourceDocument, self).save(*args, **kwargs)

    def __str__(self):
        return '%s: %s' % (self.file1, self.orig_filename)

def orgunit_cleanup_name(name_str):
    """
    Convert name to DHIS2 standard form and fix any whitespace issues (leading,
    trailing or repeated)
    """
    if name_str:
        name_str = name_str.strip() # remove leading/trailing whitespace
        name_str = re.sub('\s+', ' ', name_str) # "compress" consecutive whitespace
        name_str = name_str.replace('H/C I', 'HC I') # matches 'HC II', 'HC III' and 'HC IV'
        name_str = name_str.replace('Health Centre I', 'HC I')
        name_str = name_str.replace('Health Center I', 'HC I')
        name_str = name_str.replace('HCIV', 'HC IV')
        name_str = name_str.replace('HCII', 'HC II') # matches HC II and HC III
    return name_str

class OrgUnit(MPTTModel):
    name = models.CharField(max_length=64, db_index=True)
    parent = TreeForeignKey('self', null=True, blank=True, related_name='children', db_index=True)

    class MPTTMeta:
        order_insertion_by = ['name']

    class Meta:
        unique_together = (('name', 'parent'),)
        verbose_name = 'organisation unit'

    @classmethod
    def from_path_str(cls, path, path_sep='/'):
        return cls.from_path(path.split(path_sep))

    @classmethod
    def from_path(cls, *path_parts):
        current_node = None
        for p in path_parts:
            p = orgunit_cleanup_name(p)
            if p:
                ou_p, created = cls.objects.get_or_create(name__iexact=str(p), parent=current_node, defaults={'name':str(p)})
                current_node = ou_p
            else:
                break # stop processing when you find blank/empty path component/name

        return current_node

    @classmethod
    @lru_cache(maxsize=1024)
    def from_path_recurse(cls, *path_parts):
        if len(path_parts) == 0:
            return None
        *parent_path, node_name = path_parts
        node_name = orgunit_cleanup_name(node_name)
        if len(parent_path) == 0:
            ou, created = cls.objects.get_or_create(name__iexact=node_name, parent=None, defaults={'name':node_name})
        else:
            ou_parent = cls.from_path_recurse(*parent_path)
            ou, created = cls.objects.get_or_create(name__iexact=node_name, parent=ou_parent, defaults={'name':node_name})
        return ou

    @staticmethod
    def level_names(max_level=None):
        if max_level is None:
            level_count = len(settings.ORG_UNIT_LEVELS)
        else:
            level_count = max_level + 1
        return tuple((settings.ORG_UNIT_LEVELS.get(i) for i in range(level_count) if settings.ORG_UNIT_LEVELS.get(i) is not None))

    @staticmethod
    def level_fields(max_level=None):
        #TODO: escape/replace any characters that would be an invalid field name
        return tuple(map(lambda x: x.lower().replace(' ', '_'), OrgUnit.level_names(max_level)))

    @staticmethod
    def level_dbfields(max_level=None, *ignore, prefix=''):
        if max_level is None:
            level_count = len(settings.ORG_UNIT_LEVELS)
        else:
            level_count = max_level + 1
        return tuple(reversed([prefix+(''.join(('parent__',)*i)+'name') for i in range(level_count)]))

    @staticmethod
    def level_annotations(max_level=None, *ignore, prefix=''):
        return dict(zip(OrgUnit.level_fields(max_level), [F(f) for f in OrgUnit.level_dbfields(max_level, prefix=prefix)]))

    @staticmethod
    def get_level_field(level):
        #TODO: escape/replace any characters that would be an invalid field name
        return settings.ORG_UNIT_LEVELS[level].lower().replace(' ', '_')

    def __str__(self):
        return '%s [parent_id: %s]' % (self.name, str(self.parent_id),)

class DataElement(models.Model):
    VALUE_TYPES = (
        ('NUMBER', 'Number'),
        ('INTEGER', 'Integer (whole numbers only)'),
        ('POS_INT', 'Positive Integer'),
        # ('PERCENT', 'Percentage'), # implies the Average aggregation method
        #TODO: for boolean types auto-create a (hidden?) category with two options corresponding to the labels we have in the boolean
        # ('BOOLEAN', 'Boolean (True/False)'), # two values (0/1) with default labels (category options?) of 'True'/'False'
        # ('CHOICE', 'Selection from fixed set') # implies the Count aggregation method
    )
    AGG_METHODS = (
        ('SUM', 'Sum()'),
        # ('COUNT', 'Count()'), # eg. we have a facility reporting patient aggregates but also a facility level field (has_incinerator)
        # ('AVG', 'Average()'), # needs a corresponding 'population' data element that provides the ratios when we combine two (or more) averages
    )
    
    dhis2_uid = models.CharField(max_length=11, blank=True, null=True) #TODO: add unique check and check for a minimum length of 11 as well
    name = models.CharField(max_length=160, unique=True, db_index=True)
    alias = models.CharField(max_length=128, blank=True, null=True, db_index=True)
    value_type = models.CharField(max_length=8, choices=VALUE_TYPES)
    value_min = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Minimum Value', blank=True, null=True)
    value_max = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Maximum Value', blank=True, null=True)
    aggregation_method = models.CharField(max_length=8, choices=AGG_METHODS)

    def validate_unique(self, exclude=None):
        super(DataElement, self).validate_unique(exclude=exclude)

        # name already exists as an alias
        if DataElement.objects.filter(Q(alias__iexact=self.name), ~Q(id=self.id)).exists():
            raise ValidationError({'name': 'Name already used as an alias: \'%s\'' % (self.name,)})
        if self.alias:
            # alias already exists as a name/alias
            if DataElement.objects.filter(Q(name__iexact=self.alias)|Q(alias__iexact=self.alias), ~Q(id=self.id)).exists():
                raise ValidationError({'alias': 'Alias already used as a name/alias: \'%s\'' % (self.alias,)})

    def save(self, *args, **kwargs):
        self.validate_unique()
        super(DataElement, self).save(*args, **kwargs)

    def __repr__(self):
        return 'DataElement<%s>' % (str(self),)

    def __str__(self):
        return '%s' % (self.name,)

class Category(models.Model):
    name = models.CharField(max_length=128, unique=True)

    class Meta:
        verbose_name_plural = 'categories'

    def __str__(self):
        return self.name

class CategoryCombo(models.Model):
    name = models.CharField(max_length=512)
    categories = models.ManyToManyField(Category)

    @classmethod
    def from_cat_names(cls, cat_names):
        sorted_names = sorted(cat_names) #TODO: sort based on the name of the classification the Category belongs to
        cat_list = [Category.objects.get_or_create(name__iexact=cat_name, defaults={'name':cat_name})[0] for cat_name in sorted_names]
        cc_name = '(%s)' % ', '.join(sorted_names)
        cat_combo, created = cls.objects.get_or_create(name__iexact=cc_name, defaults={'name':cc_name})
        if created:
            for categ in cat_list:
                cat_combo.categories.add(categ)
            cat_combo.save()

        return cat_combo

    def __str__(self):
        return self.name


# TODO: Consider tracking which data element each subcategory is from (reduce false matches and other? benefits)
CATEGORIES = [
    'Male',
    'Female',
    'Outreach',
    'Death',
    '18 Mths-<5 Years',
    '5-<10 Years',
    '10-<15 Years',
    '15-<19 Years',
    '19-<49 Years',
    '>49 Years',
    '10-19 Years',
    '20-24 Years',
    '>=25 Years',

    '<2 Years',
    '2 - < 5 Years (HIV Care)',
    '5 - 14 Years',
    '< 15 Years',
    '15 Years and above',
    # HMIS 106a: 1B ART QUARTERLY COHORT ANALYSIS REPORT: FOLLOW-UP
    # 'Alive on ART in Cohort',
    # 'Died',
    # 'Lost  to Followup',
    # 'Lost to Followup',
    # 'Lost',
    # 'Started on ART-Cohort',
    # 'Stopped',
    # 'Transfered In',
    # 'Transferred Out',
    '0-4 Years',
    '5-14 Yrs',
    # HMIS 105: 2 MNCH
    'Under 1',
    '1-4 Yrs',
    '6-11 Months',
    '12-59 Months',
    '1-4 Years',
    '5-14 Years',
    'Static',
    # HMIS 105: 1.3 OPD
    '0-28 Days',
    '29 Days-4 Years',
    '5-59 Years',
    '60andAbove Years',

    # VMMC
    '2<5 Years',
    '5-<15 Years',
    '15-49 Years',

    # HMIS 105 Laboratory
    'Under 5 years',
    '5 years and above',

    # HMIS 105 Family Planning
    'Group I',
    'Group J',
    'Group O',
    'Group Z',
    'New Users',
    'Revisits',

    # HMIS 106a Nutrition
    '< 6 Months',
    '6-59 Months',
    '5-12 Years',
    '12+ Years',

    # HMIS 108
    'Case',

    '<15',
    '15+',

    # GEND_GBV TARGET: GBV Care, from USAID
    '<10',
    '10-14',
    '15-17',
    '18-19',
    '20-24',
    '25-49',
    '50+',

    # TX_CURR TARGET, from USAID
    '<1',
    '<1-9',
    '1-9',
    '10-14',
    '15-19',

    # TX_PVLS TARGET: 12 Months Viral Load < 1000, from USAID
    '25-40'
]

SEP_REGEX_STR = '[\s,]+' # one or more of these characters in sequence
CATEGORY_REGEX_STR = '|'.join('%s?(%s)(?:[^\w]|$)' % (SEP_REGEX_STR, re.escape(categ)) for categ in sorted(CATEGORIES, key=lambda x: (len(x), x), reverse=True))
CATEGORY_REGEX = re.compile(CATEGORY_REGEX_STR)
SEXLESS_CATEGORY_REGEX_STR = '|'.join('%s?(%s)(?:[^\w]|$)' % (SEP_REGEX_STR, re.escape(categ)) for categ in sorted(CATEGORIES[4:], key=lambda x: (len(x), x), reverse=True)) #TODO: even more horrible a hack
SEXLESS_CATEGORY_REGEX = re.compile(SEXLESS_CATEGORY_REGEX_STR)

ICKY_CATEGS = (
    'Number of Male',
    'Male partners',
    'Female Condom',
    'Male Condom',
    'Device Based (DC)',
    'Surgical(SC)',
    'OPD Death',
)

def unpack_data_element(de_long):
    if any([de_long.upper().find(s.upper()) >=0 for s in ICKY_CATEGS]):
        m = SEXLESS_CATEGORY_REGEX.split(de_long)
    else:
        m = CATEGORY_REGEX.split(de_long)
    # squash list of matches by removing blank and None entries (and False and numeric zeroes)
    de_name, *category_list = tuple(filter(None, m))
    category_list = [c.strip() for c in category_list]
    cat_str = ', '.join(category_list)

    # deals with cases where the data element name includes a subcategory ('105-2.1a Male partners received HIV test results in eMTCT')
    # and matches multiple subcategories ('Lost' and 'Lost  to Followup' in '106a Cohort  All patients 12 months Lost  to Followup')
    #TODO: reimplement this, it is a really ugly hack
    if any([cat not in CATEGORIES for cat in category_list]):
        cat_str = ' '.join(category_list)
        if cat_str not in CATEGORIES:
            de_name = de_long
            cat_str = ''
            category_list = () # empty tuple

    field_info = DataElement._meta.fields[2].deconstruct()
    if field_info[0] == 'name':
        MAX_NAME_LEN = field_info[3]['max_length']
    de_instance, created = DataElement.objects.get_or_create(name__iexact=de_name[:MAX_NAME_LEN], value_type='NUMBER', value_min=None, value_max=None, aggregation_method='SUM', defaults={'name':de_name[:MAX_NAME_LEN]})
    if len(category_list):
        return (de_instance, CategoryCombo.from_cat_names(category_list))
    else:
        return (de_instance, None)

from django.db.models import Func

class Collate_C(Func):
    function = 'COLLATE "C"'
    template = '%(expressions)s %(function)s'

class DataValueQuerySet(models.QuerySet):
    """Convenience queryset methods for handling datavalues"""
    def what(self, *names):
        de_filters = None
        if names:
            for de in names:
                if de is None:
                    continue # skip any names/uids with value of None
                if de_filters:
                    de_filters = de_filters | Q(data_element__name__iexact=de) | Q(data_element__alias__iexact=de)
                else:
                    de_filters = Q(data_element__name__iexact=de) | Q(data_element__alias__iexact=de)

        qs = self.annotate(de_name=Collate_C('data_element__name'))
        qs = qs.annotate(de_uid=F('data_element__dhis2_uid'))
        if de_filters:
            qs = qs.filter(de_filters)
        return qs

    def where(self, *names_or_objects):
        import functools
        import operator
        if isinstance(names_or_objects[0], OrgUnit):
            # we were passed a list of OrgUnit instances
            orgunits = names_or_objects
        else:
            # we were passed a list of OrgUnit names
            ou_filters = [Q(name__iexact=ou_name) for ou_name in names_or_objects if ou_name is not None]
            ou_filters_combined = functools.reduce(operator.__or__, ou_filters)
            orgunits = list(OrgUnit.objects.filter(ou_filters_combined))
            if len(orgunits) == 0:
                return self.none()

        qs = self
        for ou in orgunits:
            qs = qs.filter(Q(org_unit__lft__gte=ou.lft) & Q(org_unit__rght__lte=ou.rght))
        return qs

    def when(self, *periods):
        PERIOD_REGEX = re.compile(r'(\d{4}-\d{2})|(\d{4}-?Q\d{1})|(\d{4})')

        qs = self
        period_filters = None
        for p in periods:
            res = PERIOD_REGEX.match(p)
            if res is None:
                continue

            (g_month, g_quarter, g_year) = res.group(1, 2, 3)
            if g_month:
                if period_filters:
                    period_filters = period_filters | Q(month=g_month)
                else:
                    period_filters = Q(month=g_month)
                qs = qs.annotate(period=F('month'))
            if g_quarter:
                if period_filters:
                    period_filters = period_filters | Q(quarter=g_quarter)
                else:
                    period_filters = Q(quarter=g_quarter)
                qs = qs.annotate(period=F('quarter'))
            if g_year:
                if period_filters:
                    period_filters = period_filters | Q(year=g_year)
                else:
                    period_filters = Q(year=g_year)
                qs = qs.annotate(period=F('year'))
        if period_filters:
            qs = qs.filter(period_filters)
        #TODO: Can we reduce the annotate calls to just one? Should we?
        return qs

class DataValueManager(models.Manager):
    """Attach our custom queryset methods to the model manager"""
    def get_queryset(self):
        return DataValueQuerySet(self.model, using=self._db)

    def what(self, *names):
        return self.get_queryset().what(*names)

    def where(self, *names_or_objects):
        return self.get_queryset().where(*names_or_objects)

    def when(self, *periods):
        return self.get_queryset().where(*periods)

def get_default_category_combo():
    return CategoryCombo.objects.get(id=1)

class DataValue(models.Model):
    data_element = models.ForeignKey(DataElement, related_name='data_values')
    category_combo = models.ForeignKey(CategoryCombo, related_name='data_values', default=1)
    site_str = models.CharField(max_length=128)
    org_unit = models.ForeignKey(OrgUnit, related_name='data_values')
    numeric_value = models.DecimalField(max_digits=17, decimal_places=4)
    month = models.CharField(max_length=7, blank=True, null=True, db_index=True) # ISO 8601 format '2017-09'
    quarter = models.CharField(max_length=7, blank=True, null=True, db_index=True) # ISO 8601 format '2017-Q3'
    year = models.CharField(max_length=4, blank=True, null=True, db_index=True) # ISO 8601 format '2017'
    source_doc = models.ForeignKey(SourceDocument, related_name='data_values')

    objects = DataValueManager() # override the default manager

    class Meta():
        unique_together = (('data_element', 'category_combo', 'org_unit', 'year', 'quarter', 'month'),)

    def __repr__(self):
        return 'DataValue<%s [%s], %s, %s, %d>' % (str(self.data_element), self.category_combo, self.site_str,  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

    def __str__(self):
        return '%s [%s], %s, %s, %d' % (str(self.data_element), self.category_combo, self.site_str.split(' => ')[-1],  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

@lru_cache(maxsize=16) # memoize to reduce cost of "parsing"
def extract_periods(period_str):
    from .grabbag import period_to_dates, dates_to_iso_periods
    dates = period_to_dates(period_str)
    return dates_to_iso_periods(*dates)

def load_excel_to_datavalues(source_doc):
    from django.db import connection
    from collections import defaultdict
    import calendar

    from .grabbag import MONTH_TO_MONTH_REGEX

    MONTH_PREFIX_REGEX = re.compile(r'^[\s]*(%s) ([0-9]{4})?[\s]*' % ('|'.join(calendar.month_name[1:]),))

    DE_COLUMN_START = 4 # 0-based index of first dataelement column in worksheet

    db_cursor = connection.cursor()

    wb = openpyxl.load_workbook(source_doc.file1.path, read_only=True)
    logger.debug(wb.get_sheet_names())

    for ws_name in wb.get_sheet_names():
        if ws_name in ['Validations']:
            continue
        ws = wb[ws_name]
        if ws.calculate_dimension() == 'A1:A1': # check for (one kind of) invalid dimensions
            ws.max_row = ws.max_column = None
        logger.debug((ws_name, ws.max_row, ws.max_column))

        iter_rows = iter(ws.rows)
        first_row = next(iter_rows)
        headers = [cell.value for cell in first_row]
        clean_headers = (MONTH_TO_MONTH_REGEX.sub('', MONTH_PREFIX_REGEX.sub('', h)).lstrip() for h in headers[DE_COLUMN_START:] if h is not None)
        data_elements = tuple(unpack_data_element(de) for de in clean_headers)


        for row in iter_rows:
            period_cell, *location_cells = row[:DE_COLUMN_START]
            location_parts = [c.value for c in location_cells]
            if not period_cell.value or not any(location_parts):
                continue # ignore rows where period or location is missing
            if period_cell.is_date:
                # convert to ISO 8601 month notation
                period = '{0.year}-{0.month:02}'.format(period_cell.value)
            else:
                period = period_cell.value
            iso_year, iso_quarter, iso_month = extract_periods(str(period).strip())
            location_parts = [x.strip() for x in location_parts if x is not None]
            DISTRICT_SUFFIX_REGEX = re.compile(r'\s*[Dd]istrict\s*$')
            if DISTRICT_SUFFIX_REGEX.search(location_parts[0]) is None:
                location_parts[0] = location_parts[0] + ' District'
            location_parts = (settings.ORG_UNIT_ROOT_NAME, *location_parts) # turn to tuple and prepend name of root OrgUnit
            current_ou = OrgUnit.from_path_recurse(*location_parts)
            location = ' => '.join(location_parts)
            logger.debug((period, location))

            site_val_cells = row[DE_COLUMN_START:]
            site_values = zip(data_elements, (c.value for c in site_val_cells))
            where_when = {
                'site_str': location,
                'org_unit': current_ou,
                'month': iso_month,
                'quarter': iso_quarter,
                'year': iso_year,
            }

            for (de, cc), dv in site_values:
                if dv is None or (isinstance(dv, str) and dv.strip() == ''):
                    continue # skip rows with empty values
                try:
                    if cc:
                        cc_id = cc.id
                    else:
                        cc_id = 1
                    conflict_fields = ['data_element_id', 'category_combo_id', 'org_unit_id']
                    maybe_null = list(zip(('year', 'quarter', 'month'), (iso_year, iso_quarter, iso_month)))
                    conflict_fields += [f[0] for f in maybe_null if f[1]]
                    conflict_condition = ' AND '.join(['{0} IS NULL'.format(f[0]) for f in maybe_null if not f[1]])
                    if conflict_condition:
                        conflict_condition = 'WHERE ' + conflict_condition
                    
                    upsert_sql = '''INSERT INTO cannula_datavalue (data_element_id, category_combo_id, org_unit_id, site_str, year, quarter, month, source_doc_id, numeric_value)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT ({0}) {1} DO UPDATE SET source_doc_id=%s, numeric_value=%s
                    '''.format(', '.join(conflict_fields), conflict_condition)
                    what_where_when = (de.id, cc_id, current_ou.id, location, iso_year, iso_quarter, iso_month)
                    db_cursor.execute(upsert_sql, (what_where_when+(source_doc.id, Decimal(dv))*2))
                except decimal.InvalidOperation as e:
                    pass # not convertible to a Decimal, ignore


    return dict() # TODO: remove this and update callers

def de_pivot_col(de):
    return 'DE_%d' % (de.id,)

def pivot_clause(data_elements):
    return ',\n'.join("SUM(CASE WHEN de.name = '%s' THEN dv.numeric_value ELSE 0 END) as '%s'" % (de.name, de_pivot_col(de)) for de in data_elements)

def validation_expr(left, right, operator):
    pass

def validation_expr_elements(expr):
    names = list()
    for de_tup in DataElement.objects.all().values_list('name', 'alias'):
        names.extend(de_tup)
    names = filter(None, names)
    sorted_names = list(sorted(names, key=len, reverse=True)) # sort puts longest matches first
    DE_REGEX = '|'.join('%s' % (re.escape(de_name),) for de_name in sorted_names)
    m = re.findall(DE_REGEX, expr, flags=re.IGNORECASE)
    print(m)
    return tuple(filter(None, m))

def load_excel_to_validations(source_doc):
    wb = openpyxl.load_workbook(source_doc.file1.path) #TODO: ensure we close the workbook file. use a context manager?
    logger.debug(wb.get_sheet_names())

    for ws_name in wb.get_sheet_names():
        if ws_name != 'Validations':
            continue

        ws = wb[ws_name]
        logger.debug((ws_name, ws.max_row, ws.max_column))

        
        for row in ws.rows[1:]: # skip header row
            validation_name, l_exp, op, r_exp, *_ = [c.value for c in row]
            if not l_exp or not op or not r_exp:
                continue # ignore rows where any part of the rule is missing
            op = op.replace(' ', '')
            if op not in ('>', '>=', '=', '<', '<='):
                continue # ignore rows with an invalid operator
            l_exp, op, r_exp = str(l_exp), str(op), str(r_exp) # ensure expressions and operator are strings
            print(validation_name, l_exp, op, r_exp)
            l_element_names = validation_expr_elements(l_exp)
            r_element_names = validation_expr_elements(r_exp)
            element_names = l_element_names + r_element_names
            try:
                vr = ValidationRule.objects.get(name=validation_name)
                vr.left_expr, vr.right_expr, vr.operator = l_exp, r_exp, op
            except ValidationRule.DoesNotExist as e:
                vr = ValidationRule(name=validation_name, left_expr=l_exp, right_expr=r_exp, operator=op)
            vr.save()
            print(vr.view_name())

    return

    ou_level = month_multiple = None
    search_periods = ['2016']
    de_meta_list = query_de_meta(tt_names)
    print(de_meta_list)
    calc_exprs = (
        ('DE_5*100/DE_22', ['DE_22',]),
        ('DE_6*100/DE_22', ['DE_22',]),
    )
    if ou_level is None: # if not explicitly given use the highest orgunit level
        ou_level = min(map(lambda x: x.ou_level, de_meta_list))
    if month_multiple is None:
        month_multiple = max(map(lambda x: x.month_multiple, de_meta_list))
    calc_query = mk_calculation_sql(calc_exprs, de_meta_list, [], ou_level, search_periods, month_multiple)
    print(calc_query)

def mk_validation_rule_sql(rule_expr, data_elements):
    de_meta_list = query_de_meta(data_elements)
    ou_level = min(map(lambda x: x.ou_level, de_meta_list))
    month_multiple = max(map(lambda x: x.month_multiple, de_meta_list))

    subst_rule_expr = rule_expr
    for de_meta in sorted(de_meta_list, key=lambda x: x.name, reverse=True):
        subst_rule_expr = re.sub('(?i)' + re.escape(de_meta.name), 'DE_%d' % (de_meta.id,), subst_rule_expr)
        if de_meta.alias:
            subst_rule_expr = re.sub('(?i)' + re.escape(de_meta.alias), 'DE_%d' % (de_meta.id,), subst_rule_expr)

    return mk_calculation_sql([(subst_rule_expr, [])], de_meta_list, [], ou_level, [], month_multiple)

def query_de_meta(de_names):
    """
    Given a sequence of dataelement names, return a corresponding sequence of
    tuples containing the name, id, highest orgunit level it is collected at,
    and the largest period type it is collected for (as a multiple of month)
    """
    from functools import reduce
    from collections import namedtuple

    if len(de_names) == 0:
        return tuple()
    
    q_objs = reduce(lambda x, y: x | y, (Q(alias__iexact=de_name)|Q(name__iexact=de_name) for de_name in de_names))
    qs = DataElement.objects.filter(q_objs)
    qs = qs.annotate(ou_level=Min(F('data_values__org_unit__level')))
    qs = qs.annotate(month_multiple=Min(Case(When(data_values__month__isnull=False, then=1), When(data_values__quarter__isnull=False, then=4), When(data_values__year__isnull=False, then=12), default=None, output_field=models.IntegerField())))
    qs = qs.order_by('name', 'id', 'ou_level', 'month_multiple')
    
    DataElementMeta = namedtuple('DataElementMeta', ['name', 'alias', 'id', 'ou_level', 'month_multiple'])
    return tuple(DataElementMeta(**v) for v in qs.values('name', 'alias', 'id', 'ou_level', 'month_multiple'))

def fields_for_month_multiple(month_mul):
    return ('year', 'quarter', 'month')[:(12, 3, 1).index(month_mul)+1]

def gen_pairs(iterable):
    item2_iter = iter(iterable)
    _ = next(item2_iter) # skip one ahead
    for item1 in iterable:
        yield (item1, next(item2_iter))

def mk_de_group_sql(de_meta_list, all_fields, ou_level):
    select_clause = ' '.join(['SELECT', ', '.join(all_fields)])
    
    _tables = ['cannula_datavalue dv', 'cannula_orgunit ou', 'cannula_dataelement de'] + ['cannula_orgunit ou%d' % i for i in range(ou_level+1)]
    from_clause = ' '.join(['FROM', ', '.join(_tables)])

    join_filter_subclause = ' AND '.join(('dv.org_unit_id=ou.id', 'dv.data_element_id=de.id'))
    ou_traversals = ['ou0.parent_id IS NULL'] + ['ou%d.parent_id=ou%d.id' % (c, p) for p,c in gen_pairs(range(ou_level+1))]
    ou_traversals.append('ou.id = ou%d.id' % (ou_level,))
    de_filters = ['de.name=\'%s\'' % (de_m.name,) for de_m in de_meta_list] #TODO: switch to data element IDs/UIDs/CODEs to avoid SQL injection
    de_filter_clause = '(%s)' % ('\nOR '.join(de_filters),)
    where_parts = [join_filter_subclause, *ou_traversals, de_filter_clause]
    where_clause = 'WHERE ' + ('\nAND '.join(where_parts))

    return '\n'.join([select_clause, from_clause, where_clause])

def mk_union_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    from itertools import groupby

    ou_fields = OrgUnit.level_fields(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    print(ou_fields, period_fields)

    hier_ou_pairs = tuple(('ou%d' % (i,), f) for i, f in enumerate(ou_fields))
    hier_ou_fields = tuple('%s.name as %s' % (code, desc) for code, desc in hier_ou_pairs)

    placeholder_fields = ', '.join(['NULL as %s' % (f,) for f in (period_fields + ou_fields)])
    union_parts = ['SELECT %s, NULL as de_name, NULL as numeric_value FROM cannula_datavalue dv' % (placeholder_fields,)]

    grouped_de_metas = groupby(de_meta_list, lambda x: (x.ou_level, x.month_multiple))
    for g in grouped_de_metas:
        g_ident, g_seq = g
        g_seq = list(g_seq)
        g_ou_level, g_month_multiple = g_ident

        if g_month_multiple > period_month_multiple:
            my_period_fields = fields_for_month_multiple(g_month_multiple)
            my_periods = [tuple(filter(None, grabbag.dates_to_iso_periods(*grabbag.period_to_dates(p)))) for p in period_list]
            my_periods = [tuple('\'%s\'' % (p,) for p in p_tup) for p_tup in my_periods] #TODO: do proper db quoting
            print(my_periods)
            for p_tup in my_periods:
                all_fields = my_period_fields + p_tup[len(my_period_fields):] + hier_ou_fields + ('de.name as de_name' , 'numeric_value')
                group_select = mk_de_group_sql(g_seq, all_fields, g_ou_level)
                union_parts.append(group_select)
        else:
            my_period_fields = period_fields
            all_fields = my_period_fields + hier_ou_fields + ('de.name as de_name' , 'numeric_value')
            group_select = mk_de_group_sql(g_seq, all_fields, g_ou_level)
            union_parts.append(group_select)

    return '\nUNION ALL\n'.join(union_parts)

def mk_aggregate_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    union_sql = mk_union_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = OrgUnit.level_fields(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    groupby_fields = period_fields+ou_fields+('de_name',)
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+('sum(numeric_value) as numeric_sum', 'count(numeric_value) as numeric_count'))])
    group_order_clause = 'AS q_aggregate\nGROUP BY %s\nORDER BY %s' % (groupby_fields_str, groupby_fields_str)

    aggregate_sql = select_clause + '\n' + 'FROM (' + '\n' + union_sql + '\n' + ') ' + group_order_clause

    return aggregate_sql

def mk_pivot_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    aggregate_sql = mk_aggregate_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = OrgUnit.level_fields(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    pivot_fields = list()
    for de in de_meta_list:
        if de.month_multiple <= period_month_multiple:
            de_pivot_str = 'SUM(CASE WHEN de_name = \'%s\' THEN numeric_sum ELSE 0 END) as DE_%d' % (de.name, de.id)
        else:
            de_pivot_str = 'SUM(CASE WHEN de_name = \'%s\' THEN numeric_sum/%f ELSE 0 END) as DE_%d' % (de.name, de.month_multiple/period_month_multiple, de.id)
        pivot_fields.append(de_pivot_str)
    groupby_fields = period_fields + ou_fields
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+tuple(pivot_fields))])
    group_clause = 'AS q_pivot\nGROUP BY %s' % (groupby_fields_str)

    pivot_sql = select_clause + '\n' + 'FROM (' + '\n' + aggregate_sql + '\n' + ') ' + group_clause

    return pivot_sql

def mk_calc_fields(calculations):
    calc_fields = list()
    for i, (calc_exp, zero_checks) in enumerate(calculations, start=1):
        z_c_str = ' AND '.join('(%s != 0)' % (z_c_field) for z_c_field in zero_checks)
        if len(zero_checks) > 0:
            calc_str = 'CASE WHEN %s THEN %s ELSE NULL END as DE_CALC_%d' % (z_c_str, calc_exp, i)
        else:
            calc_str = '%s as DE_CALC_%d' % (calc_exp, i)
        calc_fields.append(calc_str)

    return tuple(calc_fields)

def mk_calculation_sql(calculations, de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    from collections import defaultdict

    print('OU_PARAM: %d, PERIOD_PARAM: %d' % (ou_level, period_month_multiple))
    
    pivot_sql = mk_pivot_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = OrgUnit.level_fields(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    calc_src_fields =  tuple('DE_%d' % (de.id,) for de in de_meta_list)
    calc_fields = mk_calc_fields(calculations)
    groupby_fields = period_fields + ou_fields
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+calc_src_fields+calc_fields)])

    where_groups = defaultdict(list)
    p_fields = fields_for_month_multiple(period_month_multiple)
    for p in period_list:
        p_vals = tuple(filter(None, grabbag.dates_to_iso_periods(*grabbag.period_to_dates(p))))
        for p_pair in zip(p_fields, p_vals):
            if p_pair not in where_groups[p_pair[0]]:
                where_groups[p_pair[0]].append(p_pair)
    where_parts = ['(%s)' % (' OR '.join('%s=\'%s\'' % (f, v) for f, v in l),) for k, l in where_groups.items()]

    if len(where_parts) > 0:
        where_clause = 'WHERE (%s)' % ' AND '.join(where_parts)
    else:
        where_clause = ''

    calculation_sql = select_clause + '\n' + 'FROM (' + '\n' + pivot_sql + '\n' + ') AS q_calculate' + '\n' + where_clause

    return calculation_sql

class ValidationRule(models.Model):
    name = models.CharField(max_length=128, unique=True)
    left_expr = models.CharField(max_length=256) #TODO: store the cleaned up expression with symbolic data element names
    right_expr = models.CharField(max_length=256)
    operator = models.CharField(max_length=2) #TODO: make this a choice field
    #TODO: add a description/comments field ?

    data_elements = models.ManyToManyField(DataElement)

    def expression(self):
        return ' '.join([self.left_expr, self.operator, self.right_expr])

    def view_name(self):
        return 'vw_validation_%d' % (self.id,)

    def instanciate(self):
        from django.db import connection

        # parse and collect data element names
        l_exp = self.left_expr.replace('\n', '')
        r_exp = self.right_expr.replace('\n', '')
        l_element_names = validation_expr_elements(l_exp)
        r_element_names = validation_expr_elements(r_exp)
        element_names = l_element_names + r_element_names

        remainder = l_exp
        for name in element_names:
            remainder = remainder.replace(name, '')
        remainder, l_number_count = re.subn(r'\d+[\.\d+]?', '', remainder) # remove numbers
        remainder = remainder.replace('+', '') # remove operators
        remainder = remainder.replace('-', '')
        remainder = remainder.replace('*', '')
        remainder = remainder.replace('/', '')
        remainder = re.sub(r'\s', '', remainder) # remove whitespace
        while True:
            remainder, substitution_count = re.subn(r'\(([^\)]*)\)', r'\1', remainder) # remove matching parentheses pairs
            if not substitution_count:
                break # no more parentheses pairs
        if len(remainder):
            print('REMAINDER: ', remainder)
            return # short-circuit, rule can be instanciated later
        remainder = r_exp
        for name in element_names:
            remainder = remainder.replace(name, '')
        remainder, r_number_count = re.subn(r'\d+[\.\d+]?', '', remainder) # remove numbers
        remainder = remainder.replace('+', '') # remove operators
        remainder = remainder.replace('-', '')
        remainder = remainder.replace('*', '')
        remainder = remainder.replace('/', '')
        remainder = re.sub(r'\s', '', remainder) # remove whitespace
        while True:
            remainder, substitution_count = re.subn(r'\(([^\)]*)\)', r'\1', remainder) # remove matching parentheses pairs
            if not substitution_count:
                break # no more parentheses pairs
        if len(remainder):
            print('REMAINDER: ', remainder)
            return # short-circuit, rule can be instanciated later
        if not(len(l_element_names) or l_number_count) and not (len(r_element_names) or r_number_count):
            return # short-circuit, rule can be instanciated later
        
        # modify list of data elements
        new_meta_list = query_de_meta(element_names)
        curr_meta_list = query_de_meta(self.data_elements.all().values_list('name'))
        for de_meta in curr_meta_list: # remove, if not in new list
            if de_meta not in new_meta_list:
                self.data_elements.remove(DataElement.objects.get(id=de_meta.id))
        for de_meta in new_meta_list: # add, if not in current list
            if de_meta not in curr_meta_list:
                self.data_elements.add(DataElement.objects.get(id=de_meta.id))

        # create the view
        sql = mk_validation_rule_sql(self.expression(), element_names)
        view_sql = 'CREATE OR REPLACE VIEW %s AS\n%s' % (self.view_name(), sql)
        cursor = connection.cursor()
        cursor.execute(view_sql, [])

    def save(self, *args, **kwargs):
        super(ValidationRule, self).save(*args, **kwargs)

        self.instanciate()

        super(ValidationRule, self).save(*args, **kwargs)

    def __str__(self):
        return self.name

def get_validation_view_names():
    from django.db import connection
    cursor = connection.cursor()
    cursor.execute('SELECT viewname FROM pg_catalog.pg_views WHERE viewowner=%s and viewname LIKE %s;', (settings.DATABASES['default']['USER'], 'vw_validation_%'))
    return [x[0] for x in cursor]

#sumplementary tools
class TbPrev(models.Model):
    period = models.CharField("YYYY-MM",max_length=200)
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)

    tbprevc1d = models.IntegerField("tb_prev (d,agesex)  _10-14 years, female",default=0)
    tbprevc2d = models.IntegerField("tb_prev (d,agesex)  __lt 10 years, female",default=0)
    tbprevc3d = models.IntegerField("tb_prev (d,agesex)  __lt15 yrs, female",default=0)
    tbprevc4d = models.IntegerField("tb_prev (d,agesex)  _15+ yrs, female",default=0)
    tbprevc5d = models.IntegerField("tb_prev (d,agesex)  _female, 15-17 years",default=0)
    tbprevc6d = models.IntegerField("tb_prev (d,agesex)  _female, 18-Â­24 years",default=0)
    tbprevc7d = models.IntegerField("tb_prev (d,agesex)  _female, 25+ yrs",default=0)
    tbprevc8d = models.IntegerField("tb_prev (d,agesex)  _10-14 years, male",default=0)
    tbprevc9d = models.IntegerField("tb_prev (d,agesex)  __lt 10 years, male",default=0)
    tbprevc10d = models.IntegerField("tb_prev (d,agesex)  __lt15 yrs, male",default=0)
    tbprevc11d = models.IntegerField("tb_prev (d,agesex)  _15+ yrs, male",default=0)
    tbprevc12d = models.IntegerField("tb_prev (d,agesex)  _male, 15-17 years",default=0)
    tbprevc13d = models.IntegerField("tb_prev (d,agesex)  _male, 18-Â­24 years",default=0)
    tbprevc14d = models.IntegerField("tb_prev (d,agesex)  _male, 25+ yrs",default=0)
    tbprevc15d = models.IntegerField("tb_prev (d,agesex)  _ipt by newly enrolled on art, female",default=0)
    tbprevc16d = models.IntegerField("tb_prev (d,agesex)  _ipt by newly enrolled on art, male",default=0)
    tbprevc17d = models.IntegerField("tb_prev (d,agesex)  _ipt by newly enrolled on art, total",default=0)
    tbprevc18d = models.IntegerField("tb_prev (d,agesex)  _ipt by previously enrolled on   art, female",default=0)
    tbprevc19d = models.IntegerField("tb_prev (d,agesex)  _ipt by previously enrolled on   art, male",default=0)
    tbprevc20d = models.IntegerField("tb_prev (d,agesex)  _ipt by previously enrolled on   art, total",default=0)
    tbprevc21d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, female",default=0)
    tbprevc22d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, male",default=0)
    tbprevc23d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, total",default=0)
    tbprevc24d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, female",default=0)
    tbprevc25d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, male",default=0)
    tbprevc26d = models.IntegerField("tb_prev (d,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, total",default=0)

    tbprevc1n = models.IntegerField("tb_prev (n,agesex)  _10-14 years, female",default=0)
    tbprevc2n = models.IntegerField("tb_prev (n,agesex)  __lt 10 years, female",default=0)
    tbprevc3n = models.IntegerField("tb_prev (n,agesex)  __lt15 yrs, female",default=0)
    tbprevc4n = models.IntegerField("tb_prev (n,agesex)  _15+ yrs, female",default=0)
    tbprevc5n = models.IntegerField("tb_prev (n,agesex)  _female, 15-17 years",default=0)
    tbprevc6n = models.IntegerField("tb_prev (n,agesex)  _female, 18-Â­24 years",default=0)
    tbprevc7n = models.IntegerField("tb_prev (n,agesex)  _female, 25+ yrs",default=0)
    tbprevc8n = models.IntegerField("tb_prev (n,agesex)  _10-14 years, male",default=0)
    tbprevc9n = models.IntegerField("tb_prev (n,agesex)  __lt 10 years, male",default=0)
    tbprevc10n = models.IntegerField("tb_prev (n,agesex)  __lt15 yrs, male",default=0)
    tbprevc11n = models.IntegerField("tb_prev (n,agesex)  _15+ yrs, male",default=0)
    tbprevc12n = models.IntegerField("tb_prev (n,agesex)  _male, 15-17 years",default=0)
    tbprevc13n = models.IntegerField("tb_prev (n,agesex)  _male, 18-Â­24 years",default=0)
    tbprevc14n = models.IntegerField("tb_prev (n,agesex)  _male, 25+ yrs",default=0)
    tbprevc15n = models.IntegerField("tb_prev (n,agesex)  _ipt by newly enrolled on art, female",default=0)
    tbprevc16n = models.IntegerField("tb_prev (n,agesex)  _ipt by newly enrolled on art, male",default=0)
    tbprevc17n = models.IntegerField("tb_prev (n,agesex)  _ipt by newly enrolled on art, total",default=0)
    tbprevc18n = models.IntegerField("tb_prev (n,agesex)  _ipt by previously enrolled on   art, female",default=0)
    tbprevc19n = models.IntegerField("tb_prev (n,agesex)  _ipt by previously enrolled on   art, male",default=0)
    tbprevc20n = models.IntegerField("tb_prev (n,agesex)  _ipt by previously enrolled on   art, total",default=0)
    tbprevc21n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, female",default=0)
    tbprevc22n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, male",default=0)
    tbprevc23n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by newly enrolled on art, total",default=0)
    tbprevc24n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, female",default=0)
    tbprevc25n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, male",default=0)
    tbprevc26n = models.IntegerField("tb_prev (n,type of tb preventative (tpt)_alternative tpt regimen (eg, 3 month inh and rifapentine) by  previously enrolled on art, total",default=0)

    def __str__(self):
        return self.period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility

class TbPrevTargets(models.Model):

    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)

    tbprev_tc1n = models.IntegerField("TB_PREV (N) Numerator",default=0)
    tbprev_tc2n = models.IntegerField("TB_PREV (N) 6-12 months IPT",default=0)
    tbprev_tc3n = models.IntegerField("TB_PREV (N) Female, <15",default=0)
    tbprev_tc4n = models.IntegerField("TB_PREV (N) Female, 15+",default=0)
    tbprev_tc5n = models.IntegerField("TB_PREV (N) Male, <15",default=0)
    tbprev_tc6n = models.IntegerField("TB_PREV (N) Male, 15+",default=0)

    tbprev_tc1d = models.IntegerField("TB_PREV (D) Numerator",default=0)
    tbprev_tc2d = models.IntegerField("TB_PREV (D) 6-12 months IPT",default=0)
    tbprev_tc3d = models.IntegerField("TB_PREV (D) Female, <15",default=0)
    tbprev_tc4d = models.IntegerField("TB_PREV (D) Female, 15+",default=0)
    tbprev_tc5d = models.IntegerField("TB_PREV (D) Male, <15",default=0)
    tbprev_tc6d = models.IntegerField("TB_PREV (D) Male, 15+",default=0)

    def __str__(self):
        return self.period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility

class TbPrev_Scorecard(object):
    period = ""
    district = ""
    subcounty = ""
    healthfacility = ""

    #data should be computed on facility
    target15LessN=0
    target15PlusN=0
    targetFemaleN=0
    targetMaleN=0
    target6To12N=0

    ArtIpt15LessN=0
    ArtIpt15PlusN=0
    ArtIptFemaleN=0
    ArtIptMaleN=0
    ArtIpt6To12N=0
    ArtIptContiousN=0
    ArtIptAlternativeRegimenN=0

    Perf15LessN=0
    Perf15PlusN=0
    PerfFemaleN=0
    PerfMaleN=0
    Perf6To12N=0
    
    #denominator
    target15LessD=0
    target15PlusD=0
    targetFemaleD=0
    targetMaleD=0
    target6To12D=0

    ArtIpt15LessD=0
    ArtIpt15PlusD=0
    ArtIptFemaleD=0
    ArtIptMaleD=0
    ArtIpt6To12D=0
    ArtIptContiousD=0
    ArtIptAlternativeRegimenD=0

    Perf15LessD=0
    Perf15PlusD=0
    PerfFemaleD=0
    PerfMaleD=0
    Perf6To12D=0

    OverAll15Less=0
    OverAll15Plus=0
    OverAllFemale=0
    OverAllMale=0
    OverAll6To12=0
    OverAllContious=0
    OverAllternativeRegimen=0

class Lab(models.Model):
    period = models.DateField("YYYY-MM",max_length=200)
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)

    dataelement = models.CharField("Data Element",max_length=200)
    dataelement_value = models.IntegerField("data element value",default=0)
   
    def __str__(self):
        return self.period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility

class LabTargets(models.Model):
    period = models.CharField("YYYY-MM",max_length=200)
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)

    dataelement = models.CharField("Data Element",max_length=200)
    dataelement_target = models.IntegerField("data element value",default=0)

    def __str__(self):
        return self.period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility

class Lab_Scorecard(object):
    period = ""
    district = ""
    subcounty = ""
    healthfacility = ""

    annual_target=0
    vl_samples_sent=-1
    perct_achievement_monthly=0
    vl_samples_rejected=0
    vl_samples_returned=0
    perct_achievement_sent_annnal=0

    def __init__(self):
        self.vl_samples_returned =(vl_samples_sent - vl_samples_rejected)
        self.perct_achievement_sent = (vl_samples_returned/vl_samples_sent) * 100

class RMNCHAndMalaria(models.Model):
    reporting_period = models.DateField("YYYY-MM",max_length=200)
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)
    healthfacilitylevel = models.CharField("Health Facility",max_length=200)
    ownership = models.CharField("Health Facility",max_length=200)

    ipt2_n = models.IntegerField("IPT2_N",default=0)
    ipt2_d = models.IntegerField("IPT2_D",default=0)
    amtsl_n = models.IntegerField("AMTSL_N",default=0)
    amtsl_d = models.IntegerField("AMTSL_D",default=0)
    asphyxia_n = models.IntegerField("Asphyxia_N",default=0)
    asphyxia_d = models.IntegerField("Asphyxia_D",default=0)
    diarrhea_n = models.IntegerField("Diarrhea_N",default=0)
    diarrhea_d = models.IntegerField("Diarrhea_D",default=0)
    pneumonia_n = models.IntegerField("Pneumonia_N",default=0)
    pneumonia_d = models.IntegerField("Pneumonia_D",default=0)
    ret_n = models.IntegerField("Ret_N",default=0)
    ret_d = models.IntegerField("Ret_D",default=0)
    hei_n = models.IntegerField("HEI_N",default=0)
    hei_d = models.IntegerField("HEI_D",default=0)
    simple_malaria_n = models.IntegerField("simple_malaria_N",default=0)
    simple_malaria_d = models.IntegerField("simple_malaria_D",default=0)
    severe_malaria_n = models.IntegerField("severe_malaria_N",default=0)
    severe_malaria_d = models.IntegerField("severe_malaria_D",default=0)
    dmcondoms = models.IntegerField("DOS:Mcondoms",default=0)
    dfcondoms = models.IntegerField("DOS:Fcondoms",default=0)
    dmbeads = models.IntegerField("DOS:Mbeads",default=0)
    dimplanon = models.IntegerField("DOS:Implanon",default=0)
    djadelle = models.IntegerField("DOS:Jadelle",default=0)
    diud = models.IntegerField("DOS:IUD",default=0)
    ddepoprovera = models.IntegerField("DOS:DepoProvera",default=0)
    dsayana = models.IntegerField("DOS:Sayana",default=0)
    dpills = models.IntegerField("DOS:Pills",default=0)
    depills = models.IntegerField("DOS:Epills",default=0)

    #for computation only
    ipt2=0
    man3stage=0
    manbirth=0
    treatd=0
    treatp=0
    pmtct=0
    mtct=0
    simple=0
    severe=0

    #doos properties
    
    def __str__(self):
        return self.reporting_period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility

class mnchandmal(object):
    period=''
    ipt2=0
    man3stage=0
    manbirth=0
    treatd=0
    treatp=0
    pmtct=0
    mtct=0
    simple=0
    severe=0

class DOOS(object):

    district =''
    subcounty =''
    healthfacility = ''
    healthfacilitylevel = ''
    ownership = ''
    period=''
    dmcondoms=0
    dfcondoms=0
    dmbeads=0
    dimplanon=0
    djadelle=0
    diud=0
    ddepoprovera=0
    dsayana=0
    dpills=0
    depills=0

class mnchandmal_dashboard(object):
    period=''
    ipt2=0
    man3stage=0
    manbirth=0
    treatd=0
    treatp=0
    pmtct=0
    mtct=0
    simple=0
    severe=0

class pmtcteid(models.Model):
    period = models.DateField("YYYY-MM",max_length=200)
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)
   
    ca1 = models.IntegerField("105-2.1 A17:HIV+ Pregnant Women already on ART before 1st ANC (ART-K)",default=0)
    ca2 = models.IntegerField("105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)",default=0)
    ca3 = models.IntegerField("105-2.1 A1:ANC 1st Visit for women 10-19 Years",default=0)
    ca4 = models.IntegerField("105-2.1 A1:ANC 1st Visit for women 20-24 Years",default=0)
    ca5 = models.IntegerField("105-2.1 A1:ANC 1st Visit for women >=25 Years",default=0)
    ca6 = models.IntegerField("105-2.1 HIV+ Pregnant Women initiated on ART for EMTCT (ART)",default=0)
    ca7 = models.IntegerField("105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR) 10-19 Years",default=0)
    ca8 = models.IntegerField("105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR) 20-24 Years",default=0)
    ca9 = models.IntegerField("105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR) >=25 Years",default=0)
    ca10 = models.IntegerField("105-2.1 Pregnant Women tested HIV+ for 1st time this pregnancy (TRR) at any visit 10-19 Years",default=0)
    ca11 = models.IntegerField("105-2.1 Pregnant Women tested HIV+ for 1st time this pregnancy (TRR) at any visit 20-24 Years",default=0)
    ca12 = models.IntegerField("105-2.1 Pregnant Women tested HIV+ for 1st time this pregnancy (TRR) at any visit >=25 Years",default=0)
    ca13 = models.IntegerField("105-2.1a Pregnant Women who knew status before 1st ANC (Total (TRK + TRRK))",default=0)
    ca14 = models.IntegerField("105-2.1b Pregnant Women who knew status before 1st ANC (HIV+(TRRK))",default=0)
    ca15 = models.IntegerField("105-2.2a Women testing HIV+ in labour (1st time this Pregnancy)",default=0)
    ca16 = models.IntegerField("105-2.2b Women testing HIV+ in labour (Retest this Pregnancy)",default=0)
    ca17 = models.IntegerField("105-2.3a Breastfeeding mothers newly testing HIV+(1st test)",default=0)
    ca18 = models.IntegerField("105-2.3b Breastfeeding mothers newly testing HIV+(retest)",default=0)
    ca19 = models.IntegerField("105-2.4a Exposed Infants Tested for HIV Below 18 Months(by 1st PCR)",default=0)
    ca20 = models.IntegerField("105-2.4b 1st DNA PCR result returned(HIV+)",default=0)
    ca21 = models.IntegerField("105-2.4c Exposed Infants Tested for HIV Below 18 Months(< 2 Months old)",default=0)
    ca22 = models.IntegerField("012 1.Total number of HEI in birth cohort (born 24 months previously)",default=0)
    ca23 = models.IntegerField("012 7.A. Outcomes for HIV exposed infants: Number of HEI being discharged at 18 months as Positive",default=0)
    ca24 = models.IntegerField("012 7. Outcomes for HIV exposed infants: Total Number of HEI being discharged at 18 months",default=0)
    ca25 = models.IntegerField("012 7.A. Outcomes for HIV exposed infants: Number of HEI being discharged at 18 months as Negative",default=0)
    ca26 = models.IntegerField("012 7.D. Outcomes for HIV exposed infants: Transferred out (Number of HEI who were transferred out before 18 months)",default=0)
    ca27 = models.IntegerField("012 7.E. Outcomes for HIV exposed infants:  Lost to follow (Number of HEI who are lost to follow-up before 18 months)",default=0)
    ca28 = models.IntegerField("012 7.F. Outcomes for HIV exposed infants: Died (Number of HEI who died before 18 months)",default=0)
    ca29 = models.IntegerField("012 7.G. Outcomes for HIV exposed infants: In care but no test done at 18 months",default=0)

    def __str__(self):
        return self.period +" "+ self.district +" "+ self.subcounty +" "+ self.healthfacility


class pmtcteid_targets(models.Model):
    district = models.CharField("District",max_length=200)
    subcounty = models.CharField("Subcounty",max_length=200)
    healthfacility = models.CharField("Health Facility",max_length=200)
   
    mtct_pmtct_art = models.IntegerField("105-2.1 A17:HIV+ Pregnant Women already on ART before 1st ANC (ART-K)",default=0)
    pmp = models.IntegerField("105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)",default=0)
    hivplus_infants = models.IntegerField("105-2.1 A1:ANC 1st Visit for women 10-19 Years",default=0)
    anci = models.IntegerField("105-2.1 A1:ANC 1st Visit for women 20-24 Years",default=0)
    pmtctstartpos = models.IntegerField("105-2.1 A1:ANC 1st Visit for women >=25 Years",default=0)
    pmtctstart = models.IntegerField("105-2.1 HIV+ Pregnant Women initiated on ART for EMTCT (ART)",default=0)

class pmtcteid_dashboard(object):
    period = ''
    district = ''
    healthfacility = ''
    T_PMTCT_STAT_nANC_D=0
    ANC_1_visit_D=0
    perf_T_PMTCT_STAT_nANC_DandANC_1_visit=0

    T_PMTCT_STAT_N=0
    TRandTRR=0
    TRKandTRRK=0
    total_TRandTRRandTRKandTRRK=0 #TRandTRR+TRKandTRRK
    PMTCT_STAT_1Nand1D=0 #(total_TRandTRRandTRKandTRRK/ANC_1_visit_D)*100
    PerfPMTCT_STAT=0 #(total_TRandTRRandTRKandTRRK/T_PMTCT_STAT_N)*100

    T_PMTCT_STAT_POS_N=0
    HIVplusTRRK=0
    TRRplus=0
    TRR=0
    t_HIVplusTRRK_TRRplus_TRR_ANC=0 #HIVplusTRRK+TRRplus+TRR
    PMTCT_STAT_POS_2N_1D=0 #(t_HIVplusTRRK_TRRplus_TRR_ANC/ANC_1_visit_D)*100
    PerfPMTCT_STAT_POS=0 #(t_HIVplusTRRK_TRRplus_TRR_ANC/T_PMTCT_STAT_POS_N)*100

    T_PMTCT_ART_N=0
    ART_K=0
    ART=0
    MTCT_3N=0 #ART_K+ART
    PMTCT_ART=0 #(MTCT_3N/t_HIVplusTRRK_TRRplus_TRR_ANC)*100
    PerfPMTCT_ART=0 #(MTCT_3N/T_PMTCT_ART_N)* 100

    HEI_discharged_18m=0
    HEI_transferred_out_before_18m=0
    HEI_ltfu_before_18m=0
    HEI_died_before_18m=0
    HEI_care_but_no_test_done18m=0
    t_HEI_4N=0 #HEI_discharged_18m+HEI_transferred_out_before_18m+HEI_ltfu_before_18m+HEI_died_before_18m+HEI_care_but_no_test_done18m
    HEI_in_birth_cohort_24mp_4D=0
    PMTCT_FO=0 #(t_HEI_4N/HEI_in_birth_cohort_24mp_4D)*100

    T_PMTCT_EID_N=0
    PMTCT_EIDles2m=0
    PMTCT_EIDles2_12m=0
    t_PMTCT_EID2m2_12m=0 #PMTCT_EIDles2m+PMTCT_EIDles2_12m
    t_HIVplusTRRK_TRRplus_TRR_ANC_c=0 #similar to t_HIVplusTRRK_TRRplus_TRR_ANC
    Women_HIV_plus_LD=0
    Breastfeeding_HIV_plus=0
    t_Women_HIV_plus_LDplusBreastfeeding_HIV_plus=0 #Women_HIV_plus_LD+Breastfeeding_HIV_plus
    PMTCT_EID=0 #(t_PMTCT_EID2m2_12m/t_Women_HIV_plus_LDplusBreastfeeding_HIV_plus)*100
    PerfPMTCT_EID=0 #(t_PMTCT_EID2m2_12m/T_PMTCT_EID_N)*100

    T_PMTCT_EID_POS=0
    PMTCT_EID_POS=0
    PerfPMTCT_EID_POS=0 #(PMTCT_EID_POS/T_PMTCT_EID_POS)*100

class IFASBottleneck(models.Model):
    bottleneck=models.CharField("Enter the name of the bottleneck",max_length=200)
    level=models.CharField("Level",max_length=200)
    when_identified = models.DateField("Please enter the date when the bottleneck was identified")
    where_identified = models.IntegerField(choices=Where_Identified_CHOICES, default=1) 
    potential_solutions=models.CharField("Level",max_length=400)
    efforts_to_address_bottleneck=models.CharField("Efforts to address the bottleneck",max_length=500)
    next_steps=models.CharField("Next steps",max_length=500)
    additional_bottleneck_identified=models.CharField("Additional bottleneck identified during efforts",max_length=500)
    comments=models.CharField("Comments if any",max_length=500)

    


    



    
