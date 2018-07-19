from django.shortcuts import render, get_object_or_404, render_to_response, redirect
from django.db.models import Avg, Case, Count, F, Max, Min, Prefetch, Q, Sum, When
from django.db.models import Value, CharField
from django.db.models.functions import Substr
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.template import RequestContext
from django.core.urlresolvers import reverse

from datetime import date
from decimal import Decimal
from itertools import groupby, tee, chain, product
from collections import OrderedDict

import openpyxl

from . import dateutil, grabbag
from .grabbag import default_zero, default, sum_zero, all_not_none, grouper

from .models import DataElement, OrgUnit, DataValue, ValidationRule, SourceDocument, ou_dict_from_path, ou_path_from_dict,lqas_dataset,lqas_target
from .forms import SourceDocumentForm, DataElementAliasForm, UserProfileForm

from .dashboards import LegendSet

import os
from django.conf import settings

def index(request):
    context = {
        'validation_rules': ValidationRule.objects.all().values_list('id', 'name')
    }
    return render(request, 'cannula/index.html', context)

@login_required
def validation_rule_listing(request, thematic_area):
    from django.db.models.functions import Concat
    import functools
    import operator
    from cannula.models import get_validation_view_names

    RULE_PREFIX_MAP = {
        'hiv': ('HCT_', 'HTS_'),
        'pmtct': ('PMT_',),
        'malaria': ('MAL_',),
        'mnch': ('MNCH_',),
        'nutrition': ('NUT_',),
        'tb': ('TB_',),
        'lab': ('LAB_', 'VL_'),
        'vmmc': ('VMC_',),
        'fp': ('FP_',),
        'gbv': ('GBV_',),
        'sc': ('SCM_',),
        'qi': ('QI_',),
    }

    rule_prefixes = RULE_PREFIX_MAP.get(thematic_area)
    if rule_prefixes:
        rule_filters = [Q(name__istartswith=r_prefix) for r_prefix in rule_prefixes]
        rule_filters_combined = functools.reduce(operator.__or__, rule_filters)
        qs_vr = ValidationRule.objects.filter(rule_filters_combined)
    else:
        qs_vr = ValidationRule.objects.all()
    qs_vr = qs_vr.order_by('name').annotate(expression=Concat('left_expr', Value(' '), 'operator', Value(' '), 'right_expr'))

    context = {
        'validation_rules': qs_vr,
        'validation_views': get_validation_view_names(),
    }
    return render(request, 'cannula/validation_rule_listing.html', context)

@login_required
def user_profile_edit(request):
    from django.contrib.auth import update_session_auth_hash
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib import messages

    if request.POST and 'profile_save' in request.POST:
        profile_form = UserProfileForm(request.POST, instance=request.user)
        if profile_form.is_valid():
            profile_form.save()
            messages.success(request, 'Profile updated')
    else:
        # create form from current user profile data
        profile_form = UserProfileForm(instance=request.user)

    if request.POST and 'passwd_change' in request.POST:
        passwd_form = PasswordChangeForm(request.user, request.POST)
        if passwd_form.is_valid():
            passwd_form.save()
            update_session_auth_hash(request, request.user)  # stop user from having to login again
            messages.success(request, 'Your password has been changed')
    else:
        # create form from current user profile data
        passwd_form = PasswordChangeForm(request.user)

    context = {
        'profile_form': profile_form,
        'passwd_form': passwd_form,
    }
    return render(request, 'cannula/user_profile_edit.html', context)

@login_required
def data_elements(request):
    data_elements = DataElement.objects.order_by('name').all()
    return render(request, 'cannula/data_element_listing.html', {'data_elements': data_elements})

# avoid strange behaviour from itertools.groupby by evaluating all the group iterables as lists
def groupbylist(*args, **kwargs):
    return [[k, list(g)] for k, g in groupby(*args, **kwargs)]

def filter_empty_rows(grouped_vals):
    for row in grouped_vals:
        row_heading, row_values = row
        if any(v['numeric_sum'] is not None for v in row_values):
            yield row

def month2quarter(month_num):
    return ((month_num-1)//3+1)

def make_excel_url(request_path):
    import os
    import urllib

    parts = urllib.parse.urlparse(request_path)
    a, b, path, *others = parts
    excel_path = ''.join([os.path.splitext(path)[0], '.xls'])
    return urllib.parse.urlunparse([a, b, excel_path, *others])

def make_csv_url(request_path):
    return make_excel_url(request_path).replace('.xls', '.csv')

@login_required
def malaria_cases_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    malaria_de_names = (
        '105-1.1 OPD New Attendance',
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
        '105-2.1 A3:Total ANC visits (New clients + Re-attendances)',
        '105-2.1 A9:Pregnant Women receiving free LLINs',
        'Malaria tests - WEP Microscopy Positive Cases',
        'Malaria tests - WEP Microscopy Tested Cases',
        'Malaria tests - WEP RDT Positve Cases',
        'Malaria tests - WEP RDT Tested Cases',
        'Malaria tests - WEP Suspected Malaria (fever)',
        'Malaria treated - WEP Microscopy Negative Cases Treated',
        'Malaria treated - WEP RDT Negative Cases Treated',
    )
    malaria_short_names = (
        '2D - OPD attendance',
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
        '1D - ANC visits',
        '1N - Pregnant women receiving free LLINs',
        'Malaria tests - WEP Microscopy Positive Cases',
        'Malaria tests - WEP Microscopy Tested Cases',
        'Malaria tests - WEP RDT Positve Cases',
        'Malaria tests - WEP RDT Tested Cases',
        '4D - Suspected malaria (fever)',
        'Malaria treated - WEP Microscopy Negative Cases Treated',
        'Malaria treated - WEP RDT Negative Cases Treated',
    )
    de_malaria_meta = list(product(malaria_de_names, (None,)))
    data_element_metas += list(product(malaria_short_names, (None,)))

    qs_malaria = DataValue.objects.what(*malaria_de_names)
    qs_malaria = qs_malaria.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_malaria = qs_malaria.where(filter_district)
    qs_malaria = qs_malaria.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_malaria = qs_malaria.when(filter_period)
    qs_malaria = qs_malaria.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_malaria = qs_malaria.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_malaria = list(val_malaria)

    gen_raster = grabbag.rasterize(ou_list, de_malaria_meta, val_malaria, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_malaria2 = list(gen_raster)

    age_under5_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    age_under5_short_names = (
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
    )
    de_age_under5_meta = list(product(age_under5_de_names, ('<5 yrs',)))
    data_element_metas += list(product(age_under5_short_names, (None,)))

    qs_age_under5 = DataValue.objects.what(*age_under5_de_names)
    qs_age_under5 = qs_age_under5.filter(category_combo__categories__name='29 Days-4 Years')
    qs_age_under5 = qs_age_under5.annotate(cat_combo=Value('<5 yrs', output_field=CharField()))
    # qs_age_under5 = qs_age_under5.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_age_under5 = qs_age_under5.where(filter_district)
    qs_age_under5 = qs_age_under5.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_age_under5 = qs_age_under5.when(filter_period)
    qs_age_under5 = qs_age_under5.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_age_under5 = qs_age_under5.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_age_under5 = list(val_age_under5)

    gen_raster = grabbag.rasterize(ou_list, de_age_under5_meta, val_age_under5, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_age_under52 = list(gen_raster)

    age_5_to_59_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    age_5_to_59_short_names = (
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
    )
    de_age_5_to_59_meta = list(product(age_5_to_59_de_names, ('5-59 yrs',)))
    data_element_metas += list(product(age_5_to_59_short_names, (None,)))

    qs_age_5_to_59 = DataValue.objects.what(*age_5_to_59_de_names)
    qs_age_5_to_59 = qs_age_5_to_59.filter(category_combo__categories__name='5-59 Years')
    qs_age_5_to_59 = qs_age_5_to_59.annotate(cat_combo=Value('5-59 yrs', output_field=CharField()))
    # qs_age_5_to_59 = qs_age_5_to_59.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_age_5_to_59 = qs_age_5_to_59.where(filter_district)
    qs_age_5_to_59 = qs_age_5_to_59.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_age_5_to_59 = qs_age_5_to_59.when(filter_period)
    qs_age_5_to_59 = qs_age_5_to_59.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_age_5_to_59 = qs_age_5_to_59.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_age_5_to_59 = list(val_age_5_to_59)

    gen_raster = grabbag.rasterize(ou_list, de_age_5_to_59_meta, val_age_5_to_59, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_age_5_to_592 = list(gen_raster)

    age_60_plus_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    age_60_plus_short_names = (
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
    )
    de_age_60_plus_meta = list(product(age_60_plus_de_names, ('60+ yrs',)))
    data_element_metas += list(product(age_60_plus_short_names, (None,)))

    qs_age_60_plus = DataValue.objects.what(*age_60_plus_de_names)
    qs_age_60_plus = qs_age_60_plus.filter(category_combo__categories__name='60andAbove Years')
    qs_age_60_plus = qs_age_60_plus.annotate(cat_combo=Value('60+ yrs', output_field=CharField()))
    # qs_age_60_plus = qs_age_60_plus.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_age_60_plus = qs_age_60_plus.where(filter_district)
    qs_age_60_plus = qs_age_60_plus.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_age_60_plus = qs_age_60_plus.when(filter_period)
    qs_age_60_plus = qs_age_60_plus.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_age_60_plus = qs_age_60_plus.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_age_60_plus = list(val_age_60_plus)

    gen_raster = grabbag.rasterize(ou_list, de_age_60_plus_meta, val_age_60_plus, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_age_60_plus2 = list(gen_raster)

    female_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    female_short_names = (
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
    )
    de_female_meta = list(product(female_de_names, ('Female',)))
    data_element_metas += list(product(female_short_names, (None,)))

    qs_female = DataValue.objects.what(*female_de_names)
    qs_female = qs_female.filter(category_combo__categories__name='Female')
    qs_female = qs_female.annotate(cat_combo=Value('Female', output_field=CharField()))
    # qs_female = qs_female.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_female = qs_female.where(filter_district)
    qs_female = qs_female.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_female = qs_female.when(filter_period)
    qs_female = qs_female.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_female = qs_female.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_female = list(val_female)

    gen_raster = grabbag.rasterize(ou_list, de_female_meta, val_female, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_female2 = list(gen_raster)

    male_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    male_short_names = (
        '2N - OPD malaria cases',
        '3N - OPD malaria cases confirmed',
    )
    de_male_meta = list(product(male_de_names, ('Male',)))
    data_element_metas += list(product(male_short_names, (None,)))

    qs_male = DataValue.objects.what(*male_de_names)
    qs_male = qs_male.filter(category_combo__categories__name='Male')
    qs_male = qs_male.annotate(cat_combo=Value('Male', output_field=CharField()))
    # qs_male = qs_male.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_male = qs_male.where(filter_district)
    qs_male = qs_male.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_male = qs_male.when(filter_period)
    qs_male = qs_male.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_male = qs_male.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_male = list(val_male)

    gen_raster = grabbag.rasterize(ou_list, de_male_meta, val_male, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_male2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_malaria2, val_age_under52, val_age_5_to_592, val_age_60_plus2, val_female2, val_male2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (opd_attend, opd_malaria, opd_malaria_confirmed, anc_visits, preg_given_nets, micro_pos, micro_tested, rdt_pos, rdt_tested, suspected, micro_neg_treated, rdt_neg_treated, under_5, under_5_confirmed, from_5_to_59, from_5_to_59_confirmed, over_60, over_60_confirmed, female, female_confirmed, male, male_confirmed, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        calculated_vals.append(preg_given_nets)
        calculated_vals.append(anc_visits)

        if all_not_none(preg_given_nets['numeric_sum']) and anc_visits['numeric_sum']:
            pct_preg_given_nets = 100 * preg_given_nets['numeric_sum'] / anc_visits['numeric_sum']
        else:
            pct_preg_given_nets = None
        pct_preg_given_nets_val = {
            'de_name': '1 - % women who received ITNs at ANC clinics',
            'cat_combo': None,
            'numeric_sum': pct_preg_given_nets,
        }
        pct_preg_given_nets_val.update(_group_ou_dict)
        calculated_vals.append(pct_preg_given_nets_val)

        calculated_vals.append(opd_malaria)
        calculated_vals.append(opd_attend)

        if all_not_none(opd_malaria['numeric_sum']) and opd_attend['numeric_sum']:
            pct_opd_malaria = 100 * opd_malaria['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_opd_malaria = None
        pct_opd_malaria_val = {
            'de_name': '2 - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': None,
            'numeric_sum': pct_opd_malaria,
        }
        pct_opd_malaria_val.update(_group_ou_dict)
        calculated_vals.append(pct_opd_malaria_val)

        if all_not_none(under_5['numeric_sum']) and opd_attend['numeric_sum']:
            pct_under_5 = 100 * under_5['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_under_5 = None
        pct_under_5_val = {
            'de_name': '2a - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': '<5 yrs',
            'numeric_sum': pct_under_5,
        }
        pct_under_5_val.update(_group_ou_dict)
        calculated_vals.append(pct_under_5_val)

        if all_not_none(from_5_to_59['numeric_sum']) and opd_attend['numeric_sum']:
            pct_from_5_to_59 = 100 * from_5_to_59['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_from_5_to_59 = None
        pct_from_5_to_59_val = {
            'de_name': '2a - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': '5-59 yrs',
            'numeric_sum': pct_from_5_to_59,
        }
        pct_from_5_to_59_val.update(_group_ou_dict)
        calculated_vals.append(pct_from_5_to_59_val)

        if all_not_none(over_60['numeric_sum']) and opd_attend['numeric_sum']:
            pct_over_60 = 100 * over_60['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_over_60 = None
        pct_over_60_val = {
            'de_name': '2a - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': '60+ yrs',
            'numeric_sum': pct_over_60,
        }
        pct_over_60_val.update(_group_ou_dict)
        calculated_vals.append(pct_over_60_val)

        if all_not_none(female['numeric_sum']) and opd_attend['numeric_sum']:
            pct_female = 100 * female['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_female = None
        pct_female_val = {
            'de_name': '2b - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': 'Female',
            'numeric_sum': pct_female,
        }
        pct_female_val.update(_group_ou_dict)
        calculated_vals.append(pct_female_val)

        if all_not_none(male['numeric_sum']) and opd_attend['numeric_sum']:
            pct_male = 100 * male['numeric_sum'] / opd_attend['numeric_sum']
        else:
            pct_male = None
        pct_male_val = {
            'de_name': '2b - Proportion of OPD cases diagnosed with malaria (%)',
            'cat_combo': 'Male',
            'numeric_sum': pct_male,
        }
        pct_male_val.update(_group_ou_dict)
        calculated_vals.append(pct_male_val)
        
        calculated_vals.append(opd_malaria_confirmed)

        if all_not_none(opd_malaria_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_opd_malaria_confirmed = 100 * opd_malaria_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_opd_malaria_confirmed = None
        pct_opd_malaria_confirmed_val = {
            'de_name': '3 - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)',
            'cat_combo': None,
            'numeric_sum': pct_opd_malaria_confirmed,
        }
        pct_opd_malaria_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_opd_malaria_confirmed_val)

        if all_not_none(under_5_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_under_5_confirmed = 100 * under_5_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_under_5_confirmed = None
        pct_under_5_confirmed_val = {
            'de_name': '3a - Proportion of OPD malaria cases confirmed (%)',
            'cat_combo': '<5 yrs',
            'numeric_sum': pct_under_5_confirmed,
        }
        pct_under_5_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_under_5_confirmed_val)

        if all_not_none(from_5_to_59_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_from_5_to_59_confirmed = 100 * from_5_to_59_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_from_5_to_59_confirmed = None
        pct_from_5_to_59_confirmed_val = {
            'de_name': '3a - Proportion of OPD malaria cases confirmed (%)',
            'cat_combo': '5-59 yrs',
            'numeric_sum': pct_from_5_to_59_confirmed,
        }
        pct_from_5_to_59_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_from_5_to_59_confirmed_val)

        if all_not_none(over_60_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_over_60_confirmed = 100 * over_60_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_over_60_confirmed = None
        pct_over_60_confirmed_val = {
            'de_name': '3a - Proportion of OPD malaria cases confirmed (%)',
            'cat_combo': '60+ yrs',
            'numeric_sum': pct_over_60_confirmed,
        }
        pct_over_60_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_over_60_confirmed_val)

        if all_not_none(female_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_female_confirmed = 100 * female_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_female_confirmed = None
        pct_female_confirmed_val = {
            'de_name': '3b - Proportion of OPD malaria cases confirmed (%)',
            'cat_combo': 'Female',
            'numeric_sum': pct_female_confirmed,
        }
        pct_female_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_female_confirmed_val)

        if all_not_none(male_confirmed['numeric_sum']) and opd_malaria['numeric_sum']:
            pct_male_confirmed = 100 * male_confirmed['numeric_sum'] / opd_malaria['numeric_sum']
        else:
            pct_male_confirmed = None
        pct_male_confirmed_val = {
            'de_name': '3b - Proportion of OPD malaria cases confirmed (%)',
            'cat_combo': 'Male',
            'numeric_sum': pct_male_confirmed,
        }
        pct_male_confirmed_val.update(_group_ou_dict)
        calculated_vals.append(pct_male_confirmed_val)

        if default(micro_tested['numeric_sum'], rdt_tested['numeric_sum']):
            suspected_tested = sum_zero(micro_tested['numeric_sum'], rdt_tested['numeric_sum'])
        else:
            suspected_tested = None
        suspected_tested_val = {
            'de_name': '4N - Suspected tested',
            'cat_combo': None,
            'numeric_sum': suspected_tested,
        }
        suspected_tested_val.update(_group_ou_dict)
        calculated_vals.append(suspected_tested_val)
        
        calculated_vals.append(suspected)

        if all_not_none(suspected_tested) and suspected['numeric_sum']:
            pct_suspected_tested = 100 * suspected_tested / suspected['numeric_sum']
        else:
            pct_suspected_tested = None
        pct_suspected_tested_val = {
            'de_name': '4 - Proportion of Suspected tested (%)',
            'cat_combo': None,
            'numeric_sum': pct_suspected_tested,
        }
        pct_suspected_tested_val.update(_group_ou_dict)
        calculated_vals.append(pct_suspected_tested_val)

        if default(micro_neg_treated['numeric_sum'], rdt_neg_treated['numeric_sum']):
            neg_treated = sum_zero(micro_neg_treated['numeric_sum'], rdt_neg_treated['numeric_sum'])
        else:
            neg_treated = None
        neg_treated_val = {
            'de_name': '5N - Negative cases treated',
            'cat_combo': None,
            'numeric_sum': neg_treated,
        }
        neg_treated_val.update(_group_ou_dict)
        calculated_vals.append(neg_treated_val)

        if default(micro_tested['numeric_sum'], rdt_tested['numeric_sum']):
            negative = sum_zero(micro_tested['numeric_sum'], rdt_tested['numeric_sum']) - sum_zero(micro_pos['numeric_sum'], rdt_pos['numeric_sum'])
        else:
            negative = None
        negative_val = {
            'de_name': '5D - Negative cases',
            'cat_combo': None,
            'numeric_sum': negative,
        }
        negative_val.update(_group_ou_dict)
        calculated_vals.append(negative_val)

        if all_not_none(neg_treated) and negative:
            pct_neg_treated = 100 * neg_treated / negative
        else:
            pct_neg_treated = None
        pct_neg_treated_val = {
            'de_name': '5 - Proportion of negative cases treated (%)',
            'cat_combo': None,
            'numeric_sum': pct_neg_treated,
        }
        pct_neg_treated_val.update(_group_ou_dict)
        calculated_vals.append(pct_neg_treated_val)

        # _group[1].extend(calculated_vals)
        _group[1] = calculated_vals # replace the output indicators

    data_element_metas = list() # replace the list of output indicators
    data_element_metas += list(product(['1N - Pregnant women receiving free LLINs'], (None,)))
    data_element_metas += list(product(['1D - ANC visits'], (None,)))
    data_element_metas += list(product(['1 - % women who received ITNs at ANC clinics'], (None,)))
    data_element_metas += list(product(['2N - OPD malaria cases'], (None,)))
    data_element_metas += list(product(['2D - OPD attendance'], (None,)))
    data_element_metas += list(product(['2 - Proportion of OPD cases diagnosed with malaria (%)'], (None,)))
    data_element_metas += list(product(['2a - Proportion of OPD cases diagnosed with malaria (%)'], ('<5 yrs',)))
    data_element_metas += list(product(['2a - Proportion of OPD cases diagnosed with malaria (%)'], ('5-59 yrs',)))
    data_element_metas += list(product(['2a - Proportion of OPD cases diagnosed with malaria (%)'], ('60+ yrs',)))
    data_element_metas += list(product(['2b - Proportion of OPD cases diagnosed with malaria (%)'], ('Female',)))
    data_element_metas += list(product(['2b - Proportion of OPD cases diagnosed with malaria (%)'], ('Male',)))
    data_element_metas += list(product(['3N - OPD malaria cases confirmed'], (None,)))
    data_element_metas += list(product(['3 - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], (None,)))
    data_element_metas += list(product(['3a - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], ('<5 yrs',)))
    data_element_metas += list(product(['3a - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], ('5-59 yrs',)))
    data_element_metas += list(product(['3a - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], ('60+ yrs',)))
    data_element_metas += list(product(['3b - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], ('Female',)))
    data_element_metas += list(product(['3b - Proportion of OPD malaria cases confirmed by RDT or microscopy (%)'], ('Male',)))
    data_element_metas += list(product(['4N - Suspected tested'], (None,)))
    data_element_metas += list(product(['4D - Suspected malaria (fever)'], (None,)))
    data_element_metas += list(product(['4 - Proportion of Suspected tested (%)'], (None,)))
    data_element_metas += list(product(['5N - Negative cases treated'], (None,)))
    data_element_metas += list(product(['5D - Negative cases'], (None,)))
    data_element_metas += list(product(['5 - Proportion of negative cases treated (%)'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    tested_ls = LegendSet()
    tested_ls.name = 'Testing and Treatment'
    # tested_ls.add_interval('orange', 0, 25)
    # tested_ls.add_interval('yellow', 25, 40)
    # tested_ls.add_interval('light-green', 40, 60)
    # tested_ls.add_interval('green', 60, 100)
    tested_ls.add_interval('red', 100, None)
    tested_ls.mappings[num_path_elements+20] = True
    tested_ls.mappings[num_path_elements+23] = True
    legend_sets.append(tested_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="malaria_cases_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="malaria_cases_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/malaria_cases_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

def malaria_dashboard(request):
    this_day = date.today()
    this_quarter = '%d-Q%d' % (this_day.year, month2quarter(this_day.month))
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_day.year, this_day.year-6, -1) for q in range(4, 0, -1)]
    period_list = list(filter(lambda qtr: qtr < this_quarter, reversed(PREV_5YR_QTRS)))[-6:]
    def val_with_period_de_fun(row, col):
        period = row
        de_name = col
        return { 'de_name': de_name, 'period': period, 'numeric_sum': None }

    malaria_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
        '105-2.1 A7:Second dose IPT (IPT2)',
    )
    de_malaria_meta = list(product(malaria_de_names, (None,)))
    qs_malaria = DataValue.objects.what(*malaria_de_names)
    qs_malaria = qs_malaria.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_malaria = qs_malaria.when(*period_list)
    qs_malaria = qs_malaria.order_by('period', 'de_name')
    val_malaria = qs_malaria.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_malaria = list(val_malaria)
    val_malaria = list(grabbag.rasterize(period_list, malaria_de_names, val_malaria, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun))

    preg_de_names =(
        'Expected Pregnancies',
    )
    de_preg_meta = list(product(preg_de_names, (None,)))
    qs_preg = DataValue.objects.what(*preg_de_names)
    qs_preg = qs_preg.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_preg = qs_preg.when('2016', '2017')
    qs_preg = qs_preg.order_by('period', 'de_name')
    val_preg = qs_preg.values('period', 'de_name').annotate(numeric_sum=(Sum('numeric_value')))
    val_preg = list(val_preg)
    # convert annual expected pregnancies to quarterly
    for v in val_preg[::-1]: # in reverse, since we'll be adding to the end
        v['numeric_sum'] = v['numeric_sum']/4
        year = v['period']
        v['period'] = '{0}-Q{1}'.format(year, 1)
        for q in ['{0}-Q{1}'.format(year, i) for i in range(2, 5)]:
            v_quarter = dict(v)
            v_quarter['period'] = q
            v_quarter['numeric_sum'] = v_quarter['numeric_sum']
            val_preg.append(v_quarter)
    val_preg.sort(key=lambda x: (x['period'], x['de_name']))
    val_preg = list(grabbag.rasterize(period_list, preg_de_names, val_preg, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun))

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_preg, val_malaria), key=lambda x: (x['period'])), key=lambda x: (x['period']))
    # if True:
    #     grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (period, (expected_pregnancies, malaria_total, malaria_confirmed, ipt2, *other_vals)) = _group
        
        calculated_vals = list()

        if all_not_none(ipt2['numeric_sum'], expected_pregnancies['numeric_sum']) and expected_pregnancies['numeric_sum']:
            ipt2_rate = (ipt2['numeric_sum'] * 100) / expected_pregnancies['numeric_sum']
        else:
            ipt2_rate = None
        ipt2_rate_val = {
            'period': period,
            'de_name': 'IPT2 Rate (%)',
            'numeric_sum': ipt2_rate,
        }
        calculated_vals.append(ipt2_rate_val)

        if all_not_none(malaria_confirmed['numeric_sum'], malaria_total['numeric_sum']) and malaria_total['numeric_sum']:
            presumptive_rate = 100 - (malaria_confirmed['numeric_sum'] * 100) / malaria_total['numeric_sum']
        else:
            presumptive_rate = None
        presumptive_rate_val = {
            'period': period,
            'de_name': 'Presumptive Treatment Rate (%)',
            'numeric_sum': presumptive_rate,
        }
        calculated_vals.append(presumptive_rate_val)

        _group[1] = calculated_vals
    
    context = {
        'data_element_names': [
            ('IPT Rate (%)', None),
            ('Presumptive Treatment Rate (%)', None),
        ],
        'grouped_data': grouped_vals,
        'calculated_vals': calculated_vals,
    }
    return render(request, 'cannula/index.html', context)

@login_required
def malaria_ipt_scorecard(request, org_unit_level=2, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }
    # annotations for data collected at subcounty level
    SUBCOUNTY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(2, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    ipt_de_names = (
        '105-2.1 A6:First dose IPT (IPT1)',
        '105-2.1 A7:Second dose IPT (IPT2)',
    )
    de_ipt_meta = list(product(ipt_de_names, (None,)))

    # get IPT1 and IPT2 without subcategory disaggregation
    qs = DataValue.objects.what(*ipt_de_names)
    qs = qs.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs = qs.where(filter_district)
    qs = qs.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs = qs.when(filter_period)
    qs = qs.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_ipt_all = qs.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    
    gen_raster = grabbag.rasterize(ou_list, de_ipt_meta, val_ipt_all, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_ipt_all2 = list(gen_raster)

    # get list of subcategories for IPT2
    qs_ipt_subcat = DataValue.objects.what('105-2.1 A7:Second dose IPT (IPT2)').order_by('category_combo__name').values_list('de_name', 'category_combo__name').distinct()
    subcategory_names = tuple(qs_ipt_subcat)

    # get IPT2 with subcategory disaggregation
    qs2 = DataValue.objects.what('105-2.1 A7:Second dose IPT (IPT2)')
    qs2 = qs2.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs2 = qs2.where(filter_district)
    qs2 = qs2.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs2 = qs2.when(filter_period)
    qs2 = qs2.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_dicts2 = qs2.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, subcategory_names, val_dicts2, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_dicts2 = list(gen_raster)

    pregnancies_de_names = (
        'Expected Pregnancies',
    )
    de_pregnancies_meta = list(product(pregnancies_de_names, (None,)))
    # get expected pregnancies
    qs3 = DataValue.objects.what(*pregnancies_de_names)
    qs3 = qs3.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs3 = qs3.where(filter_district)
    # use clearer aliases for the unwieldy names
    qs3 = qs3.annotate(**SUBCOUNTY_LEVEL_ANNOTATIONS)
    # pregnancy estimates are annual (from population), so filter by year component of period and divide by 4
    qs3 = qs3.when(filter_period[:4])
    qs3 = qs3.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_preg = qs3.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(numeric_sum=(Sum('numeric_value')/4))

    gen_raster = grabbag.rasterize(ou_list, de_pregnancies_meta, val_preg, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_preg2 = list(gen_raster)

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_preg2, val_ipt_all2, val_dicts2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))
    
    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (preg_val, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))

        if preg_val['de_name'] == 'Expected Pregnancies':
            for i, val in enumerate(reversed(other_vals)):
                # calculate the IPT rate for the IPT1/IPT2 values (without subcategories)
                if val['de_name'] in ipt_de_names and val['cat_combo'] is None:
                    if all_not_none(val['numeric_sum'], preg_val['numeric_sum']) and preg_val['numeric_sum']:
                        ipt_percent = 100 * val['numeric_sum'] / preg_val['numeric_sum']
                    else:
                        ipt_percent = None
                    ipt_percent_val = {
                        'de_name': val['de_name'] + ' %',
                        'cat_combo': val['cat_combo'],
                        'numeric_sum': ipt_percent,
                    }
                    ipt_percent_val.update(_group_ou_dict)
                    _group[1].insert(len(other_vals)-i+1, ipt_percent_val)

    data_element_names = list()
    data_element_names.insert(0, ('Expected Pregnancies', None))
    for de_n in ipt_de_names:
        data_element_names.append((de_n, None))
        data_element_names.append(('%', None))
    data_element_names.extend(subcategory_names)

    num_path_elements = len(ou_headers)
    legend_sets = list()
    ipt_ls = LegendSet()
    ipt_ls.name = 'IPT rate'
    ipt_ls.add_interval('yellow', 0, 71)
    ipt_ls.add_interval('green', 71, None)
    ipt_ls.mappings[num_path_elements+2] = True
    ipt_ls.mappings[num_path_elements+4] = True
    legend_sets.append(ipt_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_names))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="malaria_ipt_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_names)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            offset = 0
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j+offset, value=g_val['numeric_sum'])
                if 'ipt_rate' in g_val:
                    offset += 1
                    ws.cell(row=i, column=j+offset, value=g_val['ipt_rate'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSet
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="malaria_ipt_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_names,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    if output_format == 'JSON':
        from django.http import JsonResponse
        
        return JsonResponse(context)

    return render(request, 'cannula/malaria_ipt_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def malaria_compliance(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'start_period' in request.GET and request.GET['start_period'] in PREV_5YR_QTRS and 'end_period' in request.GET and request.GET['end_period']:
        start_quarter = request.GET['start_period']
        end_quarter = request.GET['end_period']
    else: # default to "immediate preceding quarter" and "this quarter"
        if this_day.month <= 3:
            start_year = this_year - 1
            start_month = (this_day.month - 3 + 12)
            end_month = this_day.month
        else:
            start_year = this_year
            start_month = this_day.month - 3
            end_month = this_day.month
        start_quarter = '%d-Q%d' % (start_year, month2quarter(start_month))
        end_quarter = '%d-Q%d' % (this_year, month2quarter(end_month))

    periods = dateutil.get_quarters(start_quarter, end_quarter)
    if start_quarter == end_quarter:
        periods = periods[:1]

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_period_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, period = col
        val_dict.update({ 'period': period, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    cases_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )
    de_cases_meta = tuple(product(cases_de_names, periods))
    data_element_metas += de_cases_meta

    qs = DataValue.objects.what(*cases_de_names)
    qs = qs.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs = qs.where(filter_district)
    qs = qs.when(*periods)
    qs = qs.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs = qs.order_by(*OU_PATH_FIELDS, 'de_name', 'period')
    val_dicts = qs.values(*OU_PATH_FIELDS, 'de_name', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_cases_meta, val_dicts, ou_path_from_dict, lambda x: (x['de_name'], x['period']), orgunit_vs_de_period_default)
    val_dicts2 = gen_raster

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(val_dicts2, key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    for _group in grouped_vals:
        (_group_ou_path, other_vals) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        calculated_vals = list()

        totals, confirmeds = other_vals[:len(periods)], other_vals[len(periods):]
        for i, (total_val, confirmed_val) in enumerate(reversed(list(zip(totals, confirmeds)))):
            if (total_val['de_name'], confirmed_val['de_name']) == cases_de_names:
                if all_not_none(confirmed_val['numeric_sum'], total_val['numeric_sum']) and total_val['numeric_sum']:
                    confirmed_rate = 100 * confirmed_val['numeric_sum'] / total_val['numeric_sum']
                else:
                    confirmed_rate = None
                confirmed_rate_val = {
                    'de_name': confirmed_val['de_name'] + ' %',
                    'period': confirmed_val['period'],
                    'numeric_sum': confirmed_rate,
                }
                confirmed_rate_val.update(_group_ou_dict)
                _group[1].insert(len(other_vals)-i*2, confirmed_rate_val)
                calculated_vals.append(confirmed_rate_val)

        prev_vals = None
        for total_val, confirmed_val in zip(totals, confirmeds):
            if prev_vals:
                total_val['previous'], confirmed_val['previous'] = prev_vals
            
            prev_vals = total_val['numeric_sum'], confirmed_val['numeric_sum']
        prev_val = None
        for calc_val in reversed(calculated_vals):
            if prev_val:
                calc_val['previous'] = prev_val
            prev_val = calc_val['numeric_sum']

    data_element_names = list()
    for de_n in cases_de_names:
        data_element_names.append((de_n, None))

    for x in range(len(data_element_metas)//2):
        data_element_metas.insert(len(data_element_metas)-x*2, (cases_de_names[1], '%'))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    compliance_ls = LegendSet()
    compliance_ls.name = 'Compliance'
    compliance_ls.add_interval('green', 80, None)
    compliance_ls.add_interval('yellow', 50, 80)
    compliance_ls.add_interval('red', 0, 50)
    for i in range(len(periods)):
        compliance_ls.mappings[num_path_elements+len(periods)+i*2+1] = True
    legend_sets.append(compliance_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="malaria_compliance_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            offset = 0
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j+offset, value=g_val['numeric_sum'])
                if 'ipt_rate' in g_val:
                    offset += 1
                    ws.cell(row=i, column=j+offset, value=g_val['ipt_rate'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSet
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="malaria_compliance_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'start_period': start_quarter,
        'end_period': end_quarter,
        'periods': periods,
        'period_desc': dateutil.DateSpan.fromquarter(start_quarter).combine(dateutil.DateSpan.fromquarter(end_quarter)).format_long(),
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/malaria_compliance_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def data_workflow_new(request, menu_name):
    if request.method == 'POST':
        form = SourceDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            docs = SourceDocument.objects.all().annotate(num_values=Count('data_values'))
            docs = docs.order_by('-uploaded_at')

            context = {
                'workflows': docs,
                'menu_name': menu_name,
                'step': 2,
            }
            return render_to_response('cannula/data_workflow_new.html', context, context_instance=RequestContext(request))
    else:
        form = SourceDocumentForm()

    context = {
        'form': form,
        'menu_name': menu_name,
        'step': 1,
    }

    return render_to_response('cannula/data_workflow_new.html', context, context_instance=RequestContext(request))

@login_required
def data_workflow_detail(request):
    from .models import load_excel_to_datavalues, load_excel_to_validations

    if 'wf_id' in request.GET:
        src_doc_id = int(request.GET['wf_id'])
        src_doc = get_object_or_404(SourceDocument, id=src_doc_id)

        if request.method == 'POST':
            if 'load_values' in request.POST:
                all_values = load_excel_to_datavalues(src_doc)
                for site_name, site_vals in all_values.items():
                    DataValue.objects.bulk_create(site_vals)
            elif 'load_validations' in request.POST:
                load_excel_to_validations(src_doc)

        qs_vals = DataValue.objects.filter(source_doc__id=src_doc_id).values('id')
        doc_elements = DataElement.objects.filter(data_values__id__in=qs_vals).order_by('name').distinct('name')
        doc_rules = ValidationRule.objects.filter(data_elements__data_values__id__in=qs_vals).order_by('name').distinct('name')
        num_values = qs_vals.count()
    else:
        raise Http404("Workflow does not exist or workflow id is missing/invalid")

    editable_ids = ""
    editable_names = ""
    ids = ""
    for de in doc_elements:
        ids += " %s" % de.id
        editable_ids += " editable_%s" % de.id
        editable_names += "|%s" % de.name

    context = {
        'srcdoc': src_doc,
        'num_values': num_values,
        'data_elements': doc_elements,
        'validation_rules': doc_rules,
        'step': 3,
        'ids': ids,
        'editable': editable_ids,
        'editable_names': editable_names,
    }
    return render_to_response('cannula/data_workflow_new.html', context, context_instance=RequestContext(request))

@login_required
def data_workflow_listing(request):
    # TODO: filter based on user who uploaded file?
    docs = SourceDocument.objects.all().annotate(num_values=Count('data_values'))
    docs = docs.order_by('-uploaded_at')

    context = {
        'workflows': docs,
    }
    return render(request, 'cannula/data_workflow_listing.html', context)

def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

@login_required
def validation_rule(request, output_format='HTML'):
    from django.db import connection
    cursor = connection.cursor()
    vr_id = int(request.GET['id'])
    vr = ValidationRule.objects.get(id=vr_id)
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))
        PREV_5YR_QTRS = list(filter(lambda x: x <= filter_period, PREV_5YR_QTRS))

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    if filter_district:
        sql_str = 'SELECT * FROM %s WHERE district=\'%s\'' % (vr.view_name(), filter_district.name)
    else:
        sql_str = 'SELECT * FROM %s' % (vr.view_name(),)

    if filter_period:
        if 'WHERE' in sql_str:
            cursor.execute(sql_str+' AND quarter=%s', (filter_period,))
        else:
            cursor.execute(sql_str+' WHERE quarter=%s', (filter_period,))
    else:
        cursor.execute(sql_str)

    columns = [col[0] for col in cursor.description]
    de_name_map = dict()
    for de_id, de_name in DataElement.objects.all().values_list('id', 'name'):
        de_name_map['de_%d' % (de_id,)] = de_name
        columns = [c.replace('de_%d' % (de_id,), de_name) for c in columns] #TODO: can we include the alias, if there is one?
    results = dictfetchall(cursor)
    for r in results:
        r['data_values'] = dict()
        for k,v in r.items():
            if k in de_name_map:
                de_name = de_name_map[k]
                r['data_values'][de_name] = v
    if 'exclude_true' in request.GET:
        results = filter(lambda x: not x['de_calc_1'], results)

    validates_ls = LegendSet()
    validates_ls.add_interval('red', None, 1)
    validates_ls.add_interval('green', 1, None)
    validates_ls.mappings[4] = True

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = vr.expression().strip()[:31] # worksheet names length limit is 31
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        def format_values(v_dict):
            return '\n'.join([ '%s: %s' % (k,v) for k,v in v_dict.items()])

        headers = ['Period', 'District', 'Subcounty', 'Facility', 'Validates?', 'Source Data']
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, res in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=next(filter(lambda x: x is not None, (res['month'], res['quarter'], res['year'])), None))
            ws.cell(row=i, column=2, value=res['district'])
            ws.cell(row=i, column=3, value=res['subcounty'])
            ws.cell(row=i, column=4, value=res['facility'])
            ws.cell(row=i, column=5, value=res['de_calc_1'])
            ws.cell(row=i, column=6, value=format_values(res['data_values']))

        for rule in validates_ls.openpyxl_rules():
            # apply conditional formatting from LegendSet
            for xls_range in validates_ls.excel_ranges():
                ws.conditional_formatting.add(xls_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s_validation.xlsx"' % (vr.name.lower(),)

        return response

    context = {
        'results': results,
        'columns': columns,
        'rule': vr,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path)
    }

    return render(request, 'cannula/validation_rule.html', context)

@login_required
def data_element_alias(request):
    if 'de_id' in request.GET:
        de_id = int(request.GET['de_id'])
        de = get_object_or_404(DataElement, id=de_id)

        if request.method == 'POST':
            form = DataElementAliasForm(request.POST, instance=de)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.name = request.POST['value[name]']
                obj.alias = request.POST['value[alias]']
                obj.save()
                de_url = '%s?wf_id=%d' % (reverse('data_workflow_detail'), int(request.GET['wf_id']))
                return redirect(de_url, de)
        else:
            form = DataElementAliasForm(instance=de)

        context = {
            'form': form,
        }
    else:
        raise Http404("Data Element does not exist or data element id is missing/invalid")

    return render_to_response('cannula/data_element_edit_alias.html', context, context_instance=RequestContext(request))

@login_required
def hts_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    hts_de_names = (
        '105-4 Number of Individuals who received HIV test results',
        '105-4 Number of Individuals who tested HIV positive',
        '105-4 Number of clients who have been linked to care',
    )
    hts_short_names = (
        'Linked',
        'Tested',
        'HIV+',
    )
    subcategory_names = ['(15+, Female)', '(15+, Male)', '(<15, Female)', '(<15, Male)']
    de_positivity_meta = list(product(hts_de_names, subcategory_names))

    qs_positivity = DataValue.objects.what(*hts_de_names)
    cc_lt_15 = ['18 Mths-<5 Years', '5-<10 Years', '10-<15 Years']
    cc_ge_15 = ['15-<19 Years', '19-<49 Years', '>49 Years']
    #TODO: cc_lt_15_f = CategoryCombo.from_cat_names(['Female', '<15']) gives a CategoryCombo instance that makes the Case statement clearer/safer
    qs_positivity = qs_positivity.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    qs_positivity = qs_positivity.exclude(cat_combo__iexact=None)
    if filter_district:
        qs_positivity = qs_positivity.where(filter_district)
    qs_positivity = qs_positivity.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_positivity = qs_positivity.when(filter_period)
    qs_positivity = qs_positivity.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_positivity = qs_positivity.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    
    gen_raster = grabbag.rasterize(ou_list, de_positivity_meta, val_positivity, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_positivity2 = list(gen_raster)

    pmtct_mother_de_names = (
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.2a Women tested for HIV in labour (1st time this Pregnancy)',
        '105-2.3a Breastfeeding mothers tested for HIV(1st test)',
    )
    de_pmtct_mother_meta = list(product(('Pregnant Women tested for HIV',), (None,)))

    qs_pmtct_mother = DataValue.objects.what(*pmtct_mother_de_names)
    qs_pmtct_mother = qs_pmtct_mother.annotate(de_name=Value('Pregnant Women tested for HIV', output_field=CharField()))
    qs_pmtct_mother = qs_pmtct_mother.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_pmtct_mother = qs_pmtct_mother.where(filter_district)
    qs_pmtct_mother = qs_pmtct_mother.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_pmtct_mother = qs_pmtct_mother.when(filter_period)
    qs_pmtct_mother = qs_pmtct_mother.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_pmtct_mother = qs_pmtct_mother.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_meta, val_pmtct_mother, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_pmtct_mother2 = list(gen_raster)

    pmtct_mother_pos_de_names = (
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.2a Women testing HIV+ in labour (1st time this Pregnancy)',
        '105-2.2b Women testing HIV+ in labour (Retest this Pregnancy)',
        '105-2.3a Breastfeeding mothers newly testing HIV+(1st test)',
        '105-2.3b Breastfeeding mothers newly testing HIV+(retest)',
    )
    de_pmtct_mother_pos_meta = list(product(('Pregnant Women testing HIV+',), (None,)))

    qs_pmtct_mother_pos = DataValue.objects.what(*pmtct_mother_pos_de_names)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(de_name=Value('Pregnant Women testing HIV+', output_field=CharField()))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_pmtct_mother_pos = qs_pmtct_mother_pos.where(filter_district)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.when(filter_period)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_pmtct_mother_pos = qs_pmtct_mother_pos.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_pos_meta, val_pmtct_mother_pos, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_pmtct_mother_pos2 = list(gen_raster)

    pmtct_child_de_names = (
        '105-2.4a Exposed Infants Tested for HIV Below 18 Months(by 1st PCR) ',
        '105-2.4b 1st DNA PCR result returned(HIV+)',
        '105-2.4b 2nd DNA PCR result returned(HIV+)',
        '105-2.1a Male partners received HIV test results in eMTCT(Total)',
        '105-2.1b Male partners received HIV test results in eMTCT(HIV+)',
    )
    pmtct_child_short_names = (
        'PMTCT INFANT HIV+',
        'PMTCT CHILD PCR1 HIV+',
        'PMTCT CHILD PCR2 HIV+',
        'PMTCT MALE PARTNERS TESTED',
        'PMTCT MALE PARTNERS HIV+',
    )
    de_pmtct_child_meta = list(product(pmtct_child_de_names, (None,)))

    qs_pmtct_child = DataValue.objects.what(*pmtct_child_de_names)
    qs_pmtct_child = qs_pmtct_child.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_pmtct_child = qs_pmtct_child.where(filter_district)
    qs_pmtct_child = qs_pmtct_child.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_pmtct_child = qs_pmtct_child.when(filter_period)
    qs_pmtct_child = qs_pmtct_child.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_pmtct_child = qs_pmtct_child.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_pmtct_child = list(val_pmtct_child)

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_child_meta, val_pmtct_child, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_pmtct_child2 = list(gen_raster)

    target_de_names = (
        'HTC_TST_POS_TARGET',
        'HTC_TST_TARGET',
    )
    de_target_meta = list(product(target_de_names, subcategory_names))

    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_target = DataValue.objects.what(*target_de_names)
    qs_target = qs_target.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_target = qs_target.where(filter_district)
    qs_target = qs_target.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_target = qs_target.when(filter_period[:4])
    qs_target = qs_target.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_target = qs_target.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)

    gen_raster = grabbag.rasterize(ou_list, de_target_meta, val_target, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_target2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_positivity2, val_pmtct_mother2, val_pmtct_mother_pos2, val_pmtct_child2, val_target2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (tst_over15_f, tst_over15_m, tst_under15_f, tst_under15_m, pos_over15_f, pos_over15_m, pos_under15_f, pos_under15_m, linked_over15_f, linked_over15_m, linked_under15_f, linked_under15_m, tst_pregnant, pos_pregnant, pos_infant, pos_pcr1, pos_pcr2, tst_male_partner, pos_male_partner, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        under15_f_sum = default_zero(tst_under15_f['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_f_val = {
            'de_name': 'Tested',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_sum,
        }
        under15_f_val.update(_group_ou_dict)
        calculated_vals.append(under15_f_val)
        
        under15_m_sum = default_zero(tst_under15_m['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_m_val = {
            'de_name': 'Tested',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_sum,
        }
        under15_m_val.update(_group_ou_dict)
        calculated_vals.append(under15_m_val)
        
        over15_f_sum = default_zero(tst_over15_f['numeric_sum']) + default_zero(tst_pregnant['numeric_sum'])
        over15_f_val = {
            'de_name': 'Tested',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_sum,
        }
        over15_f_val.update(_group_ou_dict)
        calculated_vals.append(over15_f_val)
        
        over15_m_sum = default_zero(tst_over15_m['numeric_sum']) + default_zero(tst_male_partner['numeric_sum'])
        over15_m_val = {
            'de_name': 'Tested',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_sum,
        }
        over15_m_val.update(_group_ou_dict)
        calculated_vals.append(over15_m_val)
        
        half_pos_pcr = Decimal(default_zero(pos_pcr1['numeric_sum']) + default_zero(pos_pcr1['numeric_sum']))/2
        pos_under15_f_sum = default_zero(pos_under15_f['numeric_sum']) + half_pos_pcr
        pos_under15_f_val = {
            'de_name': 'HIV+',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_sum,
        }
        pos_under15_f_val.update(_group_ou_dict)
        calculated_vals.append(pos_under15_f_val)
        
        pos_under15_m_sum = default_zero(pos_under15_m['numeric_sum']) + half_pos_pcr
        pos_under15_m_val = {
            'de_name': 'HIV+',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_sum,
        }
        pos_under15_m_val.update(_group_ou_dict)
        calculated_vals.append(pos_under15_m_val)
        
        pos_over15_f_sum = default_zero(pos_over15_f['numeric_sum']) + Decimal(default_zero(pos_pregnant['numeric_sum']))
        pos_over15_f_val = {
            'de_name': 'HIV+',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_sum,
        }
        pos_over15_f_val.update(_group_ou_dict)
        calculated_vals.append(pos_over15_f_val)
        
        pos_over15_m_sum = default_zero(pos_over15_m['numeric_sum']) + Decimal(default_zero(pos_male_partner['numeric_sum']))
        pos_over15_m_val = {
            'de_name': 'HIV+',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_sum,
        }
        pos_over15_m_val.update(_group_ou_dict)
        calculated_vals.append(pos_over15_m_val)

        tested_total = sum([under15_f_sum, under15_m_sum, over15_f_sum, over15_m_sum])
        pos_total = sum([pos_under15_f_sum, pos_under15_m_sum, pos_over15_f_sum, pos_over15_m_sum])
        tested_total_val = {
            'de_name': 'Tested',
            'cat_combo': None,
            'numeric_sum': tested_total,
        }
        tested_total_val.update(_group_ou_dict)
        pos_total_val = {
            'de_name': 'HIV+',
            'cat_combo': None,
            'numeric_sum': pos_total,
        }
        pos_total_val.update(_group_ou_dict)
        calculated_vals.append(tested_total_val)
        calculated_vals.append(pos_total_val)

        # copy linked to care totals over
        calculated_vals.append(linked_under15_f)
        calculated_vals.append(linked_under15_m)
        calculated_vals.append(linked_over15_f)
        calculated_vals.append(linked_over15_m)

        # calculate the percentages
        target_pos_under15_f, target_pos_under15_m, target_pos_over15_f, target_pos_over15_m, target_over15_f, target_over15_m, target_under15_f, target_under15_m, *further_vals = other_vals

        if all_not_none(under15_f_sum, target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            under15_f_percent = (under15_f_sum * 100) / target_under15_f['numeric_sum']
        else:
            under15_f_percent = None
        under15_f_percent_val = {
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_percent,
        }
        under15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(under15_f_percent_val)

        if all_not_none(under15_m_sum, target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            under15_m_percent = (under15_m_sum * 100) / target_under15_m['numeric_sum']
        else:
            under15_m_percent = None
        under15_m_percent_val = {
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_percent,
        }
        under15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(under15_m_percent_val)

        if all_not_none(over15_f_sum, target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            over15_f_percent = (over15_f_sum * 100) / target_over15_f['numeric_sum']
        else:
            over15_f_percent = None
        over15_f_percent_val = {
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_percent,
        }
        over15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(over15_f_percent_val)

        if all_not_none(over15_m_sum, target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            over15_m_percent = (over15_m_sum * 100) / target_over15_m['numeric_sum']
        else:
            over15_m_percent = None
        over15_m_percent_val = {
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_percent,
        }
        over15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(over15_m_percent_val)

        if all_not_none(pos_under15_f_sum, target_pos_under15_f['numeric_sum']) and target_pos_under15_f['numeric_sum']:
            pos_under15_f_percent = (pos_under15_f_sum * 100) / target_pos_under15_f['numeric_sum']
        else:
            pos_under15_f_percent = None
        pos_under15_f_percent_val = {
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_percent,
        }
        pos_under15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(pos_under15_f_percent_val)

        if all_not_none(pos_under15_m_sum, target_pos_under15_m['numeric_sum']) and target_pos_under15_m['numeric_sum']:
            pos_under15_m_percent = (pos_under15_m_sum * 100) / target_pos_under15_m['numeric_sum']
        else:
            pos_under15_m_percent = None
        pos_under15_m_percent_val = {
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_percent,
        }
        pos_under15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(pos_under15_m_percent_val)

        if all_not_none(pos_over15_f_sum, target_pos_over15_f['numeric_sum']) and target_pos_over15_f['numeric_sum']:
            pos_over15_f_percent = (pos_over15_f_sum * 100) / target_pos_over15_f['numeric_sum']
        else:
            pos_over15_f_percent = None
        pos_over15_f_percent_val = {
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_percent,
        }
        pos_over15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(pos_over15_f_percent_val)

        if all_not_none(pos_over15_m_sum, target_pos_over15_m['numeric_sum']) and target_pos_over15_m['numeric_sum']:
            pos_over15_m_percent = (pos_over15_m_sum * 100) / target_pos_over15_m['numeric_sum']
        else:
            pos_over15_m_percent = None
        pos_over15_m_percent_val = {
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_percent,
        }
        pos_over15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(pos_over15_m_percent_val)

        if all_not_none(linked_under15_f['numeric_sum'], pos_under15_f['numeric_sum']) and pos_under15_f['numeric_sum']:
            linked_under15_f_percent = (linked_under15_f['numeric_sum'] * 100) / pos_under15_f['numeric_sum']
        else:
            linked_under15_f_percent = None
        linked_under15_f_percent_val = {
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': linked_under15_f_percent,
        }
        linked_under15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(linked_under15_f_percent_val)

        if all_not_none(linked_under15_m['numeric_sum'], pos_under15_m['numeric_sum']) and pos_under15_m['numeric_sum']:
            linked_under15_m_percent = (linked_under15_m['numeric_sum'] * 100) / pos_under15_m['numeric_sum']
        else:
            linked_under15_m_percent = None
        linked_under15_m_percent_val = {
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': linked_under15_m_percent,
        }
        linked_under15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(linked_under15_m_percent_val)

        if all_not_none(linked_over15_f['numeric_sum'], pos_over15_f['numeric_sum']) and pos_over15_f['numeric_sum']:
            linked_over15_f_percent = (linked_over15_f['numeric_sum'] * 100) / pos_over15_f['numeric_sum']
        else:
            linked_over15_f_percent = None
        linked_over15_f_percent_val = {
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': linked_over15_f_percent,
        }
        linked_over15_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(linked_over15_f_percent_val)

        if all_not_none(linked_over15_m['numeric_sum'], pos_over15_m['numeric_sum']) and pos_over15_m['numeric_sum']:
            linked_over15_m_percent = (linked_over15_m['numeric_sum'] * 100) / pos_over15_m['numeric_sum']
        else:
            linked_over15_m_percent = None
        linked_over15_m_percent_val = {
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': linked_over15_m_percent,
        }
        linked_over15_m_percent_val.update(_group_ou_dict)
        calculated_vals.append(linked_over15_m_percent_val)

        # _group[1].extend(calculated_vals)
        _group[1] = calculated_vals
    
    data_element_metas = list()
    calc_subcategory_names = ('(<15, Female)', '(<15, Male)', '(15+, Female)', '(15+, Male)')
    
    data_element_metas += list(product(['Tested',], calc_subcategory_names))
    data_element_metas += list(product(['HIV+',], calc_subcategory_names))
    data_element_metas += list(product(['Tested',], [None,]))
    data_element_metas += list(product(['HIV+',], [None,]))
    data_element_metas += list(product(['Linked',], calc_subcategory_names))
    data_element_metas += list(product(['Tested (%)',], calc_subcategory_names))
    data_element_metas += list(product(['HIV+ (%)',], calc_subcategory_names))
    data_element_metas += list(product(['Linked (%)',], calc_subcategory_names))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    test_and_pos_ls = LegendSet()
    test_and_pos_ls.name = 'Testing/Positivity'
    test_and_pos_ls.add_interval('red', 0, 75)
    test_and_pos_ls.add_interval('yellow', 75, 90)
    test_and_pos_ls.add_interval('green', 90, None)
    for i in range(num_path_elements+14, num_path_elements+14+8):
        test_and_pos_ls.mappings[i] = True
    legend_sets.append(test_and_pos_ls)
    linked_ls = LegendSet()
    linked_ls.name = 'Link to Care'
    linked_ls.add_interval('red', 0, 80)
    linked_ls.add_interval('yellow', 80, 90)
    linked_ls.add_interval('green', 90, 100)
    for i in range(num_path_elements+14+8, num_path_elements+14+8+4):
        linked_ls.mappings[i] = True
    legend_sets.append(linked_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="hts_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="hts_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/hts_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def hts_by_district(request, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YRS = ['%d' % (y,) for y in range(this_year, this_year-6, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))

    if 'period' in request.GET and request.GET['period'] in PREV_5YRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d' % (this_year,)

    period_desc = filter_period

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    hts_de_names = (
        '105-4 Number of clients who have been linked to care',
        '105-4 Number of Individuals who received HIV test results',
        '105-4 Number of Individuals who tested HIV positive',
    )
    hts_short_names = (
        'Linked',
        'Tested',
        'HIV+',
    )
    subcategory_names = ['(<15, Female)', '(<15, Male)', '(15+, Female)', '(15+, Male)']
    de_positivity_meta = list(product(hts_de_names, subcategory_names))

    qs_positivity = DataValue.objects.what(*hts_de_names).filter(year=filter_period)
    if filter_district:
        qs_positivity = qs_positivity.where(filter_district)

    cc_lt_15 = ['18 Mths-<5 Years', '5-<10 Years', '10-<15 Years']
    cc_ge_15 = ['15-<19 Years', '19-<49 Years', '>49 Years']
    #TODO: cc_lt_15_f = CategoryCombo.from_cat_names(['Female', '<15']) gives a CategoryCombo instance that makes the Case statement clearer/safer
    qs_positivity = qs_positivity.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    qs_positivity = qs_positivity.exclude(cat_combo__iexact=None)

    qs_positivity = qs_positivity.annotate(district=F('org_unit__parent__parent__name'))
    qs_positivity = qs_positivity.annotate(period=F('year'))
    qs_positivity = qs_positivity.order_by('district', 'de_name', 'cat_combo', 'period')
    val_positivity = qs_positivity.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_positivity = list(val_positivity)
    
    # all districts (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=1).annotate(district=F('name'))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by('district')
    ou_list = list(v for v in qs_ou.values_list('district'))
    ou_headers = ['District',]

    def val_with_subcat_fun(row, col):
        district, = row
        de_name, subcategory = col
        return { 'district': district, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, de_positivity_meta, val_positivity, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_positivity2 = list(gen_raster)

    pmtct_mother_de_names = (
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.2a Women tested for HIV in labour (1st time this Pregnancy)',
        '105-2.3a Breastfeeding mothers tested for HIV(1st test)',
    )
    de_pmtct_mother_meta = list(product(('Pregnant Women tested for HIV',), (None,)))

    qs_pmtct_mother = DataValue.objects.what(*pmtct_mother_de_names).filter(year=filter_period)
    if filter_district:
        qs_pmtct_mother = qs_pmtct_mother.where(filter_district)
    qs_pmtct_mother = qs_pmtct_mother.annotate(de_name=Value('Pregnant Women tested for HIV', output_field=CharField()))
    qs_pmtct_mother = qs_pmtct_mother.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother = qs_pmtct_mother.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_mother = qs_pmtct_mother.annotate(period=F('year'))
    qs_pmtct_mother = qs_pmtct_mother.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother = qs_pmtct_mother.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_meta, val_pmtct_mother, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother2 = list(gen_raster)

    pmtct_mother_pos_de_names = (
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.2a Women testing HIV+ in labour (1st time this Pregnancy)',
        '105-2.2b Women testing HIV+ in labour (Retest this Pregnancy)',
        '105-2.3a Breastfeeding mothers newly testing HIV+(1st test)',
        '105-2.3b Breastfeeding mothers newly testing HIV+(retest)',
    )
    de_pmtct_mother_pos_meta = list(product(('Pregnant Women testing HIV+',), (None,)))

    qs_pmtct_mother_pos = DataValue.objects.what(*pmtct_mother_pos_de_names).filter(year=filter_period)
    if filter_district:
        qs_pmtct_mother_pos = qs_pmtct_mother_pos.where(filter_district)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(de_name=Value('Pregnant Women testing HIV+', output_field=CharField()))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(period=F('year'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother_pos = qs_pmtct_mother_pos.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_pos_meta, val_pmtct_mother_pos, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother_pos2 = list(gen_raster)

    pmtct_child_de_names = (
        '105-2.4a Exposed Infants Tested for HIV Below 18 Months(by 1st PCR) ',
        '105-2.4b 1st DNA PCR result returned(HIV+)',
        '105-2.4b 2nd DNA PCR result returned(HIV+)',
        '105-2.1a Male partners received HIV test results in eMTCT(Total)',
        '105-2.1b Male partners received HIV test results in eMTCT(HIV+)',
    )
    pmtct_child_short_names = (
        'PMTCT INFANT HIV+',
        'PMTCT CHILD PCR1 HIV+',
        'PMTCT CHILD PCR2 HIV+',
        'PMTCT MALE PARTNERS TESTED',
        'PMTCT MALE PARTNERS HIV+',
    )
    de_pmtct_child_meta = list(product(pmtct_child_de_names, (None,)))

    qs_pmtct_child = DataValue.objects.what(*pmtct_child_de_names).filter(year=filter_period)
    if filter_district:
        qs_pmtct_child = qs_pmtct_child.where(filter_district)
    qs_pmtct_child = qs_pmtct_child.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_child = qs_pmtct_child.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_child = qs_pmtct_child.annotate(period=F('year'))
    qs_pmtct_child = qs_pmtct_child.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_child = qs_pmtct_child.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_child_meta, val_pmtct_child, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_child2 = list(gen_raster)

    target_de_names = (
        'HTC_TST_TARGET',
        'HTC_TST_POS_TARGET',
    )
    de_target_meta = list(product(target_de_names, subcategory_names))

    # targets are annual, so filter by year component of period
    qs_target = DataValue.objects.what(*target_de_names).filter(year=filter_period[:4])
    if filter_district:
        qs_target = qs_target.where(filter_district)

    qs_target = qs_target.annotate(cat_combo=F('category_combo__name'))
    qs_target = qs_target.annotate(district=F('org_unit__parent__parent__name'))
    qs_target = qs_target.annotate(period=F('year'))
    qs_target = qs_target.order_by('district', '-de_name', 'cat_combo', 'period') # note reversed order of data element names
    val_target = qs_target.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_target = list(val_target)

    gen_raster = grabbag.rasterize(ou_list, de_target_meta, val_target, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_target2 = list(gen_raster)

    # combine the data and group by district
    grouped_vals = groupbylist(sorted(chain(val_positivity2, val_pmtct_mother2, val_pmtct_mother_pos2, val_pmtct_child2, val_target2), key=lambda x: (x['district'],)), key=lambda x: (x['district'],))

    # perform calculations
    for _group in grouped_vals:
        (ou_path_list, (linked_under15_f, linked_under15_m, linked_over15_f, linked_over15_m, tst_under15_f, tst_under15_m, tst_over15_f, tst_over15_m, pos_under15_f, pos_under15_m, pos_over15_f, pos_over15_m, tst_pregnant, pos_pregnant, pos_infant, pos_pcr1, pos_pcr2, tst_male_partner, pos_male_partner, *other_vals)) = _group
        
        calculated_vals = list()

        under15_f_sum = default_zero(tst_under15_f['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_sum,
        }
        calculated_vals.append(under15_f_val)
        
        under15_m_sum = default_zero(tst_under15_m['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_sum,
        }
        calculated_vals.append(under15_m_val)
        
        over15_f_sum = default_zero(tst_over15_f['numeric_sum']) + default_zero(tst_pregnant['numeric_sum'])
        over15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_sum,
        }
        calculated_vals.append(over15_f_val)
        
        over15_m_sum = default_zero(tst_over15_m['numeric_sum']) + default_zero(tst_male_partner['numeric_sum'])
        over15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_sum,
        }
        calculated_vals.append(over15_m_val)
        
        half_pos_pcr = Decimal(default_zero(pos_pcr1['numeric_sum']) + default_zero(pos_pcr1['numeric_sum']))/2
        pos_under15_f_sum = default_zero(pos_under15_f['numeric_sum']) + half_pos_pcr
        pos_under15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_sum,
        }
        calculated_vals.append(pos_under15_f_val)
        
        pos_under15_m_sum = default_zero(pos_under15_m['numeric_sum']) + half_pos_pcr
        pos_under15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_sum,
        }
        calculated_vals.append(pos_under15_m_val)
        
        pos_over15_f_sum = default_zero(pos_over15_f['numeric_sum']) + Decimal(default_zero(pos_pregnant['numeric_sum']))
        pos_over15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_sum,
        }
        calculated_vals.append(pos_over15_f_val)
        
        pos_over15_m_sum = default_zero(pos_over15_m['numeric_sum']) + Decimal(default_zero(pos_male_partner['numeric_sum']))
        pos_over15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_sum,
        }
        calculated_vals.append(pos_over15_m_val)

        tested_total = sum([under15_f_sum, under15_m_sum, over15_f_sum, over15_m_sum])
        pos_total = sum([pos_under15_f_sum, pos_under15_m_sum, pos_over15_f_sum, pos_over15_m_sum])
        tested_total_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': None,
            'numeric_sum': tested_total,
        }
        pos_total_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': None,
            'numeric_sum': pos_total,
        }
        calculated_vals.append(tested_total_val)
        calculated_vals.append(pos_total_val)

        # copy linked to care totals over
        calculated_vals.append(linked_under15_f)
        calculated_vals.append(linked_under15_m)
        calculated_vals.append(linked_over15_f)
        calculated_vals.append(linked_over15_m)

        # calculate the percentages
        target_under15_f, target_under15_m, target_over15_f, target_over15_m, target_pos_under15_f, target_pos_under15_m, target_pos_over15_f, target_pos_over15_m, *further_vals = other_vals

        if all_not_none(under15_f_sum, target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            under15_f_percent = (under15_f_sum * 100) / target_under15_f['numeric_sum']
        else:
            under15_f_percent = None
        under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_percent,
        }
        calculated_vals.append(under15_f_percent_val)

        if all_not_none(under15_m_sum, target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            under15_m_percent = (under15_m_sum * 100) / target_under15_m['numeric_sum']
        else:
            under15_m_percent = None
        under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_percent,
        }
        calculated_vals.append(under15_m_percent_val)

        if all_not_none(over15_f_sum, target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            over15_f_percent = (over15_f_sum * 100) / target_over15_f['numeric_sum']
        else:
            over15_f_percent = None
        over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_percent,
        }
        calculated_vals.append(over15_f_percent_val)

        if all_not_none(over15_m_sum, target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            over15_m_percent = (over15_m_sum * 100) / target_over15_m['numeric_sum']
        else:
            over15_m_percent = None
        over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_percent,
        }
        calculated_vals.append(over15_m_percent_val)

        if all_not_none(pos_under15_f_sum, target_pos_under15_f['numeric_sum']) and target_pos_under15_f['numeric_sum']:
            pos_under15_f_percent = (pos_under15_f_sum * 100) / target_pos_under15_f['numeric_sum']
        else:
            pos_under15_f_percent = None
        pos_under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_percent,
        }
        calculated_vals.append(pos_under15_f_percent_val)

        if all_not_none(pos_under15_m_sum, target_pos_under15_m['numeric_sum']) and target_pos_under15_m['numeric_sum']:
            pos_under15_m_percent = (pos_under15_m_sum * 100) / target_pos_under15_m['numeric_sum']
        else:
            pos_under15_m_percent = None
        pos_under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_percent,
        }
        calculated_vals.append(pos_under15_m_percent_val)

        if all_not_none(pos_over15_f_sum, target_pos_over15_f['numeric_sum']) and target_pos_over15_f['numeric_sum']:
            pos_over15_f_percent = (pos_over15_f_sum * 100) / target_pos_over15_f['numeric_sum']
        else:
            pos_over15_f_percent = None
        pos_over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_percent,
        }
        calculated_vals.append(pos_over15_f_percent_val)

        if all_not_none(pos_over15_m_sum, target_pos_over15_m['numeric_sum']) and target_pos_over15_m['numeric_sum']:
            pos_over15_m_percent = (pos_over15_m_sum * 100) / target_pos_over15_m['numeric_sum']
        else:
            pos_over15_m_percent = None
        pos_over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_percent,
        }
        calculated_vals.append(pos_over15_m_percent_val)

        if all_not_none(linked_under15_f['numeric_sum'], pos_under15_f['numeric_sum']) and pos_under15_f['numeric_sum']:
            linked_under15_f_percent = (linked_under15_f['numeric_sum'] * 100) / pos_under15_f['numeric_sum']
        else:
            linked_under15_f_percent = None
        linked_under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': linked_under15_f_percent,
        }
        calculated_vals.append(linked_under15_f_percent_val)

        if all_not_none(linked_under15_m['numeric_sum'], pos_under15_m['numeric_sum']) and pos_under15_m['numeric_sum']:
            linked_under15_m_percent = (linked_under15_m['numeric_sum'] * 100) / pos_under15_m['numeric_sum']
        else:
            linked_under15_m_percent = None
        linked_under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': linked_under15_m_percent,
        }
        calculated_vals.append(linked_under15_m_percent_val)

        if all_not_none(linked_over15_f['numeric_sum'], pos_over15_f['numeric_sum']) and pos_over15_f['numeric_sum']:
            linked_over15_f_percent = (linked_over15_f['numeric_sum'] * 100) / pos_over15_f['numeric_sum']
        else:
            linked_over15_f_percent = None
        linked_over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': linked_over15_f_percent,
        }
        calculated_vals.append(linked_over15_f_percent_val)

        if all_not_none(linked_over15_m['numeric_sum'], pos_over15_m['numeric_sum']) and pos_over15_m['numeric_sum']:
            linked_over15_m_percent = (linked_over15_m['numeric_sum'] * 100) / pos_over15_m['numeric_sum']
        else:
            linked_over15_m_percent = None
        linked_over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': linked_over15_m_percent,
        }
        calculated_vals.append(linked_over15_m_percent_val)

        _group[1] = calculated_vals
    
    data_element_names = list()

    data_element_names += list(product(['Tested',], subcategory_names))
    data_element_names += list(product(['HIV+',], subcategory_names))
    data_element_names += list(product(['Tested',], [None,]))
    data_element_names += list(product(['HIV+',], [None,]))
    data_element_names += list(product(['Linked',], subcategory_names))
    data_element_names += list(product(['Tested (%)',], subcategory_names))
    data_element_names += list(product(['HIV+ (%)',], subcategory_names))
    data_element_names += list(product(['Linked (%)',], subcategory_names))

    legend_sets = list()
    test_and_pos_ls = LegendSet()
    test_and_pos_ls.name = 'Testing/Positivity'
    test_and_pos_ls.add_interval('red', 0, 75)
    test_and_pos_ls.add_interval('yellow', 75, 90)
    test_and_pos_ls.add_interval('green', 90, None)
    for i in range(15, 15+8):
        test_and_pos_ls.mappings[i] = True
    legend_sets.append(test_and_pos_ls)
    linked_ls = LegendSet()
    linked_ls.name = 'Link to Care'
    linked_ls.add_interval('red', 0, 80)
    linked_ls.add_interval('yellow', 80, 90)
    linked_ls.add_interval('green', 90, 100)
    for i in range(15+8, 15+8+4):
        linked_ls.mappings[i] = True
    legend_sets.append(linked_ls)

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_names)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="hts_districts_scorecard.xlsx"'

        return response

    context = {
        'grouped_data': grouped_vals,
        'data_element_names': data_element_names,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YRS,
        'district_list': DISTRICT_LIST,
    }

    return render(request, 'cannula/hts_districts.html', context)

@login_required
def care_tx_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    nut_assessed_de_names = (
        '106a ART No. active on ART assessed for Malnutrition at their visit in quarter',
        '106a Pre-ART No. Active on pre-ART Care  assessed for Malnutrition at their visit in quarter',
    )
    nut_assessed_short_names = (
        'PLHIV in care and treatment who had a nutrition assessment',
    )
    de_nut_assessed_meta = list(product(('PLHIV in care and treatment who had a nutrition assessment',), (None,)))
    data_element_metas += list(product(nut_assessed_short_names, (None,)))

    qs_nut_assessed = DataValue.objects.what(*nut_assessed_de_names)
    qs_nut_assessed = qs_nut_assessed.annotate(de_name=Value('PLHIV in care and treatment who had a nutrition assessment', output_field=CharField()))
    qs_nut_assessed = qs_nut_assessed.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_nut_assessed = qs_nut_assessed.where(filter_district)
    qs_nut_assessed = qs_nut_assessed.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_nut_assessed = qs_nut_assessed.when(filter_period)
    qs_nut_assessed = qs_nut_assessed.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_nut_assessed = qs_nut_assessed.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_nut_assessed_meta, val_nut_assessed, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_nut_assessed2 = list(gen_raster)
    
    care_tx_total_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
        '106a Pre-ART No. of active clients  on pre-ART Care in the  quarter',
    )
    care_tx_total_short_names = (
        'PLHIV in care and treatment',
    )
    de_care_tx_total_meta = list(product(('PLHIV in care and treatment',), (None,)))
    data_element_metas += list(product(care_tx_total_short_names, (None,)))

    qs_care_tx_total = DataValue.objects.what(*care_tx_total_de_names)
    qs_care_tx_total = qs_care_tx_total.annotate(de_name=Value('PLHIV in care and treatment', output_field=CharField()))
    qs_care_tx_total = qs_care_tx_total.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_care_tx_total = qs_care_tx_total.where(filter_district)
    qs_care_tx_total = qs_care_tx_total.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_care_tx_total = qs_care_tx_total.when(filter_period)
    qs_care_tx_total = qs_care_tx_total.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_care_tx_total = qs_care_tx_total.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_care_tx_total_meta, val_care_tx_total, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_care_tx_total2 = list(gen_raster)

    tb_screened_de_names = (
        '106a ART No. active on ART assessed for TB at last visit in the  quarter',
    )
    tb_screened_short_names = (
        'Active on Pre-ART/ART patients screened for TB',
    )
    de_tb_screened_meta = list(product(('Active on Pre-ART/ART patients screened for TB',), (None,)))
    data_element_metas += list(product(tb_screened_short_names, (None,)))

    qs_tb_screened = DataValue.objects.what(*tb_screened_de_names)
    qs_tb_screened = qs_tb_screened.annotate(de_name=Value('Active on Pre-ART/ART patients screened for TB', output_field=CharField()))
    qs_tb_screened = qs_tb_screened.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_tb_screened = qs_tb_screened.where(filter_district)
    qs_tb_screened = qs_tb_screened.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_tb_screened = qs_tb_screened.when(filter_period)
    qs_tb_screened = qs_tb_screened.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_tb_screened = qs_tb_screened.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_tb_screened_meta, val_tb_screened, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_tb_screened2 = list(gen_raster)

    art_total_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    art_total_short_names = (
        'Active on Pre-ART/ART patients',
    )
    de_art_total_meta = list(product(('Active on Pre-ART/ART patients',), (None,)))
    data_element_metas += list(product(art_total_short_names, (None,)))

    qs_art_total = DataValue.objects.what(*art_total_de_names)
    qs_art_total = qs_art_total.annotate(de_name=Value('Active on Pre-ART/ART patients', output_field=CharField()))
    qs_art_total = qs_art_total.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_art_total = qs_art_total.where(filter_district)
    qs_art_total = qs_art_total.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_total = qs_art_total.when(filter_period)
    qs_art_total = qs_art_total.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_total = qs_art_total.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_art_total_meta, val_art_total, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_total2 = list(gen_raster)

    cohort_12_months_de_names = (
        '106a Cohort  All patients 12 months Alive on ART in Cohort',
        '106a Cohort  All patients 12 months Started on ART-Cohort',
        '106a Cohort  All patients 12 months Transfered In',
        '106a Cohort  All patients 12 months Transferred Out',
    )
    cohort_12_months_short_names = (
        '12 months cohort - Alive',
        '12 months cohort - Started',
        '12 months cohort - Transfered In',
        '12 months cohort - Transferred Out',
    )
    de_cohort_12_months_meta = list(product(cohort_12_months_de_names, (None,)))
    data_element_metas += list(product(cohort_12_months_short_names, (None,)))

    qs_cohort_12_months = DataValue.objects.what(*cohort_12_months_de_names)
    qs_cohort_12_months = qs_cohort_12_months.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_cohort_12_months = qs_cohort_12_months.where(filter_district)
    qs_cohort_12_months = qs_cohort_12_months.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_cohort_12_months = qs_cohort_12_months.when(filter_period)
    qs_cohort_12_months = qs_cohort_12_months.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_cohort_12_months = qs_cohort_12_months.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_cohort_12_months_meta, val_cohort_12_months, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_cohort_12_months2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_nut_assessed2, val_care_tx_total2, val_tb_screened2, val_art_total2, val_cohort_12_months2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (care_tx_nut_assessed, care_tx_total, tb_screened, art_total, cohort_alive, cohort_started, cohort_in, cohort_out, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(care_tx_total['numeric_sum'], care_tx_nut_assessed['numeric_sum']) and care_tx_total['numeric_sum']:
            nut_assessed_percent = (care_tx_nut_assessed['numeric_sum'] * 100) / care_tx_total['numeric_sum']
        else:
            nut_assessed_percent = None
        nut_assessed_percent_val = {
            'de_name': '% of PLHIV in care and treatment who had a nutrition assessment conducted',
            'cat_combo': None,
            'numeric_sum': nut_assessed_percent,
        }
        nut_assessed_percent_val.update(_group_ou_dict)
        calculated_vals.append(nut_assessed_percent_val)

        if all_not_none(art_total['numeric_sum'], tb_screened['numeric_sum']) and art_total['numeric_sum']:
            tb_screened_percent = (tb_screened['numeric_sum'] * 100) / art_total['numeric_sum']
        else:
            tb_screened_percent = None
        tb_screened_percent_val = {
            'de_name': '% of ART patients who were screened for TB',
            'cat_combo': None,
            'numeric_sum': tb_screened_percent,
        }
        tb_screened_percent_val.update(_group_ou_dict)
        calculated_vals.append(tb_screened_percent_val)

        cohort_net = sum_zero(cohort_started['numeric_sum'], cohort_in['numeric_sum'])-default_zero(cohort_out['numeric_sum'])
        if all_not_none(cohort_alive['numeric_sum'], cohort_net) and cohort_net:
            cohort_survival_percent = (cohort_alive['numeric_sum'] * 100) / cohort_net
        else:
            cohort_survival_percent = None
        cohort_survival_percent_val = {
            'de_name': '% of adults and children known to be on Tx 12 months after initiatio of ART',
            'cat_combo': None,
            'numeric_sum': cohort_survival_percent,
        }
        cohort_survival_percent_val.update(_group_ou_dict)
        calculated_vals.append(cohort_survival_percent_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['% of PLHIV in care and treatment who had a nutrition assessment conducted'], (None,)))
    data_element_metas += list(product(['% of ART patients who were screened for TB'], (None,)))
    data_element_metas += list(product(['% of adults and children known to be on Tx 12 months after initiatio of ART'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    retention_ls = LegendSet()
    retention_ls.name = '12 month Retention'
    retention_ls.add_interval('red', 0, 80)
    retention_ls.add_interval('yellow', 80, 85)
    retention_ls.add_interval('green', 85, None)
    retention_ls.mappings[num_path_elements+10] = True
    legend_sets.append(retention_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="vmmc_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="care_tx_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/care_tx_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def pmtct_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'Target: PMTCT_ART',
        'Target: PMTCT_EID',
        'Target: PMTCT_EID_POS',
        'Target: PMTCT_STAT',
        'Target: PMTCT_STAT_POS',
        'Target: new ANC clients',
    )
    targets_short_names = (
        'Target: PMTCT_ART',
        'Target: PMTCT_EID',
        'Target: PMTCT_EID_POS',
        'Target: PMTCT_STAT',
        'Target: PMTCT_STAT_POS',
        'Target: new ANC clients',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))
    data_element_metas += list(product(targets_short_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    pmtct_de_names = (
        '105-2.1 A17:HIV+ Pregnant Women already on ART before 1st ANC (ART-K)',
        '105-2.1 A18:Pregnant Women re-tested later in pregnancy (TR+ &TRR+)',
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.1 A1:ANC 1st Visit for women',
        '105-2.1 HIV+ Pregnant Women initiated on ART for EMTCT (ART)',
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.1 Pregnant Women tested HIV+ for 1st time this pregnancy (TRR) at any visit',
        '105-2.1a Pregnant Women who knew status before 1st ANC (Total (TRK + TRRK))',
        '105-2.1b Pregnant Women who knew status before 1st ANC (HIV+(TRRK))',
    )
    pmtct_short_names = (
        '105-2.1 A17:HIV+ Pregnant Women already on ART before 1st ANC (ART-K)',
        '105-2.1 A18:Pregnant Women re-tested later in pregnancy (TR+ &TRR+)',
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.1 A1:ANC 1st Visit for women',
        '105-2.1 HIV+ Pregnant Women initiated on ART for EMTCT (ART)',
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.1 Pregnant Women tested HIV+ for 1st time this pregnancy (TRR) at any visit',
        '105-2.1a Pregnant Women who knew status before 1st ANC (Total (TRK + TRRK))',
        '105-2.1b Pregnant Women who knew status before 1st ANC (HIV+(TRRK))',
    )
    de_pmtct_meta = list(product(pmtct_short_names, (None,)))
    data_element_metas += de_pmtct_meta

    qs_pmtct = DataValue.objects.what(*pmtct_de_names)
    # qs_pmtct = qs_pmtct.annotate(de_name=Value(pmtct_short_names[0], output_field=CharField()))
    qs_pmtct = qs_pmtct.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_pmtct = qs_pmtct.where(filter_district)
    qs_pmtct = qs_pmtct.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_pmtct = qs_pmtct.when(filter_period)
    qs_pmtct = qs_pmtct.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_pmtct = qs_pmtct.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_pmtct = list(val_pmtct)

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_meta, val_pmtct, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_pmtct2 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_pmtct2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (target_art, target_eid, target_eid_pos, target_known, target_known_pos, target_new, art_already, hiv_retested, hiv_retested_pos, anc_1_visit, art_initiated, hiv_tested, hiv_tested_pos, hiv_known, hiv_known_pos, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(anc_1_visit['numeric_sum'], target_new['numeric_sum']) and target_new['numeric_sum']:
            perf_anc_1_visit = 100 * anc_1_visit['numeric_sum'] / target_new['numeric_sum']
        else:
            perf_anc_1_visit = None
        perf_anc_1_visit_val = {
            'de_name': 'Perf. %: New ANC1 clients',
            'cat_combo': None,
            'numeric_sum': perf_anc_1_visit,
        }
        perf_anc_1_visit_val.update(_group_ou_dict)
        calculated_vals.append(perf_anc_1_visit_val)

        if all_not_none(target_known['numeric_sum']) and target_known['numeric_sum']:
            perf_hiv_known = 100 * sum_zero(hiv_tested['numeric_sum'], hiv_known['numeric_sum']) / target_known['numeric_sum']
        else:
            perf_hiv_known = None
        perf_hiv_known_val = {
            'de_name': 'Perf. %: Pregnant women with known HIV status',
            'cat_combo': None,
            'numeric_sum': perf_hiv_known,
        }
        perf_hiv_known_val.update(_group_ou_dict)
        calculated_vals.append(perf_hiv_known_val)

        if all_not_none(anc_1_visit['numeric_sum']) and anc_1_visit['numeric_sum']:
            pmtct_stat = 100 * sum_zero(hiv_tested['numeric_sum'], hiv_known['numeric_sum']) / anc_1_visit['numeric_sum']
        else:
            pmtct_stat = None
        pmtct_stat_val = {
            'de_name': 'PMTCT_STAT %',
            'cat_combo': None,
            'numeric_sum': pmtct_stat,
        }
        pmtct_stat_val.update(_group_ou_dict)
        calculated_vals.append(pmtct_stat_val)

        if all_not_none(target_known_pos['numeric_sum']) and target_known_pos['numeric_sum']:
            perf_hiv_known_pos = 100 * sum_zero(hiv_tested_pos['numeric_sum'], hiv_known_pos['numeric_sum']) / target_known_pos['numeric_sum']
        else:
            perf_hiv_known_pos = None
        perf_hiv_known_pos_val = {
            'de_name': 'Perf. %: Pregnant women with known HIV+ status',
            'cat_combo': None,
            'numeric_sum': perf_hiv_known_pos,
        }
        perf_hiv_known_pos_val.update(_group_ou_dict)
        calculated_vals.append(perf_hiv_known_pos_val)

        if all_not_none(anc_1_visit['numeric_sum']) and anc_1_visit['numeric_sum']:
            pmtct_stat_pos = 100 * sum_zero(hiv_tested_pos['numeric_sum'], hiv_known_pos['numeric_sum']) / anc_1_visit['numeric_sum']
        else:
            pmtct_stat_pos = None
        pmtct_stat_pos_val = {
            'de_name': 'PMTCT_STAT_POS %',
            'cat_combo': None,
            'numeric_sum': pmtct_stat_pos,
        }
        pmtct_stat_pos_val.update(_group_ou_dict)
        calculated_vals.append(pmtct_stat_pos_val)

        if all_not_none(target_art['numeric_sum']) and target_art['numeric_sum']:
            perf_art = 100 * sum_zero(art_initiated['numeric_sum']) / target_art['numeric_sum']
        else:
            perf_art = None
        perf_art_val = {
            'de_name': 'Perf. %: HIV+ pregnant women who received ART to reduce the risk of MTCT',
            'cat_combo': None,
            'numeric_sum': perf_art,
        }
        perf_art_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_val)

        if all_not_none(hiv_tested_pos['numeric_sum'], hiv_known_pos['numeric_sum']) and sum_zero(hiv_tested_pos['numeric_sum'], hiv_known_pos['numeric_sum']):
            pmtct_art = 100 * sum_zero(art_already['numeric_sum'], art_initiated['numeric_sum']) / sum_zero(hiv_tested_pos['numeric_sum'], hiv_known_pos['numeric_sum'])
        else:
            pmtct_art = None
        pmtct_art_val = {
            'de_name': 'PMTCT_ART %',
            'cat_combo': None,
            'numeric_sum': pmtct_art,
        }
        pmtct_art_val.update(_group_ou_dict)
        calculated_vals.append(pmtct_art_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Perf. %: New ANC1 clients'], (None,)))
    data_element_metas += list(product(['Perf. %: Pregnant women with known HIV status'], (None,)))
    data_element_metas += list(product(['PMTCT_STAT %'], (None,)))
    data_element_metas += list(product(['Perf. %: Pregnant women with known HIV+ status'], (None,)))
    data_element_metas += list(product(['PMTCT_STAT_POS %'], (None,)))
    data_element_metas += list(product(['Perf. %: HIV+ pregnant women who received ART to reduce the risk of MTCT'], (None,)))
    data_element_metas += list(product(['PMTCT_ART %'], (None,)))


    num_path_elements = len(ou_headers)
    legend_sets = list()
    pmtct_ls = LegendSet()
    pmtct_ls.name = 'PMTCT (Target Performance)'
    pmtct_ls.add_interval('red', 0, 70)
    pmtct_ls.add_interval('yellow', 70, 90)
    pmtct_ls.add_interval('green', 90, None)
    pmtct_ls.mappings[num_path_elements+15] = True
    pmtct_ls.mappings[num_path_elements+16] = True
    pmtct_ls.mappings[num_path_elements+18] = True
    pmtct_ls.mappings[num_path_elements+20] = True
    legend_sets.append(pmtct_ls)
    pmtct_ls = LegendSet()
    pmtct_ls.name = 'PMTCT (PEPFAR)'
    pmtct_ls.add_interval('red', 0, 80)
    pmtct_ls.add_interval('yellow', 80, 90)
    pmtct_ls.add_interval('green', 90, 100)
    pmtct_ls.mappings[num_path_elements+17] = True
    pmtct_ls.mappings[num_path_elements+19] = True
    pmtct_ls.mappings[num_path_elements+21] = True
    legend_sets.append(pmtct_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pmtct_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="pmtct_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/pmtct_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

def vmmc_dashboard(request):
    this_day = date.today()
    this_quarter = '%d-Q%d' % (this_day.year, month2quarter(this_day.month))
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_day.year, this_day.year-6, -1) for q in range(4, 0, -1)]
    period_list = list(filter(lambda qtr: qtr < this_quarter, reversed(PREV_5YR_QTRS)))[-6:]
    def val_with_period_de_fun(row, col):
        period = row
        de_name = col
        return { 'de_name': de_name, 'period': period, 'numeric_sum': None }

    data_element_metas = list()

    vmmc_target_de_names = (
        'VMMC_CIRC_TARGET',
    )
    vmmc_target_short_names = (
        'VMMC_CIRC_TARGET',
    )
    de_vmmc_target_meta = list(product(vmmc_target_short_names, (None,)))

    qs_vmmc_target = DataValue.objects.what(*vmmc_target_de_names)
    qs_vmmc_target = qs_vmmc_target.annotate(cat_combo=Value(None, output_field=CharField()))
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    year_list = sorted(list(set(qstr[:4] for qstr in period_list)))
    qs_vmmc_target = qs_vmmc_target.when(*year_list)
    qs_vmmc_target = qs_vmmc_target.order_by('period', 'de_name')
    val_vmmc_target = qs_vmmc_target.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_vmmc_target = list(val_vmmc_target)
    q_list = sorted(['%s-Q%d' % (y, qnum,) for qnum in range(1, 5) for y in year_list])
    q_list = [x for x in q_list if x in period_list]

    def duplicate_values_over_periods(val_list, p_list):
        def assign_period(v, p):
            new_val = dict(v)
            new_val['period'] = p
            return new_val

        return [assign_period(val, period) for val in val_list for period in p_list]

    val_vmmc_target = duplicate_values_over_periods(val_vmmc_target, q_list)

    gen_raster = grabbag.rasterize(period_list, vmmc_target_de_names, val_vmmc_target, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_vmmc_target2 = list(gen_raster)

    vmmc_de_names = (
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Moderate',
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Severe',
        '105-5 Clients circumcised by circumcision Technique Device Based (DC)',
        '105-5 Clients circumcised by circumcision Technique Other VMMC techniques',
        '105-5 Clients circumcised by circumcision Technique Surgical(SC)',
        '105-5a Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 48 Hours)',
    )
    vmmc_short_names = (
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Moderate',
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Severe',
        '105-5 Clients circumcised by circumcision Technique Device Based (DC)',
        '105-5 Clients circumcised by circumcision Technique Other VMMC techniques',
        '105-5 Clients circumcised by circumcision Technique Surgical(SC)',
        '105-5a Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 48 Hours)',
    )
    de_vmmc_meta = list(product(vmmc_short_names, (None,)))

    qs_vmmc = DataValue.objects.what(*vmmc_de_names)
    qs_vmmc = qs_vmmc.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_vmmc = qs_vmmc.when(*period_list)
    qs_vmmc = qs_vmmc.order_by('period', 'de_name')
    val_vmmc = qs_vmmc.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_vmmc = list(val_vmmc)

    gen_raster = grabbag.rasterize(period_list, vmmc_de_names, val_vmmc, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_vmmc2 = list(gen_raster)

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_vmmc_target2, val_vmmc2), key=lambda x: (x['period'])), key=lambda x: (x['period']))
    # if True:
    #     grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (period, (vmmc_target, adverse_moderate, adverse_severe, vmmc_device, vmmc_other, vmmc_surgical, followup_48hrs, *other_vals)) = _group
        
        calculated_vals = list()

        if all_not_none(vmmc_target['numeric_sum']) and vmmc_target['numeric_sum']:
            vmmc_total = sum_zero(vmmc_device['numeric_sum'], vmmc_other['numeric_sum'], vmmc_surgical['numeric_sum'])
            coverage_percent = (vmmc_total * 100) / vmmc_target['numeric_sum']
        else:
            coverage_percent = None
        coverage_percent_val = {
            'period': period,
            'de_name': '% PEPFAR targeted Males circumcised',
            'numeric_sum': coverage_percent,
        }
        calculated_vals.append(coverage_percent_val)

        if all_not_none(vmmc_total) and vmmc_total:
            vmmc_adverse = sum_zero(adverse_moderate['numeric_sum'], adverse_severe['numeric_sum'])
            adverse_percent = (vmmc_adverse * 100) / vmmc_total
        else:
            adverse_percent = None
        adverse_percent_val = {
            'period': period,
            'de_name': '% Experienced Adverse Events',
            'numeric_sum': adverse_percent,
        }
        calculated_vals.append(adverse_percent_val)

        if all_not_none(followup_48hrs['numeric_sum'], vmmc_total) and vmmc_total:
            followup_percent = (followup_48hrs['numeric_sum'] * 100) / vmmc_total
        else:
            followup_percent = None
        followup_percent_val = {
            'period': period,
            'de_name': '% Followed up at 48 hours',
            'numeric_sum': followup_percent,
        }
        calculated_vals.append(followup_percent_val)

        _group[1] = calculated_vals
    
    context = {
        'data_element_names': [
            ('% PEPFAR targeted Males circumcised', None),
            ('% Experienced Adverse Events', None),
            ('% Followed up at 48 hours', None),
        ],
        'grouped_data': grouped_vals,
    }
    return render(request, 'cannula/index.html', context)

@login_required
def vmmc_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'VMMC_CIRC_TARGET',
        'VMMC_DEVICE_TARGET',
        'VMMC_SURGICAL_TARGET',
    )
    targets_short_names = (
        'TARGET: VMMC_CIRC',
        'TARGET: Device-based',
        'TARGET: Surgical',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))
    data_element_metas += list(product(targets_short_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_targets = qs_targets.when(filter_period)
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    method_de_names = (
        '105-5 Clients circumcised by circumcision Technique Device Based (DC)',
        '105-5 Clients circumcised by circumcision Technique Other VMMC techniques',
        '105-5 Clients circumcised by circumcision Technique Surgical(SC)',
    )
    method_short_names = (
        'Circumcised by technique - Device Based',
        'Circumcised by technique - Other',
        'Circumcised by technique - Surgical',
    )
    de_method_meta = list(product(method_de_names, (None,)))
    data_element_metas += list(product(method_short_names, (None,)))

    qs_method = DataValue.objects.what(*method_de_names)
    qs_method = qs_method.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_method = qs_method.where(filter_district)
    qs_method = qs_method.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_method = qs_method.when(filter_period)
    qs_method = qs_method.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_method = qs_method.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_method_meta, val_method, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_method2 = list(gen_raster)

    hiv_de_names = (
        '105-5 SMC Clients Counseled, Tested and Circumcised for HIV at SMC site HIV Negative',
        '105-5 SMC Clients Counseled, Tested and Circumcised for HIV at SMC site HIV Positive',
    )
    hiv_short_names = (
        'Circumcised by HIV status - Negative',
        'Circumcised by HIV status - Positive',
    )
    de_hiv_meta = list(product(hiv_de_names, (None,)))
    data_element_metas += list(product(hiv_short_names, (None,)))

    qs_hiv = DataValue.objects.what(*hiv_de_names)
    qs_hiv = qs_hiv.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv = qs_hiv.where(filter_district)
    qs_hiv = qs_hiv.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv = qs_hiv.when(filter_period)
    qs_hiv = qs_hiv.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv = qs_hiv.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_hiv_meta, val_hiv, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv2 = list(gen_raster)

    location_de_names = (
        '105-5 Number of Males Circumcised by Age group and Technique Facility, Device Based (DC)',
        '105-5 Number of Males Circumcised by Age group and Technique Facility, Surgical(SC)',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach, Device Based (DC)',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach, Surgical(SC)',
    )
    location_de_names2 = (
        '105-5 Number of Males Circumcised by Age group and Technique Facility',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach',
    )
    location_prefix_len = len('105-5 Number of Males Circumcised by Age group and Technique Facility')
    location_short_names = (
        'Circumcised by site type - Static',
        'Circumcised by site type - Mobile',
    )
    de_location_meta = list(product(location_de_names2, (None,)))
    data_element_metas += list(product(location_short_names, (None,)))

    qs_location = DataValue.objects.what(*location_de_names)
    # drop the technique section from the returned data element name
    qs_location = qs_location.annotate(de_name=Substr('data_element__name', 1, location_prefix_len))
    qs_location = qs_location.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_location = qs_location.where(filter_district)
    qs_location = qs_location.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_location = qs_location.when(filter_period)
    qs_location = qs_location.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_location = qs_location.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_location_meta, val_location, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_location2 = list(gen_raster)

    followup_de_names = (
        '105-5a Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 48 Hours)',
        '105-5b Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 7 Days)',
        '105-5c Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Beyond 7 Days)',
    )
    followup_short_names = (
        'Follow up - Within 48 hours',
        'Follow up - Within 7 days',
        'Follow up - Beyond 7 days',
    )
    de_followup_meta = list(product(followup_de_names, (None,)))
    data_element_metas += list(product(followup_short_names, (None,)))

    qs_followup = DataValue.objects.what(*followup_de_names)
    qs_followup = qs_followup.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_followup = qs_followup.where(filter_district)
    qs_followup = qs_followup.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_followup = qs_followup.when(filter_period)
    qs_followup = qs_followup.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_followup = qs_followup.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_followup_meta, val_followup, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_followup2 = list(gen_raster)

    adverse_de_names = (
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Moderate',
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Severe',
    )
    adverse_short_names = (
        'Adverse Events - Moderate',
        'Adverse Events - Severe',
    )
    de_adverse_meta = list(product(adverse_de_names, (None,)))
    data_element_metas += list(product(adverse_short_names, (None,)))

    qs_adverse = DataValue.objects.what(*adverse_de_names)
    qs_adverse = qs_adverse.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_adverse = qs_adverse.where(filter_district)
    qs_adverse = qs_adverse.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_adverse = qs_adverse.when(filter_period)
    qs_adverse = qs_adverse.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_adverse = qs_adverse.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_adverse_meta, val_adverse, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_adverse2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_hiv2, val_location2, val_method2, val_followup2, val_adverse2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (target_total, target_device, target_surgical, hiv_negative, hiv_positive, location_facility, location_outreach, method_device, method_other, method_surgical, followup_48hrs, followup_7days, followup_plus7days, adverse_moderate, adverse_severe, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        method_sum = default_zero(method_device['numeric_sum']) + default_zero(method_surgical['numeric_sum']) + default_zero(method_other['numeric_sum'])

        if all_not_none(target_total['numeric_sum'], method_sum) and target_total['numeric_sum']:
            target_total_percent = (method_sum * 100) / target_total['numeric_sum']
        else:
            target_total_percent = None
        target_total_percent_val = {
            'de_name': 'Perf% Circumcised',
            'cat_combo': None,
            'numeric_sum': target_total_percent,
        }
        target_total_percent_val.update(_group_ou_dict)
        calculated_vals.append(target_total_percent_val)

        if all_not_none(target_device['numeric_sum'], method_device['numeric_sum']) and target_device['numeric_sum']:
            target_device_percent = (method_device['numeric_sum'] * 100) / target_device['numeric_sum']
        else:
            target_device_percent = None
        target_device_percent_val = {
            'de_name': 'Perf% Circumcised DC',
            'cat_combo': None,
            'numeric_sum': target_device_percent,
        }
        target_device_percent_val.update(_group_ou_dict)
        calculated_vals.append(target_device_percent_val)

        if all_not_none(target_surgical['numeric_sum'], method_surgical['numeric_sum']) and target_surgical['numeric_sum']:
            target_surgical_percent = (method_surgical['numeric_sum'] * 100) / target_surgical['numeric_sum']
        else:
            target_surgical_percent = None
        target_surgical_percent_val = {
            'de_name': 'Perf% Circumcised Surgical',
            'cat_combo': None,
            'numeric_sum': target_surgical_percent,
        }
        target_surgical_percent_val.update(_group_ou_dict)
        calculated_vals.append(target_surgical_percent_val)

        if all_not_none(followup_48hrs['numeric_sum'], method_sum) and method_sum:
            followup_48hrs_percent = (followup_48hrs['numeric_sum'] * 100) / method_sum
        else:
            followup_48hrs_percent = None
        followup_48hrs_percent_val = {
            'de_name': '% who returned within 48 hours',
            'cat_combo': None,
            'numeric_sum': followup_48hrs_percent,
        }
        followup_48hrs_percent_val.update(_group_ou_dict)
        calculated_vals.append(followup_48hrs_percent_val)

        adverse_sum = default_zero(adverse_moderate['numeric_sum']) + default_zero(adverse_severe['numeric_sum'])

        if all_not_none(adverse_sum, method_sum) and method_sum:
            adverse_percent = (adverse_sum * 100) / method_sum
        else:
            adverse_percent = None
        adverse_percent_val = {
            'de_name': '% with at least one adverse event',
            'cat_combo': None,
            'numeric_sum': adverse_percent,
        }
        adverse_percent_val.update(_group_ou_dict)
        calculated_vals.append(adverse_percent_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Perf% Circumcised'], (None,)))
    data_element_metas += list(product(['Perf% Circumcised DC'], (None,)))
    data_element_metas += list(product(['Perf% Circumcised Surgical'], (None,)))
    data_element_metas += list(product(['% who returned within 48 hours'], (None,)))
    data_element_metas += list(product(['% with at least one adverse event'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    vmmc_ls = LegendSet()
    vmmc_ls.name = 'Perf Circumcised'
    vmmc_ls.add_interval('orange', 0, 25)
    vmmc_ls.add_interval('yellow', 25, 40)
    vmmc_ls.add_interval('light-green', 50, 60)
    vmmc_ls.add_interval('green', 60, None)
    for i in range(num_path_elements+15, num_path_elements+15+3):
        vmmc_ls.mappings[i] = True
    legend_sets.append(vmmc_ls)
    adverse_ls = LegendSet()
    adverse_ls.name = 'Adverse Events'
    adverse_ls.add_interval('red', 0.5, None)
    adverse_ls.mappings[num_path_elements+19] = True
    legend_sets.append(adverse_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="vmmc_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="vmmc_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/vmmc_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

def lab_dashboard(request):
    this_day = date.today()
    this_quarter = '%d-Q%d' % (this_day.year, month2quarter(this_day.month))
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_day.year, this_day.year-6, -1) for q in range(4, 0, -1)]
    period_list = list(filter(lambda qtr: qtr < this_quarter, reversed(PREV_5YR_QTRS)))[-6:]
    def val_with_period_de_fun(row, col):
        period = row
        de_name = col
        return { 'de_name': de_name, 'period': period, 'numeric_sum': None }

    data_element_metas = list()

    viral_target_de_names = (
        'VL_TARGET',
    )
    viral_target_short_names = (
        'Samples target',
    )
    de_viral_target_meta = list(product(viral_target_short_names, (None,)))

    qs_viral_target = DataValue.objects.what(*viral_target_de_names)
    qs_viral_target = qs_viral_target.annotate(cat_combo=Value(None, output_field=CharField()))
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    year_list = sorted(list(set(qstr[:4] for qstr in period_list)))
    qs_viral_target = qs_viral_target.when(*year_list)
    qs_viral_target = qs_viral_target.order_by('period', 'de_name')
    val_viral_target = qs_viral_target.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_viral_target = list(val_viral_target)
    q_list = sorted(['%s-Q%d' % (y, qnum,) for qnum in range(1, 5) for y in year_list])
    q_list = [x for x in q_list if x in period_list]

    def duplicate_values_over_periods(val_list, p_list):
        def assign_period(v, p):
            new_val = dict(v)
            new_val['period'] = p
            return new_val

        return [assign_period(val, period) for val in val_list for period in p_list]

    val_viral_target = duplicate_values_over_periods(val_viral_target, q_list)

    gen_raster = grabbag.rasterize(period_list, viral_target_de_names, val_viral_target, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_viral_target2 = list(gen_raster)

    viral_load_de_names = (
        'VL samples rejected',
        'VL samples sent',
    )
    viral_load_short_names = (
        'VL samples rejected',
        'VL samples sent',
    )
    de_viral_load_meta = list(product(viral_load_short_names, (None,)))

    qs_viral_load = DataValue.objects.what(*viral_load_de_names)
    qs_viral_load = qs_viral_load.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_viral_load = qs_viral_load.when(*period_list)
    qs_viral_load = qs_viral_load.order_by('period', 'de_name')
    val_viral_load = qs_viral_load.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_viral_load = list(val_viral_load)

    gen_raster = grabbag.rasterize(period_list, viral_load_de_names, val_viral_load, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_viral_load2 = list(gen_raster)

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_viral_target2, val_viral_load2), key=lambda x: (x['period'])), key=lambda x: (x['period']))
    # if True:
    #     grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (period, (vl_target, vl_rejected, vl_sent, *other_vals)) = _group
        
        calculated_vals = list()

        if all_not_none(vl_sent['numeric_sum'], vl_target['numeric_sum']) and vl_target['numeric_sum']:
            coverage_percent = (vl_sent['numeric_sum'] * 100) / vl_target['numeric_sum']
        else:
            coverage_percent = None
        coverage_percent_val = {
            'period': period,
            'de_name': '% VL testing coverage',
            'numeric_sum': coverage_percent,
        }
        calculated_vals.append(coverage_percent_val)

        if all_not_none(vl_rejected['numeric_sum'], vl_sent['numeric_sum']) and vl_sent['numeric_sum']:
            reject_percent = (vl_rejected['numeric_sum'] * 100) / vl_sent['numeric_sum']
        else:
            reject_percent = None
        reject_percent_val = {
            'period': period,
            'de_name': '% VL sample rejection',
            'numeric_sum': reject_percent,
        }
        calculated_vals.append(reject_percent_val)

        _group[1] = calculated_vals
    
    context = {
        'data_element_names': [
            ('% VL testing coverage', None),
            ('% VL sample rejection', None),
        ],
        'grouped_data': grouped_vals,
    }
    return render(request, 'cannula/index.html', context)

@login_required
def lab_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    malaria_de_names = (
        '105-7.3 Lab Malaria Microscopy  Number Done',
        '105-7.3 Lab Malaria RDTs Number Done',
    )
    malaria_short_names = (
        'Malaria Microscopy Done',
        'Malaria RDTs Done',
    )
    de_malaria_meta = list(product(malaria_de_names, (None,)))
    data_element_metas += list(product(malaria_short_names, (None,)))

    qs_malaria = DataValue.objects.what(*malaria_de_names)
    qs_malaria = qs_malaria.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_malaria = qs_malaria.where(filter_district)
    qs_malaria = qs_malaria.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_malaria = qs_malaria.when(filter_period)
    qs_malaria = qs_malaria.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_malaria = qs_malaria.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_malaria = list(val_malaria)

    gen_raster = grabbag.rasterize(ou_list, de_malaria_meta, val_malaria, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_malaria2 = list(gen_raster)

    hiv_determine_de_names = (
        '105-7.8 Lab Determine Clinical Diagnosis',
        '105-7.8 Lab Determine HCT',
        '105-7.8 Lab Determine PMTCT',
        '105-7.8 Lab Determine Quality Control',
        '105-7.8 Lab Determine SMC',
    )
    hiv_determine_short_names = (
        'HIV tests done using Determine',
    )
    de_hiv_determine_meta = list(product(hiv_determine_short_names, (None,)))
    data_element_metas += de_hiv_determine_meta

    qs_hiv_determine = DataValue.objects.what(*hiv_determine_de_names)
    qs_hiv_determine = qs_hiv_determine.annotate(de_name=Value(hiv_determine_short_names[0], output_field=CharField()))
    qs_hiv_determine = qs_hiv_determine.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_determine = qs_hiv_determine.where(filter_district)
    qs_hiv_determine = qs_hiv_determine.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_determine = qs_hiv_determine.when(filter_period)
    qs_hiv_determine = qs_hiv_determine.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_determine = qs_hiv_determine.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_determine = list(val_hiv_determine)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_determine_meta, val_hiv_determine, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_determine2 = list(gen_raster)

    hiv_statpak_de_names = (
        '105-7.8 Lab Stat pak  Clinical Diagnosis',
        '105-7.8 Lab Stat pak  HCT',
        '105-7.8 Lab Stat pak  PMTCT',
        '105-7.8 Lab Stat pak  Quality Control',
        '105-7.8 Lab Stat pak  SMC',
    )
    hiv_statpak_short_names = (
        'HIV tests done using Stat Pak',
    )
    de_hiv_statpak_meta = list(product(hiv_statpak_short_names, (None,)))
    data_element_metas += de_hiv_statpak_meta

    qs_hiv_statpak = DataValue.objects.what(*hiv_statpak_de_names)
    qs_hiv_statpak = qs_hiv_statpak.annotate(de_name=Value(hiv_statpak_short_names[0], output_field=CharField()))
    qs_hiv_statpak = qs_hiv_statpak.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_statpak = qs_hiv_statpak.where(filter_district)
    qs_hiv_statpak = qs_hiv_statpak.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_statpak = qs_hiv_statpak.when(filter_period)
    qs_hiv_statpak = qs_hiv_statpak.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_statpak = qs_hiv_statpak.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_statpak = list(val_hiv_statpak)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_statpak_meta, val_hiv_statpak, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_statpak2 = list(gen_raster)

    hiv_unigold_de_names = (
        '105-7.8 Lab Unigold Clinical Diagnosis',
        '105-7.8 Lab Unigold HCT',
        '105-7.8 Lab Unigold PMTCT',
        '105-7.8 Lab Unigold Quality Control',
        '105-7.8 Lab Unigold SMC',
    )
    hiv_unigold_short_names = (
        'HIV tests done using Unigold',
    )
    de_hiv_unigold_meta = list(product(hiv_unigold_short_names, (None,)))
    data_element_metas += de_hiv_unigold_meta

    qs_hiv_unigold = DataValue.objects.what(*hiv_unigold_de_names)
    qs_hiv_unigold = qs_hiv_unigold.annotate(de_name=Value(hiv_unigold_short_names[0], output_field=CharField()))
    qs_hiv_unigold = qs_hiv_unigold.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_unigold = qs_hiv_unigold.where(filter_district)
    qs_hiv_unigold = qs_hiv_unigold.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_unigold = qs_hiv_unigold.when(filter_period)
    qs_hiv_unigold = qs_hiv_unigold.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_unigold = qs_hiv_unigold.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_unigold = list(val_hiv_unigold)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_unigold_meta, val_hiv_unigold, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_unigold2 = list(gen_raster)

    tb_smear_de_names = (
        '105-7.6 Lab ZN for AFBs  Number Done',
    )
    tb_smear_short_names = (
        'TB Smear',
    )
    de_tb_smear_meta = list(product(tb_smear_de_names, (None,)))
    data_element_metas += list(product(tb_smear_short_names, (None,)))

    qs_tb_smear = DataValue.objects.what(*tb_smear_de_names)
    qs_tb_smear = qs_tb_smear.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_tb_smear = qs_tb_smear.where(filter_district)
    qs_tb_smear = qs_tb_smear.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_tb_smear = qs_tb_smear.when(filter_period)
    qs_tb_smear = qs_tb_smear.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_tb_smear = qs_tb_smear.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_tb_smear = list(val_tb_smear)

    gen_raster = grabbag.rasterize(ou_list, de_tb_smear_meta, val_tb_smear, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_tb_smear2 = list(gen_raster)

    syphilis_de_names = (
        '105-7.4 Lab VDRL/RPR Number Done',
        '105-7.4 Lab TPHA  Number Done',
    )
    syphilis_short_names = (
        'Syphilis tests',
    )
    de_syphilis_meta = list(product(syphilis_short_names, (None,)))
    data_element_metas += de_syphilis_meta

    qs_syphilis = DataValue.objects.what(*syphilis_de_names)
    qs_syphilis = qs_syphilis.annotate(de_name=Value(syphilis_short_names[0], output_field=CharField()))
    qs_syphilis = qs_syphilis.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_syphilis = qs_syphilis.where(filter_district)
    qs_syphilis = qs_syphilis.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_syphilis = qs_syphilis.when(filter_period)
    qs_syphilis = qs_syphilis.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_syphilis = qs_syphilis.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_syphilis = list(val_syphilis)

    gen_raster = grabbag.rasterize(ou_list, de_syphilis_meta, val_syphilis, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_syphilis2 = list(gen_raster)

    liver_de_names = (
        '105-7.7 Lab ALT Number Done',
        '105-7.7 Lab AST Number Done',
        '105-7.7 Lab Albumin  Number Done',
    )
    liver_short_names = (
        'LFTs',
    )
    de_liver_meta = list(product(liver_short_names, (None,)))
    data_element_metas += de_liver_meta

    qs_liver = DataValue.objects.what(*liver_de_names)
    qs_liver = qs_liver.annotate(de_name=Value(liver_short_names[0], output_field=CharField()))
    qs_liver = qs_liver.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_liver = qs_liver.where(filter_district)
    qs_liver = qs_liver.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_liver = qs_liver.when(filter_period)
    qs_liver = qs_liver.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_liver = qs_liver.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_liver = list(val_liver)

    gen_raster = grabbag.rasterize(ou_list, de_liver_meta, val_liver, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_liver2 = list(gen_raster)

    renal_de_names = (
        '105-7.7 Lab Calcium  Number Done',
        '105-7.7 Lab Creatinine Number Done',
        '105-7.7 Lab Potassium Number Done',
        '105-7.7 Lab Sodium Number Done',
        '105-7.7 Lab Total Protein Number Done',
        '105-7.7 Lab Urea Number Done',
    )
    renal_short_names = (
        'RFTs',
    )
    de_renal_meta = list(product(renal_short_names, (None,)))
    data_element_metas += de_renal_meta

    qs_renal = DataValue.objects.what(*renal_de_names)
    qs_renal = qs_renal.annotate(de_name=Value(renal_short_names[0], output_field=CharField()))
    qs_renal = qs_renal.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_renal = qs_renal.where(filter_district)
    qs_renal = qs_renal.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_renal = qs_renal.when(filter_period)
    qs_renal = qs_renal.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_renal = qs_renal.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_renal = list(val_renal)

    gen_raster = grabbag.rasterize(ou_list, de_renal_meta, val_renal, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_renal2 = list(gen_raster)

    other_haem_de_names = (
        'All Other Haematology - Lab - OPD  Number Done',
    )
    other_haem_short_names = (
        'All other Haematology',
    )
    de_other_haem_meta = list(product(other_haem_de_names, (None,)))
    data_element_metas += list(product(other_haem_short_names, (None,)))

    qs_other_haem = DataValue.objects.what(*other_haem_de_names)
    qs_other_haem = qs_other_haem.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_other_haem = qs_other_haem.where(filter_district)
    qs_other_haem = qs_other_haem.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_other_haem = qs_other_haem.when(filter_period)
    qs_other_haem = qs_other_haem.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_other_haem = qs_other_haem.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_other_haem = list(val_other_haem)

    gen_raster = grabbag.rasterize(ou_list, de_other_haem_meta, val_other_haem, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_other_haem2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_malaria2, val_hiv_determine2, val_hiv_statpak2, val_hiv_unigold2, val_tb_smear2, val_syphilis2, val_liver2,val_renal2, val_other_haem2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (malaria_microscopy, malaria_rdt, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        malaria_sum = default_zero(malaria_microscopy['numeric_sum']) + default_zero(malaria_rdt['numeric_sum'])
        malaria_val = {
            'de_name': 'Malaria (Smear & RDTs)',
            'cat_combo': None,
            'numeric_sum': malaria_sum,
        }
        malaria_val.update(_group_ou_dict)
        calculated_vals.append(malaria_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Malaria (Smear & RDTs)'], (None,)))

    legend_sets = list()
    # lab_ls = LegendSet()
    # lab_ls.add_interval('orange', 0, 25)
    # lab_ls.add_interval('yellow', 25, 40)
    # lab_ls.add_interval('light-green', 50, 60)
    # lab_ls.add_interval('green', 60, None)
    # legend_sets.append(lab_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="lab_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="lab_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/lab_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def fp_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)
    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))

    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    condoms_new_de_names = (
        '105-2.5 Female Condom',
        '105-2.5 Male Condom',
    )
    condoms_new_short_names = (
        'New users - Condoms',
    )
    de_condoms_new_meta = list(product(condoms_new_short_names, (None,)))
    data_element_metas += de_condoms_new_meta

    qs_condoms_new = DataValue.objects.what(*condoms_new_de_names)
    qs_condoms_new = qs_condoms_new.annotate(de_name=Value(condoms_new_short_names[0], output_field=CharField()))
    qs_condoms_new = qs_condoms_new.filter(category_combo__categories__name='New Users')
    qs_condoms_new = qs_condoms_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_condoms_new = qs_condoms_new.where(filter_district)
    qs_condoms_new = qs_condoms_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_condoms_new = qs_condoms_new.when(filter_period)
    qs_condoms_new = qs_condoms_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_condoms_new = qs_condoms_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_condoms_new = list(val_condoms_new)

    gen_raster = grabbag.rasterize(ou_list, de_condoms_new_meta, val_condoms_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_condoms_new2 = list(gen_raster)

    fp_new_de_names = (
        '105-2.5 Injectable',
        '105-2.5 IUDs',
        '105-2.5 Natural',
        '105-2.7 Implant',
    )
    fp_new_short_names = (
        'New users - Injectables',
        'New users - IUDs',
        'New users - Natural methods',
        'New users - Implants',
    )
    de_fp_new_meta = list(product(fp_new_de_names, (None,)))
    data_element_metas += list(product(fp_new_short_names, (None,)))

    qs_fp_new = DataValue.objects.what(*fp_new_de_names)
    qs_fp_new = qs_fp_new.filter(category_combo__categories__name='New Users')
    qs_fp_new = qs_fp_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_fp_new = qs_fp_new.where(filter_district)
    qs_fp_new = qs_fp_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_fp_new = qs_fp_new.when(filter_period)
    qs_fp_new = qs_fp_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_fp_new = qs_fp_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_fp_new = list(val_fp_new)

    gen_raster = grabbag.rasterize(ou_list, de_fp_new_meta, val_fp_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_fp_new2 = list(gen_raster)

    oral_new_de_names = (
        '105-2.5 Oral: Microgynon',
        '105-2.5 Oral: Lo-Feminal',
        '105-2.5 Oral : Ovrette or Another POP',
    )
    oral_new_short_names = (
        'New users - Oral',
    )
    de_oral_new_meta = list(product(oral_new_short_names, (None,)))
    data_element_metas += de_oral_new_meta

    qs_oral_new = DataValue.objects.what(*oral_new_de_names)
    qs_oral_new = qs_oral_new.annotate(de_name=Value(oral_new_short_names[0], output_field=CharField()))
    qs_oral_new = qs_oral_new.filter(category_combo__categories__name='New Users')
    qs_oral_new = qs_oral_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_oral_new = qs_oral_new.where(filter_district)
    qs_oral_new = qs_oral_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_oral_new = qs_oral_new.when(filter_period)
    qs_oral_new = qs_oral_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_oral_new = qs_oral_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_oral_new = list(val_oral_new)

    gen_raster = grabbag.rasterize(ou_list, de_oral_new_meta, val_oral_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_oral_new2 = list(gen_raster)

    other_new_de_names = (
        '105-2.5 Other Method',
    )
    other_new_short_names = (
        'New users - Other methods',
    )
    de_other_new_meta = list(product(other_new_short_names, (None,)))
    data_element_metas += de_other_new_meta

    qs_other_new = DataValue.objects.what(*other_new_de_names)
    qs_other_new = qs_other_new.annotate(de_name=Value(other_new_short_names[0], output_field=CharField()))
    qs_other_new = qs_other_new.filter(category_combo__categories__name='New Users')
    qs_other_new = qs_other_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_other_new = qs_other_new.where(filter_district)
    qs_other_new = qs_other_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_other_new = qs_other_new.when(filter_period)
    qs_other_new = qs_other_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_other_new = qs_other_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_other_new = list(val_other_new)

    gen_raster = grabbag.rasterize(ou_list, de_other_new_meta, val_other_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_other_new2 = list(gen_raster)

    sterile_new_de_names = (
        '105-2.7 Female Sterilisation (TubeLigation)',
        '105-2.7 Male Sterilisation (Vasectomy)',
    )
    sterile_new_short_names = (
        'New users - Sterilisation (male and female)',
    )
    de_sterile_new_meta = list(product(sterile_new_short_names, (None,)))
    data_element_metas += de_sterile_new_meta

    qs_sterile_new = DataValue.objects.what(*sterile_new_de_names)
    qs_sterile_new = qs_sterile_new.annotate(de_name=Value(sterile_new_short_names[0], output_field=CharField()))
    qs_sterile_new = qs_sterile_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_sterile_new = qs_sterile_new.where(filter_district)
    qs_sterile_new = qs_sterile_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_sterile_new = qs_sterile_new.when(filter_period)
    qs_sterile_new = qs_sterile_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_sterile_new = qs_sterile_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sterile_new = list(val_sterile_new)

    gen_raster = grabbag.rasterize(ou_list, de_sterile_new_meta, val_sterile_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_sterile_new2 = list(gen_raster)

    condoms_revisit_de_names = (
        '105-2.5 Female Condom',
        '105-2.5 Male Condom',
    )
    condoms_revisit_short_names = (
        'Revisits - Condoms',
    )
    de_condoms_revisit_meta = list(product(condoms_revisit_short_names, (None,)))
    data_element_metas += de_condoms_revisit_meta

    qs_condoms_revisit = DataValue.objects.what(*condoms_revisit_de_names)
    qs_condoms_revisit = qs_condoms_revisit.annotate(de_name=Value(condoms_revisit_short_names[0], output_field=CharField()))
    qs_condoms_revisit = qs_condoms_revisit.filter(category_combo__categories__name='Revisits')
    qs_condoms_revisit = qs_condoms_revisit.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_condoms_revisit = qs_condoms_revisit.where(filter_district)
    qs_condoms_revisit = qs_condoms_revisit.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_condoms_revisit = qs_condoms_revisit.when(filter_period)
    qs_condoms_revisit = qs_condoms_revisit.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_condoms_revisit = qs_condoms_revisit.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_condoms_revisit = list(val_condoms_revisit)

    gen_raster = grabbag.rasterize(ou_list, de_condoms_revisit_meta, val_condoms_revisit, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_condoms_revisit2 = list(gen_raster)

    fp_revisit_de_names = (
        '105-2.5 Injectable',
        '105-2.5 IUDs',
        '105-2.5 Natural',
        '105-2.7 Implant',
    )
    fp_revisit_short_names = (
        'Revisits - Injectables',
        'Revisits - IUDs',
        'Revisits - Natural methods',
        'Revisits - Implants',
    )
    de_fp_revisit_meta = list(product(fp_revisit_de_names, (None,)))
    data_element_metas += list(product(fp_revisit_short_names, (None,)))

    qs_fp_revisit = DataValue.objects.what(*fp_revisit_de_names)
    qs_fp_revisit = qs_fp_revisit.filter(category_combo__categories__name='Revisits')
    qs_fp_revisit = qs_fp_revisit.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_fp_revisit = qs_fp_revisit.where(filter_district)
    qs_fp_revisit = qs_fp_revisit.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_fp_revisit = qs_fp_revisit.when(filter_period)
    qs_fp_revisit = qs_fp_revisit.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_fp_revisit = qs_fp_revisit.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_fp_revisit = list(val_fp_revisit)

    gen_raster = grabbag.rasterize(ou_list, de_fp_revisit_meta, val_fp_revisit, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_fp_revisit2 = list(gen_raster)

    oral_revisit_de_names = (
        '105-2.5 Oral: Microgynon',
        '105-2.5 Oral: Lo-Feminal',
        '105-2.5 Oral : Ovrette or Another POP',
    )
    oral_revisit_short_names = (
        'Revisits - Oral',
    )
    de_oral_revisit_meta = list(product(oral_revisit_short_names, (None,)))
    data_element_metas += de_oral_revisit_meta

    qs_oral_revisit = DataValue.objects.what(*oral_revisit_de_names)
    qs_oral_revisit = qs_oral_revisit.annotate(de_name=Value(oral_revisit_short_names[0], output_field=CharField()))
    qs_oral_revisit = qs_oral_revisit.filter(category_combo__categories__name='Revisits')
    qs_oral_revisit = qs_oral_revisit.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_oral_revisit = qs_oral_revisit.where(filter_district)
    qs_oral_revisit = qs_oral_revisit.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_oral_revisit = qs_oral_revisit.when(filter_period)
    qs_oral_revisit = qs_oral_revisit.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_oral_revisit = qs_oral_revisit.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_oral_revisit = list(val_oral_revisit)

    gen_raster = grabbag.rasterize(ou_list, de_oral_revisit_meta, val_oral_revisit, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_oral_revisit2 = list(gen_raster)

    other_revisit_de_names = (
        '105-2.5 Other Method',
    )
    other_revisit_short_names = (
        'Revisits - Other methods',
    )
    de_other_revisit_meta = list(product(other_revisit_short_names, (None,)))
    data_element_metas += de_other_revisit_meta

    qs_other_revisit = DataValue.objects.what(*other_revisit_de_names)
    qs_other_revisit = qs_other_revisit.annotate(de_name=Value(other_revisit_short_names[0], output_field=CharField()))
    qs_other_revisit = qs_other_revisit.filter(category_combo__categories__name='Revisits')
    qs_other_revisit = qs_other_revisit.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_other_revisit = qs_other_revisit.where(filter_district)
    qs_other_revisit = qs_other_revisit.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_other_revisit = qs_other_revisit.when(filter_period)
    qs_other_revisit = qs_other_revisit.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_other_revisit = qs_other_revisit.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_other_revisit = list(val_other_revisit)

    gen_raster = grabbag.rasterize(ou_list, de_other_revisit_meta, val_other_revisit, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_other_revisit2 = list(gen_raster)

    hiv_new_de_names = (
        '105-2.5 Number HIV+ FP users',
    )
    hiv_new_short_names = (
        'New users - HIV+',
    )
    de_hiv_new_meta = list(product(hiv_new_short_names, (None,)))
    data_element_metas += de_hiv_new_meta

    qs_hiv_new = DataValue.objects.what(*hiv_new_de_names)
    qs_hiv_new = qs_hiv_new.annotate(de_name=Value(hiv_new_short_names[0], output_field=CharField()))
    qs_hiv_new = qs_hiv_new.filter(category_combo__categories__name='New Users')
    qs_hiv_new = qs_hiv_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_new = qs_hiv_new.where(filter_district)
    qs_hiv_new = qs_hiv_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_new = qs_hiv_new.when(filter_period)
    qs_hiv_new = qs_hiv_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_new = qs_hiv_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_new = list(val_hiv_new)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_new_meta, val_hiv_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_new2 = list(gen_raster)

    hiv_revisit_de_names = (
        '105-2.5 Number HIV+ FP users',
    )
    hiv_revisit_short_names = (
        'Revisits - HIV+',
    )
    de_hiv_revisit_meta = list(product(hiv_revisit_short_names, (None,)))
    data_element_metas += de_hiv_revisit_meta

    qs_hiv_revisit = DataValue.objects.what(*hiv_revisit_de_names)
    qs_hiv_revisit = qs_hiv_revisit.annotate(de_name=Value(hiv_revisit_short_names[0], output_field=CharField()))
    qs_hiv_revisit = qs_hiv_revisit.filter(category_combo__categories__name='Revisits')
    qs_hiv_revisit = qs_hiv_revisit.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_revisit = qs_hiv_revisit.where(filter_district)
    qs_hiv_revisit = qs_hiv_revisit.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_revisit = qs_hiv_revisit.when(filter_period)
    qs_hiv_revisit = qs_hiv_revisit.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_revisit = qs_hiv_revisit.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_revisit = list(val_hiv_revisit)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_revisit_meta, val_hiv_revisit, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_revisit2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_condoms_new2, val_fp_new2, val_oral_new2, val_other_new2, val_sterile_new2, val_condoms_revisit2, val_fp_revisit2, val_oral_revisit2, val_other_revisit2, val_hiv_new2, val_hiv_revisit2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (condom_new, inject_new, iud_new, natural_new, implant_new, oral_new, other_new, sterile_new, condom_revisit, inject_revisit, iud_revisit, natural_revisit, implant_revisit, oral_revisit, other_revisit, hiv_new, hiv_revisit, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        total_new_sum = default_zero(condom_new['numeric_sum']) + default_zero(inject_new['numeric_sum']) + default_zero(iud_new['numeric_sum']) + default_zero(natural_new['numeric_sum']) + default_zero(implant_new['numeric_sum']) + default_zero(oral_new['numeric_sum']) + default_zero(other_new['numeric_sum']) + default_zero(sterile_new['numeric_sum'])
        total_new_val = {
            'de_name': 'New Users - TOTAL',
            'cat_combo': None,
            'numeric_sum': total_new_sum,
        }
        total_new_val.update(_group_ou_dict)
        calculated_vals.append(total_new_val)

        total_revisit_sum = default_zero(condom_revisit['numeric_sum']) + default_zero(inject_revisit['numeric_sum']) + default_zero(iud_revisit['numeric_sum']) + default_zero(natural_revisit['numeric_sum']) + default_zero(implant_revisit['numeric_sum']) + default_zero(oral_revisit['numeric_sum']) + default_zero(other_revisit['numeric_sum'])
        total_revisit_val = {
            'de_name': 'Revisits - TOTAL',
            'cat_combo': None,
            'numeric_sum': total_revisit_sum,
        }
        total_revisit_val.update(_group_ou_dict)
        calculated_vals.append(total_revisit_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['New Users - TOTAL'], (None,)))
    data_element_metas += list(product(['Revisits - TOTAL'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    # fp_ls = LegendSet()
    # fp_ls.add_interval('orange', 0, 25)
    # fp_ls.add_interval('yellow', 25, 40)
    # fp_ls.add_interval('light-green', 50, 60)
    # fp_ls.add_interval('green', 60, None)
    # legend_sets.append(fp_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="family_planning_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="family_planning_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/fp_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def fp_cyp_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    oral_de_names = (
        '105-2.5 Oral: Microgynon',
        '105-2.5 Oral: Lo-Feminal',
        '105-2.5 Oral : Ovrette or Another POP',
    )
    oral_short_names = (
        'Oral dispensed (cycles)',
    )
    de_oral_meta = list(product(oral_short_names, (None,)))
    data_element_metas += de_oral_meta

    qs_oral = DataValue.objects.what(*oral_de_names)
    qs_oral = qs_oral.annotate(de_name=Value(oral_short_names[0], output_field=CharField()))
    qs_oral = qs_oral.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_oral = qs_oral.where(filter_district)
    qs_oral = qs_oral.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_oral = qs_oral.when(filter_period)
    qs_oral = qs_oral.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_oral = qs_oral.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_oral = list(val_oral)

    gen_raster = grabbag.rasterize(ou_list, de_oral_meta, val_oral, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_oral2 = list(gen_raster)

    condoms_de_names = (
        '105-2.5 Female Condom',
        '105-2.5 Male Condom',
    )
    condoms_short_names = (
        'Condoms dispensed (pieces)',
    )
    de_condoms_meta = list(product(condoms_short_names, (None,)))
    data_element_metas += de_condoms_meta

    qs_condoms = DataValue.objects.what(*condoms_de_names)
    qs_condoms = qs_condoms.annotate(de_name=Value(condoms_short_names[0], output_field=CharField()))
    qs_condoms = qs_condoms.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_condoms = qs_condoms.where(filter_district)
    qs_condoms = qs_condoms.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_condoms = qs_condoms.when(filter_period)
    qs_condoms = qs_condoms.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_condoms = qs_condoms.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_condoms = list(val_condoms)

    gen_raster = grabbag.rasterize(ou_list, de_condoms_meta, val_condoms, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_condoms2 = list(gen_raster)

    implants_new_de_names = (
        '105-2.7 Implant',
    )
    implants_new_short_names = (
        'New users - Implants',
    )
    de_implants_new_meta = list(product(implants_new_short_names, (None,)))
    data_element_metas += de_implants_new_meta

    qs_implants_new = DataValue.objects.what(*implants_new_de_names)
    qs_implants_new = qs_implants_new.annotate(de_name=Value(implants_new_short_names[0], output_field=CharField()))
    qs_implants_new = qs_implants_new.filter(category_combo__categories__name='New Users')
    qs_implants_new = qs_implants_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_implants_new = qs_implants_new.where(filter_district)
    qs_implants_new = qs_implants_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_implants_new = qs_implants_new.when(filter_period)
    qs_implants_new = qs_implants_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_implants_new = qs_implants_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_implants_new = list(val_implants_new)

    gen_raster = grabbag.rasterize(ou_list, de_implants_new_meta, val_implants_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_implants_new2 = list(gen_raster)

    injectable_de_names = (
        '105-2.5 Injectable',
    )
    injectable_short_names = (
        'Injectable dispensed (doses)',
    )
    de_injectable_meta = list(product(injectable_short_names, (None,)))
    data_element_metas += de_injectable_meta

    qs_injectable = DataValue.objects.what(*injectable_de_names)
    qs_injectable = qs_injectable.annotate(de_name=Value(injectable_short_names[0], output_field=CharField()))
    qs_injectable = qs_injectable.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_injectable = qs_injectable.where(filter_district)
    qs_injectable = qs_injectable.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_injectable = qs_injectable.when(filter_period)
    qs_injectable = qs_injectable.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_injectable = qs_injectable.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_injectable = list(val_injectable)

    gen_raster = grabbag.rasterize(ou_list, de_injectable_meta, val_injectable, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_injectable2 = list(gen_raster)

    iud_de_names = (
        '105-2.5 IUDs',
    )
    iud_short_names = (
        'IUD inserted',
    )
    de_iud_meta = list(product(iud_short_names, (None,)))
    data_element_metas += de_iud_meta

    qs_iud = DataValue.objects.what(*iud_de_names)
    qs_iud = qs_iud.annotate(de_name=Value(iud_short_names[0], output_field=CharField()))
    qs_iud = qs_iud.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_iud = qs_iud.where(filter_district)
    qs_iud = qs_iud.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_iud = qs_iud.when(filter_period)
    qs_iud = qs_iud.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_iud = qs_iud.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_iud = list(val_iud)

    gen_raster = grabbag.rasterize(ou_list, de_iud_meta, val_iud, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_iud2 = list(gen_raster)

    sterile_new_de_names = (
        '105-2.7 Female Sterilisation (TubeLigation)',
        '105-2.7 Male Sterilisation (Vasectomy)',
    )
    sterile_new_short_names = (
        'New users - Sterilisation (male and female)',
    )
    de_sterile_new_meta = list(product(sterile_new_short_names, (None,)))
    data_element_metas += de_sterile_new_meta

    qs_sterile_new = DataValue.objects.what(*sterile_new_de_names)
    qs_sterile_new = qs_sterile_new.annotate(de_name=Value(sterile_new_short_names[0], output_field=CharField()))
    qs_sterile_new = qs_sterile_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_sterile_new = qs_sterile_new.where(filter_district)
    qs_sterile_new = qs_sterile_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_sterile_new = qs_sterile_new.when(filter_period)
    qs_sterile_new = qs_sterile_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_sterile_new = qs_sterile_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sterile_new = list(val_sterile_new)

    gen_raster = grabbag.rasterize(ou_list, de_sterile_new_meta, val_sterile_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_sterile_new2 = list(gen_raster)

    natural_de_names = (
        '105-2.5 Natural',
    )
    natural_short_names = (
        'Natural methods',
    )
    de_natural_meta = list(product(natural_short_names, (None,)))
    data_element_metas += de_natural_meta

    qs_natural = DataValue.objects.what(*natural_de_names)
    qs_natural = qs_natural.annotate(de_name=Value(natural_short_names[0], output_field=CharField()))
    qs_natural = qs_natural.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_natural = qs_natural.where(filter_district)
    qs_natural = qs_natural.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_natural = qs_natural.when(filter_period)
    qs_natural = qs_natural.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_natural = qs_natural.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_natural = list(val_natural)

    gen_raster = grabbag.rasterize(ou_list, de_natural_meta, val_natural, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_natural2 = list(gen_raster)

    emergency_de_names = (
        '105-2.6 Emergency contraceptives  No. Dispensed by CBD',
        '105-2.6 Emergency contraceptives  No. Dispensed at Unit',
        '105-2.6 Emergency contraceptives  No. Disp. At Outreach',
    )
    emergency_short_names = (
        'Emergency contraceptives dispensed (doses)',
    )
    de_emergency_meta = list(product(emergency_short_names, (None,)))
    data_element_metas += de_emergency_meta

    qs_emergency = DataValue.objects.what(*emergency_de_names)
    qs_emergency = qs_emergency.annotate(de_name=Value(emergency_short_names[0], output_field=CharField()))
    qs_emergency = qs_emergency.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_emergency = qs_emergency.where(filter_district)
    qs_emergency = qs_emergency.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_emergency = qs_emergency.when(filter_period)
    qs_emergency = qs_emergency.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_emergency = qs_emergency.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_emergency = list(val_emergency)

    gen_raster = grabbag.rasterize(ou_list, de_emergency_meta, val_emergency, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_emergency2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_oral2, val_condoms2, val_implants_new2, val_injectable2, val_iud2, val_sterile_new2, val_natural2, val_emergency2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (oral, condoms, implants_new, injectable, iud, sterile_new, natural, emergency, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(oral['numeric_sum']):
            cyp_oral = oral['numeric_sum'] / 15
        else:
            cyp_oral = None
        cyp_oral_val = {
            'de_name': 'CYPs Oral',
            'cat_combo': None,
            'numeric_sum': cyp_oral,
        }
        cyp_oral_val.update(_group_ou_dict)
        calculated_vals.append(cyp_oral_val)

        if all_not_none(condoms['numeric_sum']):
            cyp_condoms = condoms['numeric_sum'] / 120
        else:
            cyp_condoms = None
        cyp_condoms_val = {
            'de_name': 'CYPs Condoms',
            'cat_combo': None,
            'numeric_sum': cyp_condoms,
        }
        cyp_condoms_val.update(_group_ou_dict)
        calculated_vals.append(cyp_condoms_val)

        if all_not_none(implants_new['numeric_sum']):
            cyp_implants = implants_new['numeric_sum'] * Decimal(2.5)
        else:
            cyp_implants = None
        cyp_implants_val = {
            'de_name': 'CYPs Implants',
            'cat_combo': None,
            'numeric_sum': cyp_implants,
        }
        cyp_implants_val.update(_group_ou_dict)
        calculated_vals.append(cyp_implants_val)

        if all_not_none(injectable['numeric_sum']):
            cyp_injectable = injectable['numeric_sum'] / 4
        else:
            cyp_injectable = None
        cyp_injectable_val = {
            'de_name': 'CYPs Injectable',
            'cat_combo': None,
            'numeric_sum': cyp_injectable,
        }
        cyp_injectable_val.update(_group_ou_dict)
        calculated_vals.append(cyp_injectable_val)

        if all_not_none(iud['numeric_sum']):
            cyp_iud = iud['numeric_sum'] * Decimal(4.6)
        else:
            cyp_iud = None
        cyp_iud_val = {
            'de_name': 'CYPs IUD',
            'cat_combo': None,
            'numeric_sum': cyp_iud,
        }
        cyp_iud_val.update(_group_ou_dict)
        calculated_vals.append(cyp_iud_val)

        if all_not_none(sterile_new['numeric_sum']):
            cyp_sterile = sterile_new['numeric_sum'] * Decimal(10)
        else:
            cyp_sterile = None
        cyp_sterile_val = {
            'de_name': 'CYPs sterile',
            'cat_combo': None,
            'numeric_sum': cyp_sterile,
        }
        cyp_sterile_val.update(_group_ou_dict)
        calculated_vals.append(cyp_sterile_val)

        if all_not_none(natural['numeric_sum']):
            cyp_natural = natural['numeric_sum'] / 4
        else:
            cyp_natural = None
        cyp_natural_val = {
            'de_name': 'CYPs Natural Methods',
            'cat_combo': None,
            'numeric_sum': cyp_natural,
        }
        cyp_natural_val.update(_group_ou_dict)
        calculated_vals.append(cyp_natural_val)

        if all_not_none(emergency['numeric_sum']):
            cyp_emergency = emergency['numeric_sum'] / 20
        else:
            cyp_emergency = None
        cyp_emergency_val = {
            'de_name': 'CYPs Emergency Contraceptives',
            'cat_combo': None,
            'numeric_sum': cyp_emergency,
        }
        cyp_emergency_val.update(_group_ou_dict)
        calculated_vals.append(cyp_emergency_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['CYPs Oral'], (None,)))
    data_element_metas += list(product(['CYPs Condoms'], (None,)))
    data_element_metas += list(product(['CYPs Implants'], (None,)))
    data_element_metas += list(product(['CYPs Injectable'], (None,)))
    data_element_metas += list(product(['CYPs IUD'], (None,)))
    data_element_metas += list(product(['CYP Sterilisation'], (None,)))
    data_element_metas += list(product(['CYPs Natural Methods'], (None,)))
    data_element_metas += list(product(['CYPs Emergency contraceptives'], (None,)))

    legend_sets = list()
    # fp_cyp_ls = LegendSet()
    # fp_cyp_ls.name = 'FP CYP'
    # fp_cyp_ls.add_interval('orange', 0, 25)
    # fp_cyp_ls.add_interval('yellow', 25, 40)
    # fp_cyp_ls.add_interval('light-green', 50, 60)
    # fp_cyp_ls.add_interval('green', 60, None)
    # legend_sets.append(fp_cyp_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fp_cyp_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="fp_cyp_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/fp_cyp_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def fp_cyp_by_district(request, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all districts (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=1).annotate(district=F('name'))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by('district')
    ou_list = list(qs_ou.values_list('district'))
    ou_headers = ['District',]

    def val_with_subcat_fun(row, col):
        district, = row
        de_name, subcategory = col
        return { 'district': district, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }

    def get_ou_path(val):
        return (val['district'],)

    oral_de_names = (
        '105-2.5 Oral: Microgynon',
        '105-2.5 Oral: Lo-Feminal',
        '105-2.5 Oral : Ovrette or Another POP',
    )
    oral_short_names = (
        'Oral dispensed (cycles)',
    )
    de_oral_meta = list(product(oral_short_names, (None,)))

    qs_oral = DataValue.objects.what(*oral_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_oral = qs_oral.where(filter_district)
    qs_oral = qs_oral.annotate(de_name=Value(oral_short_names[0], output_field=CharField()))
    qs_oral = qs_oral.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_oral = qs_oral.annotate(district=F('org_unit__parent__parent__name'))
    qs_oral = qs_oral.annotate(period=F('quarter'))
    qs_oral = qs_oral.order_by('district', 'de_name', 'cat_combo', 'period')
    val_oral = qs_oral.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_oral = list(val_oral)

    gen_raster = grabbag.rasterize(ou_list, de_oral_meta, val_oral, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_oral2 = list(gen_raster)

    condoms_de_names = (
        '105-2.5 Female Condom',
        '105-2.5 Male Condom',
    )
    condoms_short_names = (
        'Condoms dispensed (pieces)',
    )
    de_condoms_meta = list(product(condoms_short_names, (None,)))

    qs_condoms = DataValue.objects.what(*condoms_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_condoms = qs_condoms.where(filter_district)
    qs_condoms = qs_condoms.annotate(de_name=Value(condoms_short_names[0], output_field=CharField()))
    qs_condoms = qs_condoms.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_condoms = qs_condoms.annotate(district=F('org_unit__parent__parent__name'))
    qs_condoms = qs_condoms.annotate(period=F('quarter'))
    qs_condoms = qs_condoms.order_by('district', 'de_name', 'cat_combo', 'period')
    val_condoms = qs_condoms.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_condoms = list(val_condoms)

    gen_raster = grabbag.rasterize(ou_list, de_condoms_meta, val_condoms, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_condoms2 = list(gen_raster)

    implants_new_de_names = (
        '105-2.7 Implant',
    )
    implants_new_short_names = (
        'New users - Implants',
    )
    de_implants_new_meta = list(product(implants_new_short_names, (None,)))

    qs_implants_new = DataValue.objects.what(*implants_new_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_implants_new = qs_implants_new.where(filter_district)
    qs_implants_new = qs_implants_new.annotate(de_name=Value(implants_new_short_names[0], output_field=CharField()))
    qs_implants_new = qs_implants_new.filter(category_combo__categories__name='New Users')
    qs_implants_new = qs_implants_new.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_implants_new = qs_implants_new.annotate(district=F('org_unit__parent__parent__name'))
    qs_implants_new = qs_implants_new.annotate(period=F('quarter'))
    qs_implants_new = qs_implants_new.order_by('district', 'de_name', 'cat_combo', 'period')
    val_implants_new = qs_implants_new.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_implants_new = list(val_implants_new)

    gen_raster = grabbag.rasterize(ou_list, de_implants_new_meta, val_implants_new, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_implants_new2 = list(gen_raster)

    injectable_de_names = (
        '105-2.5 Injectable',
    )
    injectable_short_names = (
        'Injectable dispensed (doses)',
    )
    de_injectable_meta = list(product(injectable_short_names, (None,)))

    qs_injectable = DataValue.objects.what(*injectable_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_injectable = qs_injectable.where(filter_district)
    qs_injectable = qs_injectable.annotate(de_name=Value(injectable_short_names[0], output_field=CharField()))
    qs_injectable = qs_injectable.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_injectable = qs_injectable.annotate(district=F('org_unit__parent__parent__name'))
    qs_injectable = qs_injectable.annotate(period=F('quarter'))
    qs_injectable = qs_injectable.order_by('district', 'de_name', 'cat_combo', 'period')
    val_injectable = qs_injectable.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_injectable = list(val_injectable)

    gen_raster = grabbag.rasterize(ou_list, de_injectable_meta, val_injectable, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_injectable2 = list(gen_raster)

    iud_de_names = (
        '105-2.5 IUDs',
    )
    iud_short_names = (
        'IUD inserted',
    )
    de_iud_meta = list(product(iud_short_names, (None,)))

    qs_iud = DataValue.objects.what(*iud_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_iud = qs_iud.where(filter_district)
    qs_iud = qs_iud.annotate(de_name=Value(iud_short_names[0], output_field=CharField()))
    qs_iud = qs_iud.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_iud = qs_iud.annotate(district=F('org_unit__parent__parent__name'))
    qs_iud = qs_iud.annotate(period=F('quarter'))
    qs_iud = qs_iud.order_by('district', 'de_name', 'cat_combo', 'period')
    val_iud = qs_iud.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_iud = list(val_iud)

    gen_raster = grabbag.rasterize(ou_list, de_iud_meta, val_iud, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_iud2 = list(gen_raster)

    sterile_new_de_names = (
        '105-2.7 Female Sterilisation (TubeLigation)',
        '105-2.7 Male Sterilisation (Vasectomy)',
    )
    sterile_new_short_names = (
        'New users - Sterilisation (male and female)',
    )
    de_sterile_new_meta = list(product(sterile_new_short_names, (None,)))

    qs_sterile_new = DataValue.objects.what(*sterile_new_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_sterile_new = qs_sterile_new.where(filter_district)
    qs_sterile_new = qs_sterile_new.annotate(de_name=Value(sterile_new_short_names[0], output_field=CharField()))
    qs_sterile_new = qs_sterile_new.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_sterile_new = qs_sterile_new.annotate(district=F('org_unit__parent__parent__name'))
    qs_sterile_new = qs_sterile_new.annotate(period=F('quarter'))
    qs_sterile_new = qs_sterile_new.order_by('district', 'de_name', 'cat_combo', 'period')
    val_sterile_new = qs_sterile_new.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sterile_new = list(val_sterile_new)

    gen_raster = grabbag.rasterize(ou_list, de_sterile_new_meta, val_sterile_new, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_sterile_new2 = list(gen_raster)

    natural_de_names = (
        '105-2.5 Natural',
    )
    natural_short_names = (
        'Natural methods',
    )
    de_natural_meta = list(product(natural_short_names, (None,)))

    qs_natural = DataValue.objects.what(*natural_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_natural = qs_natural.where(filter_district)
    qs_natural = qs_natural.annotate(de_name=Value(natural_short_names[0], output_field=CharField()))
    qs_natural = qs_natural.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_natural = qs_natural.annotate(district=F('org_unit__parent__parent__name'))
    qs_natural = qs_natural.annotate(period=F('quarter'))
    qs_natural = qs_natural.order_by('district', 'de_name', 'cat_combo', 'period')
    val_natural = qs_natural.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_natural = list(val_natural)

    gen_raster = grabbag.rasterize(ou_list, de_natural_meta, val_natural, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_natural2 = list(gen_raster)

    emergency_de_names = (
        '105-2.6 Emergency contraceptives  No. Dispensed by CBD',
        '105-2.6 Emergency contraceptives  No. Dispensed at Unit',
        '105-2.6 Emergency contraceptives  No. Disp. At Outreach',
    )
    emergency_short_names = (
        'Emergency contraceptives dispensed (doses)',
    )
    de_emergency_meta = list(product(emergency_short_names, (None,)))

    qs_emergency = DataValue.objects.what(*emergency_de_names).filter(quarter=filter_period)
    if filter_district:
        qs_emergency = qs_emergency.where(filter_district)
    qs_emergency = qs_emergency.annotate(de_name=Value(emergency_short_names[0], output_field=CharField()))
    qs_emergency = qs_emergency.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_emergency = qs_emergency.annotate(district=F('org_unit__parent__parent__name'))
    qs_emergency = qs_emergency.annotate(period=F('quarter'))
    qs_emergency = qs_emergency.order_by('district', 'de_name', 'cat_combo', 'period')
    val_emergency = qs_emergency.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_emergency = list(val_emergency)

    gen_raster = grabbag.rasterize(ou_list, de_emergency_meta, val_emergency, get_ou_path, lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_emergency2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_oral2, val_condoms2, val_implants_new2, val_injectable2, val_iud2, val_sterile_new2, val_natural2, val_emergency2), key=lambda x: (x['district'],)), key=lambda x: (x['district'],))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (ou_path, (oral, condoms, implants_new, injectable, iud, sterile_new, natural, emergency, *other_vals)) = _group
        
        calculated_vals = list()

        if all_not_none(oral['numeric_sum']):
            cyp_oral = oral['numeric_sum'] / 15
        else:
            cyp_oral = None
        cyp_oral_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Oral',
            'cat_combo': None,
            'numeric_sum': cyp_oral,
        }
        calculated_vals.append(cyp_oral_val)

        if all_not_none(condoms['numeric_sum']):
            cyp_condoms = condoms['numeric_sum'] / 120
        else:
            cyp_condoms = None
        cyp_condoms_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Condoms',
            'cat_combo': None,
            'numeric_sum': cyp_condoms,
        }
        calculated_vals.append(cyp_condoms_val)

        if all_not_none(implants_new['numeric_sum']):
            cyp_implants = implants_new['numeric_sum'] * Decimal(2.5)
        else:
            cyp_implants = None
        cyp_implants_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Implants',
            'cat_combo': None,
            'numeric_sum': cyp_implants,
        }
        calculated_vals.append(cyp_implants_val)

        if all_not_none(injectable['numeric_sum']):
            cyp_injectable = injectable['numeric_sum'] / 4
        else:
            cyp_injectable = None
        cyp_injectable_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Injectable',
            'cat_combo': None,
            'numeric_sum': cyp_injectable,
        }
        calculated_vals.append(cyp_injectable_val)

        if all_not_none(iud['numeric_sum']):
            cyp_iud = iud['numeric_sum'] * Decimal(4.6)
        else:
            cyp_iud = None
        cyp_iud_val = {
            'district': ou_path[0],
            'de_name': 'CYPs IUD',
            'cat_combo': None,
            'numeric_sum': cyp_iud,
        }
        calculated_vals.append(cyp_iud_val)

        if all_not_none(sterile_new['numeric_sum']):
            cyp_sterile = sterile_new['numeric_sum'] * Decimal(10)
        else:
            cyp_sterile = None
        cyp_sterile_val = {
            'district': ou_path[0],
            'de_name': 'CYPs sterile',
            'cat_combo': None,
            'numeric_sum': cyp_sterile,
        }
        calculated_vals.append(cyp_sterile_val)

        if all_not_none(natural['numeric_sum']):
            cyp_natural = natural['numeric_sum'] / 4
        else:
            cyp_natural = None
        cyp_natural_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Natural Methods',
            'cat_combo': None,
            'numeric_sum': cyp_natural,
        }
        calculated_vals.append(cyp_natural_val)

        if all_not_none(emergency['numeric_sum']):
            cyp_emergency = emergency['numeric_sum'] / 20
        else:
            cyp_emergency = None
        cyp_emergency_val = {
            'district': ou_path[0],
            'de_name': 'CYPs Emergency Contraceptives',
            'cat_combo': None,
            'numeric_sum': cyp_emergency,
        }
        calculated_vals.append(cyp_emergency_val)

        _group[1].extend(calculated_vals)

    data_element_names = list()
    data_element_names += list(product(oral_short_names, (None,)))
    data_element_names += list(product(condoms_short_names, (None,)))
    data_element_names += list(product(implants_new_short_names, (None,)))
    data_element_names += list(product(injectable_short_names, (None,)))
    data_element_names += list(product(iud_short_names, (None,)))
    data_element_names += list(product(sterile_new_short_names, (None,)))
    data_element_names += list(product(natural_short_names, (None,)))
    data_element_names += list(product(emergency_short_names, (None,)))

    data_element_names += list(product(['CYPs Oral'], (None,)))
    data_element_names += list(product(['CYPs Condoms'], (None,)))
    data_element_names += list(product(['CYPs Implants'], (None,)))
    data_element_names += list(product(['CYPs Injectable'], (None,)))
    data_element_names += list(product(['CYPs IUD'], (None,)))
    data_element_names += list(product(['CYP Sterilisation'], (None,)))
    data_element_names += list(product(['CYPs Natural Methods'], (None,)))
    data_element_names += list(product(['CYPs Emergency contraceptives'], (None,)))

    legend_sets = list()
    # fp_cyp_ls = LegendSet()
    # fp_cyp_ls.name = 'FP CYP'
    # fp_cyp_ls.add_interval('orange', 0, 25)
    # fp_cyp_ls.add_interval('yellow', 25, 40)
    # fp_cyp_ls.add_interval('light-green', 50, 60)
    # fp_cyp_ls.add_interval('green', 60, None)
    # legend_sets.append(fp_cyp_ls)

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_names)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="fp_cyp_districts_scorecard.xlsx"'

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_names,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path)
    }

    return render(request, 'cannula/fp_cyp_districts.html', context)

@login_required
def tb_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)
    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'TB_STAT (D, DSD) TARGET: New/Relapsed TB default',
    )
    targets_short_names = (
        'TARGET: New/Relapsed TB default',
    )
    de_targets_meta = list(product(targets_short_names, (None,)))
    data_element_metas += de_targets_meta

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(de_name=Value(targets_short_names[0], output_field=CharField()))
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_targets = qs_targets.when(filter_period)

    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    notif_new_de_names = (
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] New',
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] Relapse',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] New',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] Relapse',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] New',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] Relapse',
    )
    notif_new_short_names = (
        'Notification (New and Relapse)',
    )
    de_notif_new_meta = list(product(notif_new_short_names, (None,)))
    data_element_metas += de_notif_new_meta

    qs_notif_new = DataValue.objects.what(*notif_new_de_names)
    qs_notif_new = qs_notif_new.annotate(de_name=Value(notif_new_short_names[0], output_field=CharField()))
    qs_notif_new = qs_notif_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_notif_new = qs_notif_new.where(filter_district)
    qs_notif_new = qs_notif_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_notif_new = qs_notif_new.when(filter_period)

    qs_notif_new = qs_notif_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_notif_new = qs_notif_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_notif_new = list(val_notif_new)

    gen_raster = grabbag.rasterize(ou_list, de_notif_new_meta, val_notif_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_notif_new2 = list(gen_raster)

    notif_all_de_names = (
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] New',
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] Relapse',
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] Lost to Followup',
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] Failure',
        '106a 3.1.a.1 Bacteriologically confirmed, PTB (P-BC) [Cases] Trt History Unknown',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] New',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] Relapse',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] Lost to Followup',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] Failure',
        '106a 3.1.a.2 Clinically diagnosed PTB, (P-CD) [Cases] Trt History Unknown',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] New',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] Relapse',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] Lost to Followup',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] Failure',
        '106a 3.1.a.3 EPTB, (bacteriologically or clinically diagnosed) [Cases] Trt History Unknown',
    )
    notif_all_short_names = (
        'Notification (All cases)',
    )
    de_notif_all_meta = list(product(notif_all_short_names, (None,)))
    data_element_metas += de_notif_all_meta

    qs_notif_all = DataValue.objects.what(*notif_all_de_names)
    qs_notif_all = qs_notif_all.annotate(de_name=Value(notif_all_short_names[0], output_field=CharField()))
    qs_notif_all = qs_notif_all.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_notif_all = qs_notif_all.where(filter_district)
    qs_notif_all = qs_notif_all.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_notif_all = qs_notif_all.when(filter_period)

    qs_notif_all = qs_notif_all.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_notif_all = qs_notif_all.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_notif_all = list(val_notif_all)

    gen_raster = grabbag.rasterize(ou_list, de_notif_all_meta, val_notif_all, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_notif_all2 = list(gen_raster)

    hiv_tested_de_names = (
        '106a 3.1.c.1 New HIV/TB Patients Registered, PTB (P-BC) Tested for HIV',
        '106a 3.1.c.2 New HIV/TB Patients Registered, Clinically diagnosed PTB (P-CD) Tested for HIV',
        '106a 3.1.c.3 New HIV/TB Patients Registered, EPTB (BC or CD) Tested for HIV',
        '106a 3.1.c.4 New HIV/TB Patients Registered, Other types of TB Tested for HIV',
    )
    hiv_tested_short_names = (
        'Tested for HIV',
    )
    de_hiv_tested_meta = list(product(hiv_tested_short_names, (None,)))
    data_element_metas += de_hiv_tested_meta

    qs_hiv_tested = DataValue.objects.what(*hiv_tested_de_names)
    qs_hiv_tested = qs_hiv_tested.annotate(de_name=Value(hiv_tested_short_names[0], output_field=CharField()))
    qs_hiv_tested = qs_hiv_tested.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_tested = qs_hiv_tested.where(filter_district)
    qs_hiv_tested = qs_hiv_tested.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_tested = qs_hiv_tested.when(filter_period)

    qs_hiv_tested = qs_hiv_tested.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_tested = qs_hiv_tested.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_tested = list(val_hiv_tested)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_tested_meta, val_hiv_tested, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_tested2 = list(gen_raster)

    hiv_pos_de_names = (
        '106a 3.1.c.1 New HIV/TB Patients Registered, PTB (P-BC) HIV Positive',
        '106a 3.1.c.2 New HIV/TB Patients Registered, Clinically diagnosed PTB (P-CD) HIV Positive',
        '106a 3.1.c.3 New HIV/TB Patients Registered, EPTB (BC or CD) HIV Positive',
        '106a 3.1.c.4 New HIV/TB Patients Registered, Other types of TB HIV Positive',
    )
    hiv_pos_short_names = (
        'Tested HIV+',
    )
    de_hiv_pos_meta = list(product(hiv_pos_short_names, (None,)))
    data_element_metas += de_hiv_pos_meta

    qs_hiv_pos = DataValue.objects.what(*hiv_pos_de_names)
    qs_hiv_pos = qs_hiv_pos.annotate(de_name=Value(hiv_pos_short_names[0], output_field=CharField()))
    qs_hiv_pos = qs_hiv_pos.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_pos = qs_hiv_pos.where(filter_district)
    qs_hiv_pos = qs_hiv_pos.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_pos = qs_hiv_pos.when(filter_period)

    qs_hiv_pos = qs_hiv_pos.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_pos = qs_hiv_pos.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_pos = list(val_hiv_pos)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_pos_meta, val_hiv_pos, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_pos2 = list(gen_raster)

    hiv_art_de_names = (
        '106a 3.1.c.1 New HIV/TB Patients Registered, PTB (P-BC) On ART',
        '106a 3.1.c.2 New HIV/TB Patients Registered, Clinically diagnosed PTB (P-CD) On ART',
        '106a 3.1.c.3 New HIV/TB Patients Registered, EPTB (BC or CD) On ART',
        '106a 3.1.c.4 New HIV/TB Patients Registered, Other types of TB On ART',
    )
    hiv_art_short_names = (
        'HIV+ on ART',
    )
    de_hiv_art_meta = list(product(hiv_art_short_names, (None,)))
    data_element_metas += de_hiv_art_meta

    qs_hiv_art = DataValue.objects.what(*hiv_art_de_names)
    qs_hiv_art = qs_hiv_art.annotate(de_name=Value(hiv_art_short_names[0], output_field=CharField()))
    qs_hiv_art = qs_hiv_art.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_hiv_art = qs_hiv_art.where(filter_district)
    qs_hiv_art = qs_hiv_art.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_hiv_art = qs_hiv_art.when(filter_period)

    qs_hiv_art = qs_hiv_art.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_hiv_art = qs_hiv_art.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_art = list(val_hiv_art)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_art_meta, val_hiv_art, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_hiv_art2 = list(gen_raster)

    registered_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC)',
    )
    registered_short_names = (
        'Number registered',
    )
    de_registered_meta = list(product(registered_short_names, (None,)))
    data_element_metas += de_registered_meta

    qs_registered = DataValue.objects.what(*registered_de_names)
    qs_registered = qs_registered.annotate(de_name=Value(registered_short_names[0], output_field=CharField()))
    qs_registered = qs_registered.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_registered = qs_registered.where(filter_district)
    qs_registered = qs_registered.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_registered = qs_registered.when(filter_period)

    qs_registered = qs_registered.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_registered = qs_registered.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_registered = list(val_registered)

    gen_raster = grabbag.rasterize(ou_list, de_registered_meta, val_registered, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_registered2 = list(gen_raster)

    evaluated_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Cured',
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Trt Completed',
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Died',
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Failure',
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Lost to Followup',
    )
    evaluated_short_names = (
        'Number evaluated',
    )
    de_evaluated_meta = list(product(evaluated_short_names, (None,)))
    data_element_metas += de_evaluated_meta

    qs_evaluated = DataValue.objects.what(*evaluated_de_names)
    qs_evaluated = qs_evaluated.annotate(de_name=Value(evaluated_short_names[0], output_field=CharField()))
    qs_evaluated = qs_evaluated.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_evaluated = qs_evaluated.where(filter_district)
    qs_evaluated = qs_evaluated.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_evaluated = qs_evaluated.when(filter_period)

    qs_evaluated = qs_evaluated.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_evaluated = qs_evaluated.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_evaluated = list(val_evaluated)

    gen_raster = grabbag.rasterize(ou_list, de_evaluated_meta, val_evaluated, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_evaluated2 = list(gen_raster)

    cured_completed_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Cured',
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Trt Completed',
    )
    cured_completed_short_names = (
        'Number cured or completed',
    )
    de_cured_completed_meta = list(product(cured_completed_short_names, (None,)))
    data_element_metas += de_cured_completed_meta

    qs_cured_completed = DataValue.objects.what(*cured_completed_de_names)
    qs_cured_completed = qs_cured_completed.annotate(de_name=Value(cured_completed_short_names[0], output_field=CharField()))
    qs_cured_completed = qs_cured_completed.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_cured_completed = qs_cured_completed.where(filter_district)
    qs_cured_completed = qs_cured_completed.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_cured_completed = qs_cured_completed.when(filter_period)

    qs_cured_completed = qs_cured_completed.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_cured_completed = qs_cured_completed.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_cured_completed = list(val_cured_completed)

    gen_raster = grabbag.rasterize(ou_list, de_cured_completed_meta, val_cured_completed, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_cured_completed2 = list(gen_raster)

    cured_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Cured',
    )
    cured_short_names = (
        'Number Cured',
    )
    de_cured_meta = list(product(cured_short_names, (None,)))
    data_element_metas += de_cured_meta

    qs_cured = DataValue.objects.what(*cured_de_names)
    qs_cured = qs_cured.annotate(de_name=Value(cured_short_names[0], output_field=CharField()))
    qs_cured = qs_cured.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_cured = qs_cured.where(filter_district)
    qs_cured = qs_cured.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_cured = qs_cured.when(filter_period)

    qs_cured = qs_cured.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_cured = qs_cured.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_cured = list(val_cured)

    gen_raster = grabbag.rasterize(ou_list, de_cured_meta, val_cured, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_cured2 = list(gen_raster)

    ltfu_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Lost to Followup',
    )
    ltfu_short_names = (
        'LTFU',
    )
    de_ltfu_meta = list(product(ltfu_short_names, (None,)))
    data_element_metas += de_ltfu_meta

    qs_ltfu = DataValue.objects.what(*ltfu_de_names)
    qs_ltfu = qs_ltfu.annotate(de_name=Value(ltfu_short_names[0], output_field=CharField()))
    qs_ltfu = qs_ltfu.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_ltfu = qs_ltfu.where(filter_district)
    qs_ltfu = qs_ltfu.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_ltfu = qs_ltfu.when(filter_period)

    qs_ltfu = qs_ltfu.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_ltfu = qs_ltfu.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_ltfu = list(val_ltfu)

    gen_raster = grabbag.rasterize(ou_list, de_ltfu_meta, val_ltfu, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_ltfu2 = list(gen_raster)

    notif_under15_de_names = (
        '106a 3.1.b.1 Bacteriologically confirmed, PTB (P-BC) New and Relapse [Age Groups]',
        '106a 3.1.b.2 Clinically diagnosed PTB (P-CD) [Age Groups]',
        '106a 3.1.b.3 EPTB, (bacteriologically or clinically diagnosed) [Age Groups]',
    )
    notif_under15_short_names = (
        '<15 Years Notified',
    )
    de_notif_under15_meta = list(product(notif_under15_short_names, (None,)))
    data_element_metas += de_notif_under15_meta

    qs_notif_under15 = DataValue.objects.what(*notif_under15_de_names)
    qs_notif_under15 = qs_notif_under15.annotate(de_name=Value(notif_under15_short_names[0], output_field=CharField()))
    qs_notif_under15 = qs_notif_under15.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_notif_under15 = qs_notif_under15.where(filter_district)
    qs_notif_under15 = qs_notif_under15.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_notif_under15 = qs_notif_under15.when(filter_period)

    qs_notif_under15 = qs_notif_under15.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_notif_under15 = qs_notif_under15.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_notif_under15 = list(val_notif_under15)

    gen_raster = grabbag.rasterize(ou_list, de_notif_under15_meta, val_notif_under15, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_notif_under152 = list(gen_raster)

    failed_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Failure',
    )
    failed_short_names = (
        'Number failed',
    )
    de_failed_meta = list(product(failed_short_names, (None,)))
    data_element_metas += de_failed_meta

    qs_failed = DataValue.objects.what(*failed_de_names)
    qs_failed = qs_failed.annotate(de_name=Value(failed_short_names[0], output_field=CharField()))
    qs_failed = qs_failed.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_failed = qs_failed.where(filter_district)
    qs_failed = qs_failed.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_failed = qs_failed.when(filter_period)

    qs_failed = qs_failed.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_failed = qs_failed.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_failed = list(val_failed)

    gen_raster = grabbag.rasterize(ou_list, de_failed_meta, val_failed, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_failed2 = list(gen_raster)

    died_de_names = (
        '106a 3.1.h.1 TB Treat. Outcome (All): New Patients Category I (PTB-BC) Died',
    )
    died_short_names = (
        'Number died',
    )
    de_died_meta = list(product(died_short_names, (None,)))
    data_element_metas += de_died_meta

    qs_died = DataValue.objects.what(*died_de_names)
    qs_died = qs_died.annotate(de_name=Value(died_short_names[0], output_field=CharField()))
    qs_died = qs_died.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_died = qs_died.where(filter_district)
    qs_died = qs_died.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_died = qs_died.when(filter_period)

    qs_died = qs_died.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_died = qs_died.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_died = list(val_died)

    gen_raster = grabbag.rasterize(ou_list, de_died_meta, val_died, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_died2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_notif_new2, val_notif_all2, val_hiv_tested2, val_hiv_pos2, val_hiv_art2, val_registered2, val_evaluated2, val_cured_completed2, val_cured2, val_ltfu2, val_notif_under152, val_failed2, val_died2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (g_ou_path, (target_notif_new, notif_new, notif_all, hiv_tested, hiv_pos, hiv_art, registered, evaluated, cured_completed, cured, ltfu, notif_under15, failed, died, *other_vals)) = _group
        
        calculated_vals = list()
        g_ou_dict = ou_dict_from_path(g_ou_path)

        if all_not_none(cured_completed['numeric_sum'], evaluated['numeric_sum']) and evaluated['numeric_sum']:
            tsr_percent = 100 * cured_completed['numeric_sum'] / evaluated['numeric_sum']
        else:
            tsr_percent = None
        tsr_percent_val = {
            'de_name': '% TSR',
            'cat_combo': None,
            'numeric_sum': tsr_percent,
        }
        tsr_percent_val.update(g_ou_dict)
        calculated_vals.append(tsr_percent_val)

        if all_not_none(ltfu['numeric_sum'], evaluated['numeric_sum']) and evaluated['numeric_sum']:
            ltfu_percent = 100 * ltfu['numeric_sum'] / evaluated['numeric_sum']
        else:
            ltfu_percent = None
        ltfu_percent_val = {
            'de_name': '% LTFU',
            'cat_combo': None,
            'numeric_sum': ltfu_percent,
        }
        ltfu_percent_val.update(g_ou_dict)
        calculated_vals.append(ltfu_percent_val)

        if all_not_none(notif_new['numeric_sum'], target_notif_new['numeric_sum']) and target_notif_new['numeric_sum']:
            notif_new_percent = 100 * notif_new['numeric_sum'] / target_notif_new['numeric_sum']
        else:
            notif_new_percent = None
        notif_new_percent_val = {
            'de_name': '% of cases notified (NEW & Relapse)',
            'cat_combo': None,
            'numeric_sum': notif_new_percent,
        }
        notif_new_percent_val.update(g_ou_dict)
        calculated_vals.append(notif_new_percent_val)

        if all_not_none(hiv_tested['numeric_sum'], notif_all['numeric_sum']) and notif_all['numeric_sum']:
            hiv_tested_percent = 100 * hiv_tested['numeric_sum'] / notif_all['numeric_sum']
        else:
            hiv_tested_percent = None
        hiv_tested_percent_val = {
            'de_name': '% Tested for HIV',
            'cat_combo': None,
            'numeric_sum': hiv_tested_percent,
        }
        hiv_tested_percent_val.update(g_ou_dict)
        calculated_vals.append(hiv_tested_percent_val)

        if all_not_none(hiv_art['numeric_sum'], hiv_pos['numeric_sum']) and hiv_pos['numeric_sum']:
            hiv_art_percent = 100 * hiv_art['numeric_sum'] / hiv_pos['numeric_sum']
        else:
            hiv_art_percent = None
        hiv_art_percent_val = {
            'de_name': '% HIV+ on ART',
            'cat_combo': None,
            'numeric_sum': hiv_art_percent,
        }
        hiv_art_percent_val.update(g_ou_dict)
        calculated_vals.append(hiv_art_percent_val)

        if all_not_none(cured['numeric_sum'], evaluated['numeric_sum']) and evaluated['numeric_sum']:
            cure_percent = 100 * cured['numeric_sum'] / evaluated['numeric_sum']
        else:
            cure_percent = None
        cure_percent_val = {
            'de_name': '% Cure Rate',
            'cat_combo': None,
            'numeric_sum': cure_percent,
        }
        cure_percent_val.update(g_ou_dict)
        calculated_vals.append(cure_percent_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['% TSR'], (None,)))
    data_element_metas += list(product(['% LTFU'], (None,)))
    data_element_metas += list(product(['% of cases notified (NEW & Relapse)'], (None,)))
    data_element_metas += list(product(['% Tested for HIV'], (None,)))
    data_element_metas += list(product(['% HIV+ on ART'], (None,)))
    data_element_metas += list(product(['% Cure Rate'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    notif_ls = LegendSet()
    notif_ls.name = 'Notification, Testing and ART'
    notif_ls.add_interval('red', 0, 75)
    notif_ls.add_interval('yellow', 75, 95)
    notif_ls.add_interval('green', 95, None)
    notif_ls.mappings[num_path_elements+16] = True
    notif_ls.mappings[num_path_elements+17] = True
    notif_ls.mappings[num_path_elements+18] = True
    legend_sets.append(notif_ls)
    cure_ls = LegendSet()
    cure_ls.name = 'Cure Rate'
    cure_ls.add_interval('red', 0, 50)
    cure_ls.add_interval('yellow', 50, 60)
    cure_ls.add_interval('green', 60, None)
    cure_ls.mappings[num_path_elements+19] = True
    legend_sets.append(cure_ls)
    tsr_ls = LegendSet()
    tsr_ls.name = 'TSR'
    tsr_ls.add_interval('red', 0, 80)
    tsr_ls.add_interval('yellow', 80, 85)
    tsr_ls.add_interval('green', 85, None)
    tsr_ls.mappings[num_path_elements+14] = True
    legend_sets.append(tsr_ls)
    cnr_ls = LegendSet()
    cnr_ls.name = 'CNR'
    cnr_ls.add_interval('red', 0, 85)
    cnr_ls.add_interval('yellow', 85, 115)
    cnr_ls.add_interval('green', 115, None)
    legend_sets.append(cnr_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="tb_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="tb_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/tb_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

def nutrition_dashboard(request):
    this_day = date.today()
    this_quarter = '%d-Q%d' % (this_day.year, month2quarter(this_day.month))
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_day.year, this_day.year-6, -1) for q in range(4, 0, -1)]
    period_list = list(filter(lambda qtr: qtr < this_quarter, reversed(PREV_5YR_QTRS)))[-6:]
    def val_with_period_de_fun(row, col):
        period = row
        de_name = col
        return { 'de_name': de_name, 'period': period, 'numeric_sum': None }

    data_element_metas = list()
   
    opd_attend_de_names = (
        '105-1.1 OPD New Attendance',
        '105-1.1 OPD Re-Attendance',
    )
    opd_attend_short_names = (
        'Total OPD attendence',
    )
    de_opd_attend_meta = list(product(opd_attend_short_names, (None,)))
    data_element_metas += de_opd_attend_meta

    qs_opd_attend = DataValue.objects.what(*opd_attend_de_names)
    qs_opd_attend = qs_opd_attend.annotate(de_name=Value(opd_attend_short_names[0], output_field=CharField()))
    qs_opd_attend = qs_opd_attend.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_opd_attend = qs_opd_attend.when(*period_list)
    qs_opd_attend = qs_opd_attend.order_by('period', 'de_name')
    val_opd_attend = qs_opd_attend.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(period_list, opd_attend_short_names, val_opd_attend, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_opd_attend2 = list(gen_raster)
   
    muac_de_names = (
        '106a Nutri No. 1 of clients who received nutrition assessment in this quarter using color coded MUAC tapes/Z score chart',
    )
    muac_short_names = (
        'Clients assessed using MUAC/Z score in OPD',
    )
    de_muac_meta = list(product(muac_short_names, (None,)))
    data_element_metas += de_muac_meta

    qs_muac = DataValue.objects.what(*muac_de_names)
    qs_muac = qs_muac.annotate(de_name=Value(muac_short_names[0], output_field=CharField()))
    qs_muac = qs_muac.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_muac = qs_muac.when(*period_list)
    qs_muac = qs_muac.order_by('period', 'de_name')
    val_muac = qs_muac.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(period_list, muac_short_names, val_muac, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_muac2 = list(gen_raster)
   
    active_art_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    active_art_short_names = (
        'Total No. active on ART in the quarter',
    )
    de_active_art_meta = list(product(active_art_short_names, (None,)))
    data_element_metas += de_active_art_meta

    qs_active_art = DataValue.objects.what(*active_art_de_names)
    qs_active_art = qs_active_art.annotate(de_name=Value(active_art_short_names[0], output_field=CharField()))
    qs_active_art = qs_active_art.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_active_art = qs_active_art.when(*period_list)
    qs_active_art = qs_active_art.order_by('period', 'de_name')
    val_active_art = qs_active_art.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(period_list, active_art_short_names, val_active_art, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_active_art2 = list(gen_raster)
   
    active_art_malnourish_de_names = (
        '106a ART No. active on ART assessed for Malnutrition at their visit in quarter',
    )
    active_art_malnourish_short_names = (
        '106a ART No. active on ART assessed for Malnutrition at their visit in quarter',
    )
    de_active_art_malnourish_meta = list(product(active_art_malnourish_short_names, (None,)))
    data_element_metas += de_active_art_malnourish_meta

    qs_active_art_malnourish = DataValue.objects.what(*active_art_malnourish_de_names)
    qs_active_art_malnourish = qs_active_art_malnourish.annotate(de_name=Value(active_art_malnourish_short_names[0], output_field=CharField()))
    qs_active_art_malnourish = qs_active_art_malnourish.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_active_art_malnourish = qs_active_art_malnourish.when(*period_list)
    qs_active_art_malnourish = qs_active_art_malnourish.order_by('period', 'de_name')
    val_active_art_malnourish = qs_active_art_malnourish.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(period_list, active_art_malnourish_short_names, val_active_art_malnourish, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_active_art_malnourish2 = list(gen_raster)
   
    new_malnourish_de_names = (
        '106a Nutri N4-No. of newly identified malnourished cases in this quarter - Total',
    )
    new_malnourish_short_names = (
        'No of newly identified malnourished cases in this quarter',
    )
    de_new_malnourish_meta = list(product(new_malnourish_short_names, (None,)))
    data_element_metas += de_new_malnourish_meta

    qs_new_malnourish = DataValue.objects.what(*new_malnourish_de_names)
    qs_new_malnourish = qs_new_malnourish.annotate(de_name=Value(new_malnourish_short_names[0], output_field=CharField()))
    qs_new_malnourish = qs_new_malnourish.annotate(cat_combo=Value(None, output_field=CharField()))
    qs_new_malnourish = qs_new_malnourish.when(*period_list)
    qs_new_malnourish = qs_new_malnourish.order_by('period', 'de_name')
    val_new_malnourish = qs_new_malnourish.values('period', 'de_name').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(period_list, new_malnourish_short_names, val_new_malnourish, lambda x: x['period'], lambda x: x['de_name'], val_with_period_de_fun)
    val_new_malnourish2 = list(gen_raster)

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_opd_attend2, val_muac2, val_active_art2, val_active_art_malnourish2, val_new_malnourish2), key=lambda x: (x['period'])), key=lambda x: (x['period']))
    # if True:
    #     grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (period, (opd_attend, muac, active_art, active_art_malnourish, new_malnourish, *other_vals)) = _group
        
        calculated_vals = list()

        if all_not_none(muac['numeric_sum'], opd_attend['numeric_sum']) and opd_attend['numeric_sum']:
            assessment_percent = (muac['numeric_sum'] * 100) / opd_attend['numeric_sum']
        else:
            assessment_percent = None
        assessment_percent_val = {
            'period': period,
            'de_name': '% of clients who received nutrition asssessment  in OPD',
            'numeric_sum': assessment_percent,
        }
        calculated_vals.append(assessment_percent_val)

        if all_not_none(active_art['numeric_sum'], active_art_malnourish['numeric_sum']) and active_art['numeric_sum']:
            active_art_malnourish_percent = (active_art_malnourish['numeric_sum'] * 100) / active_art['numeric_sum']
        else:
            active_art_malnourish_percent = None
        active_art_malnourish_percent_val = {
            'period': period,
            'de_name': '% of active on ART assessed for Malnutrition at their visit in quarter',
            'numeric_sum': active_art_malnourish_percent,
        }
        calculated_vals.append(active_art_malnourish_percent_val)

        _group[1] = calculated_vals
    
    context = {
        'data_element_names': [
            ('% of clients who received nutrition asssessment  in OPD', None),
            ('% of active on ART assessed for Malnutrition at their visit in quarter', None),
        ],
        'grouped_data': grouped_vals,
    }
    return render(request, 'cannula/index.html', context)

@login_required
def nutrition_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()
   
    opd_attend_de_names = (
        '105-1.1 OPD New Attendance',
        '105-1.1 OPD Re-Attendance',
    )
    opd_attend_short_names = (
        'Total OPD attendence',
    )
    de_opd_attend_meta = list(product(opd_attend_short_names, (None,)))
    data_element_metas += de_opd_attend_meta

    qs_opd_attend = DataValue.objects.what(*opd_attend_de_names)
    qs_opd_attend = qs_opd_attend.annotate(de_name=Value(opd_attend_short_names[0], output_field=CharField()))
    qs_opd_attend = qs_opd_attend.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_opd_attend = qs_opd_attend.where(filter_district)
    qs_opd_attend = qs_opd_attend.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_opd_attend = qs_opd_attend.when(filter_period)
    qs_opd_attend = qs_opd_attend.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_opd_attend = qs_opd_attend.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_opd_attend_meta, val_opd_attend, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_opd_attend2 = list(gen_raster)
   
    muac_de_names = (
        '106a Nutri No. 1 of clients who received nutrition assessment in this quarter using color coded MUAC tapes/Z score chart',
    )
    muac_short_names = (
        'Clients assessed using MUAC/Z score in OPD',
    )
    de_muac_meta = list(product(muac_short_names, (None,)))
    data_element_metas += de_muac_meta

    qs_muac = DataValue.objects.what(*muac_de_names)
    qs_muac = qs_muac.annotate(de_name=Value(muac_short_names[0], output_field=CharField()))
    qs_muac = qs_muac.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_muac = qs_muac.where(filter_district)
    qs_muac = qs_muac.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_muac = qs_muac.when(filter_period)
    qs_muac = qs_muac.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_muac = qs_muac.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_muac_meta, val_muac, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_muac2 = list(gen_raster)
   
    muac_mothers_de_names = (
        '106a Nutri No. 1 of clients who received nutrition assessment in this quarter using color coded MUAC tapes/Z score chart Pregnant/Lactating Women',
    )
    muac_mothers_short_names = (
        'Clients assessed using MUAC/Z score in OPD - Pregnant/Lactating Women',
    )
    de_muac_mothers_meta = list(product(muac_mothers_short_names, (None,)))
    data_element_metas += de_muac_mothers_meta

    qs_muac_mothers = DataValue.objects.what(*muac_mothers_de_names)
    qs_muac_mothers = qs_muac_mothers.annotate(de_name=Value(muac_mothers_short_names[0], output_field=CharField()))
    qs_muac_mothers = qs_muac_mothers.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_muac_mothers = qs_muac_mothers.where(filter_district)
    qs_muac_mothers = qs_muac_mothers.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_muac_mothers = qs_muac_mothers.when(filter_period)
    qs_muac_mothers = qs_muac_mothers.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_muac_mothers = qs_muac_mothers.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_muac_mothers_meta, val_muac_mothers, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_muac_mothers2 = list(gen_raster)
   
    mothers_total_de_names = (
        '105-2.1 A3:Total ANC visits (New clients + Re-attendances)',
        '105-2.3 Postnatal Attendances',
    )
    mothers_total_short_names = (
        'Total number of pregnant and lactating mothers',
    )
    de_mothers_total_meta = list(product(mothers_total_short_names, (None,)))
    data_element_metas += de_mothers_total_meta

    qs_mothers_total = DataValue.objects.what(*mothers_total_de_names)
    qs_mothers_total = qs_mothers_total.annotate(de_name=Value(mothers_total_short_names[0], output_field=CharField()))
    qs_mothers_total = qs_mothers_total.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_mothers_total = qs_mothers_total.where(filter_district)
    qs_mothers_total = qs_mothers_total.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_mothers_total = qs_mothers_total.when(filter_period)
    qs_mothers_total = qs_mothers_total.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_mothers_total = qs_mothers_total.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_mothers_total_meta, val_mothers_total, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_mothers_total2 = list(gen_raster)
   
    i_f_counsel_de_names = (
        '106a Nutri N7-No. of pregnant and lactating women who received infant feeding counseling - Total',
    )
    i_f_counsel_short_names = (
        '106a Nutri N7-No. of pregnant and lactating women who received infant feeding counseling - Total',
    )
    de_i_f_counsel_meta = list(product(i_f_counsel_short_names, (None,)))
    data_element_metas += de_i_f_counsel_meta

    qs_i_f_counsel = DataValue.objects.what(*i_f_counsel_de_names)
    qs_i_f_counsel = qs_i_f_counsel.annotate(de_name=Value(i_f_counsel_short_names[0], output_field=CharField()))
    qs_i_f_counsel = qs_i_f_counsel.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_i_f_counsel = qs_i_f_counsel.where(filter_district)
    qs_i_f_counsel = qs_i_f_counsel.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_i_f_counsel = qs_i_f_counsel.when(filter_period)
    qs_i_f_counsel = qs_i_f_counsel.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_i_f_counsel = qs_i_f_counsel.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_i_f_counsel_meta, val_i_f_counsel, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_i_f_counsel2 = list(gen_raster)
   
    m_n_counsel_de_names = (
        '106a Nutri N6-No. of pregnant and lactating women who received maternal nutrition counseling - Total',
    )
    m_n_counsel_short_names = (
        '106a Nutri N6-No. of pregnant and lactating women who received maternal nutrition counseling - Total',
    )
    de_m_n_counsel_meta = list(product(m_n_counsel_short_names, (None,)))
    data_element_metas += de_m_n_counsel_meta

    qs_m_n_counsel = DataValue.objects.what(*m_n_counsel_de_names)
    qs_m_n_counsel = qs_m_n_counsel.annotate(de_name=Value(m_n_counsel_short_names[0], output_field=CharField()))
    qs_m_n_counsel = qs_m_n_counsel.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_m_n_counsel = qs_m_n_counsel.where(filter_district)
    qs_m_n_counsel = qs_m_n_counsel.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_m_n_counsel = qs_m_n_counsel.when(filter_period)
    qs_m_n_counsel = qs_m_n_counsel.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_m_n_counsel = qs_m_n_counsel.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_m_n_counsel_meta, val_m_n_counsel, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_m_n_counsel2 = list(gen_raster)
   
    active_art_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    active_art_short_names = (
        'Total No. active on ART in the quarter',
    )
    de_active_art_meta = list(product(active_art_short_names, (None,)))
    data_element_metas += de_active_art_meta

    qs_active_art = DataValue.objects.what(*active_art_de_names)
    qs_active_art = qs_active_art.annotate(de_name=Value(active_art_short_names[0], output_field=CharField()))
    qs_active_art = qs_active_art.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_active_art = qs_active_art.where(filter_district)
    qs_active_art = qs_active_art.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_active_art = qs_active_art.when(filter_period)
    qs_active_art = qs_active_art.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_active_art = qs_active_art.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_active_art_meta, val_active_art, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_active_art2 = list(gen_raster)
   
    active_art_malnourish_de_names = (
        '106a ART No. active on ART assessed for Malnutrition at their visit in quarter',
    )
    active_art_malnourish_short_names = (
        '106a ART No. active on ART assessed for Malnutrition at their visit in quarter',
    )
    de_active_art_malnourish_meta = list(product(active_art_malnourish_short_names, (None,)))
    data_element_metas += de_active_art_malnourish_meta

    qs_active_art_malnourish = DataValue.objects.what(*active_art_malnourish_de_names)
    qs_active_art_malnourish = qs_active_art_malnourish.annotate(de_name=Value(active_art_malnourish_short_names[0], output_field=CharField()))
    qs_active_art_malnourish = qs_active_art_malnourish.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_active_art_malnourish = qs_active_art_malnourish.where(filter_district)
    qs_active_art_malnourish = qs_active_art_malnourish.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_active_art_malnourish = qs_active_art_malnourish.when(filter_period)
    qs_active_art_malnourish = qs_active_art_malnourish.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_active_art_malnourish = qs_active_art_malnourish.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_active_art_malnourish_meta, val_active_art_malnourish, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_active_art_malnourish2 = list(gen_raster)
   
    new_malnourish_de_names = (
        '106a Nutri N4-No. of newly identified malnourished cases in this quarter - Total',
    )
    new_malnourish_short_names = (
        'No of newly identified malnourished cases in this quarter',
    )
    de_new_malnourish_meta = list(product(new_malnourish_short_names, (None,)))
    data_element_metas += de_new_malnourish_meta

    qs_new_malnourish = DataValue.objects.what(*new_malnourish_de_names)
    qs_new_malnourish = qs_new_malnourish.annotate(de_name=Value(new_malnourish_short_names[0], output_field=CharField()))
    qs_new_malnourish = qs_new_malnourish.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_new_malnourish = qs_new_malnourish.where(filter_district)
    qs_new_malnourish = qs_new_malnourish.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_new_malnourish = qs_new_malnourish.when(filter_period)
    qs_new_malnourish = qs_new_malnourish.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_new_malnourish = qs_new_malnourish.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_new_malnourish_meta, val_new_malnourish, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_new_malnourish2 = list(gen_raster)
   
    supp_feeding_de_names = (
        '106a Nutri N5-No. of clients who received nutrition supplementary / therapeutic feeds - Total',
    )
    supp_feeding_short_names = (
        'No. of clients who received nutrition suplementary/therapeutic feeds',
    )
    de_supp_feeding_meta = list(product(supp_feeding_short_names, (None,)))
    data_element_metas += de_supp_feeding_meta

    qs_supp_feeding = DataValue.objects.what(*supp_feeding_de_names)
    qs_supp_feeding = qs_supp_feeding.annotate(de_name=Value(supp_feeding_short_names[0], output_field=CharField()))
    qs_supp_feeding = qs_supp_feeding.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_supp_feeding = qs_supp_feeding.where(filter_district)
    qs_supp_feeding = qs_supp_feeding.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_supp_feeding = qs_supp_feeding.when(filter_period)
    qs_supp_feeding = qs_supp_feeding.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_supp_feeding = qs_supp_feeding.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_supp_feeding_meta, val_supp_feeding, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_supp_feeding2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_opd_attend2, val_muac2, val_muac_mothers2, val_mothers_total2, val_i_f_counsel2, val_m_n_counsel2, val_active_art2, val_active_art_malnourish2, val_new_malnourish2, val_supp_feeding2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (opd_attend, muac, muac_mothers, mothers, infant_feeding, maternal_nutrition, active_art, active_art_malnourish, new_malnourish, supp_feeding, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(muac['numeric_sum'], opd_attend['numeric_sum']) and opd_attend['numeric_sum']:
            assessment_percent = (muac['numeric_sum'] * 100) / opd_attend['numeric_sum']
        else:
            assessment_percent = None
        assessment_percent_val = {
            'de_name': '% of clients who received nutrition asssessment  in OPD',
            'cat_combo': None,
            'numeric_sum': assessment_percent,
        }
        assessment_percent_val.update(_group_ou_dict)
        calculated_vals.append(assessment_percent_val)

        if all_not_none(mothers['numeric_sum'], muac_mothers['numeric_sum']) and mothers['numeric_sum']:
            assessment_mothers_percent = (muac_mothers['numeric_sum'] * 100) / mothers['numeric_sum']
        else:
            assessment_mothers_percent = None
        assessment_mothers_percent_val = {
            'de_name': '% of clients who received nutrition assessment   Pregnant/Lactating Women',
            'cat_combo': None,
            'numeric_sum': assessment_mothers_percent,
        }
        assessment_mothers_percent_val.update(_group_ou_dict)
        calculated_vals.append(assessment_mothers_percent_val)

        if all_not_none(active_art['numeric_sum'], active_art_malnourish['numeric_sum']) and active_art['numeric_sum']:
            active_art_malnourish_percent = (active_art_malnourish['numeric_sum'] * 100) / active_art['numeric_sum']
        else:
            active_art_malnourish_percent = None
        active_art_malnourish_percent_val = {
            'de_name': '% of active on ART assessed for Malnutrition at their visit in quarter',
            'cat_combo': None,
            'numeric_sum': active_art_malnourish_percent,
        }
        active_art_malnourish_percent_val.update(_group_ou_dict)
        calculated_vals.append(active_art_malnourish_percent_val)

        if all_not_none(mothers['numeric_sum'], infant_feeding['numeric_sum']) and mothers['numeric_sum']:
            mothers_i_f_percent = (infant_feeding['numeric_sum'] * 100) / mothers['numeric_sum']
        else:
            mothers_i_f_percent = None
        mothers_i_f_percent_val = {
            'de_name': '% of pregnant and lactating women who received infant feeding counseling ',
            'cat_combo': None,
            'numeric_sum': mothers_i_f_percent,
        }
        mothers_i_f_percent_val.update(_group_ou_dict)
        calculated_vals.append(mothers_i_f_percent_val)

        if all_not_none(mothers['numeric_sum'], maternal_nutrition['numeric_sum']) and mothers['numeric_sum']:
            mothers_m_n_percent = (maternal_nutrition['numeric_sum'] * 100) / mothers['numeric_sum']
        else:
            mothers_m_n_percent = None
        mothers_m_n_percent_val = {
            'de_name': '% of pregnant and lactating women who received maternal nutrition counseling ',
            'cat_combo': None,
            'numeric_sum': mothers_m_n_percent,
        }
        mothers_m_n_percent_val.update(_group_ou_dict)
        calculated_vals.append(mothers_m_n_percent_val)

        if all_not_none(new_malnourish['numeric_sum'], supp_feeding['numeric_sum']) and new_malnourish['numeric_sum']:
            supp_feeding_percent = (supp_feeding['numeric_sum'] * 100) / new_malnourish['numeric_sum']
        else:
            supp_feeding_percent = None
        supp_feeding_percent_val = {
            'de_name': '% of newly identified malnorished cases who received nutrition suplementary/ therapeutic feeds',
            'cat_combo': None,
            'numeric_sum': supp_feeding_percent,
        }
        supp_feeding_percent_val.update(_group_ou_dict)
        calculated_vals.append(supp_feeding_percent_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['% of clients who received nutrition asssessment  in OPD'], (None,)))
    data_element_metas += list(product(['% of clients who received nutrition assessment   Pregnant/Lactating Women'], (None,)))
    data_element_metas += list(product(['% of active on ART assessed for Malnutrition at their visit in quarter'], (None,)))
    data_element_metas += list(product(['% of pregnant and lactating women who received infant feeding counseling '], (None,)))
    data_element_metas += list(product(['% of pregnant and lactating women who received maternal nutrition counseling '], (None,)))
    data_element_metas += list(product(['% of newly identified malnorished cases who received nutrition suplementary/ therapeutic feeds'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    muac_ls = LegendSet()
    muac_ls.name = 'Nutrition Assessment'
    muac_ls.add_interval('red', 0, 25)
    muac_ls.add_interval('yellow', 25, 50)
    muac_ls.add_interval('green', 50, None)
    muac_ls.mappings[num_path_elements+10] = True
    legend_sets.append(muac_ls)
    malnourished_ls = LegendSet()
    malnourished_ls.name = 'Assessed for Malnutrition'
    malnourished_ls.add_interval('red', 0, 50)
    malnourished_ls.add_interval('yellow', 50, 80)
    malnourished_ls.add_interval('green', 80, None)
    for i in range(num_path_elements+10+1, num_path_elements+10+1+5):
        malnourished_ls.mappings[i] = True
    legend_sets.append(malnourished_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="nutrition_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="nutrition_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/nutrition_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def vl_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    viral_load_de_names = (
        'VL samples rejected',
        'VL samples sent',
    )
    viral_load_short_names = (
        'VL samples rejected',
        'VL samples sent',
    )
    de_viral_load_meta = list(product(viral_load_short_names, (None,)))

    qs_viral_load = DataValue.objects.what(*viral_load_de_names)
    qs_viral_load = qs_viral_load.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_viral_load = qs_viral_load.where(filter_district)
    qs_viral_load = qs_viral_load.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_viral_load = qs_viral_load.when(filter_period)
    qs_viral_load = qs_viral_load.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_viral_load = qs_viral_load.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_viral_load = list(val_viral_load)

    gen_raster = grabbag.rasterize(ou_list, de_viral_load_meta, val_viral_load, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_viral_load2 = list(gen_raster)

    viral_target_de_names = (
        'VL_TARGET',
    )
    viral_target_short_names = (
        'Samples target',
    )
    de_viral_target_meta = list(product(viral_target_short_names, (None,)))

    qs_viral_target = DataValue.objects.what(*viral_target_de_names)
    qs_viral_target = qs_viral_target.annotate(de_name=Value(viral_target_short_names[0], output_field=CharField()))
    qs_viral_target = qs_viral_target.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_viral_target = qs_viral_target.where(filter_district)
    qs_viral_target = qs_viral_target.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_viral_target = qs_viral_target.when(filter_period[:4])
    qs_viral_target = qs_viral_target.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_viral_target = qs_viral_target.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_viral_target = list(val_viral_target)

    gen_raster = grabbag.rasterize(ou_list, de_viral_target_meta, val_viral_target, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_viral_target2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_viral_target2, val_viral_load2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (vl_target, vl_rejected, vl_sent, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(vl_target['numeric_sum'], vl_sent['numeric_sum']) and vl_target['numeric_sum']:
            vl_sent_percent = (vl_sent['numeric_sum'] * 100) / vl_target['numeric_sum']
        else:
            vl_sent_percent = None
        vl_sent_percent_val = {
            'de_name': '% Achievement (sent)',
            'cat_combo': None,
            'numeric_sum': vl_sent_percent,
        }
        vl_sent_percent_val.update(_group_ou_dict)
        calculated_vals.append(vl_sent_percent_val)

        if all_not_none(vl_sent['numeric_sum'], vl_rejected['numeric_sum']) and vl_sent['numeric_sum']:
            vl_rejected_percent = (vl_rejected['numeric_sum'] * 100) / vl_sent['numeric_sum']
        else:
            vl_rejected_percent = None
        vl_rejected_percent_val = {
            'de_name': '% Sample rejection',
            'cat_combo': None,
            'numeric_sum': vl_rejected_percent,
        }
        vl_rejected_percent_val.update(_group_ou_dict)
        calculated_vals.append(vl_rejected_percent_val)

        vl_returned = default_zero(vl_sent['numeric_sum']) - default_zero(vl_rejected['numeric_sum'])
        vl_returned_val = {
            'de_name': 'Samples returned',
            'cat_combo': 'None',
            'numeric_sum': vl_returned,
        }
        vl_returned_val.update(_group_ou_dict)
        calculated_vals.append(vl_returned_val)

        if all_not_none(vl_sent['numeric_sum'], vl_returned) and vl_sent['numeric_sum']:
            vl_returned_percent = (vl_returned * 100) / vl_sent['numeric_sum']
        else:
            vl_returned_percent = None
        vl_returned_percent_val = {
            'de_name': '% Achievement (returned)',
            'cat_combo': None,
            'numeric_sum': vl_returned_percent,
        }
        vl_returned_percent_val.update(_group_ou_dict)
        calculated_vals.append(vl_returned_percent_val)

        _group[1].extend(calculated_vals)

    data_element_metas = list()
    
    data_element_metas += de_viral_target_meta
    data_element_metas += de_viral_load_meta

    data_element_metas += list(product(['% Achievement'], (None,)))
    data_element_metas += list(product(['% Sample rejection'], (None,)))
    data_element_metas += list(product(['Samples returned'], (None,)))
    data_element_metas += list(product(['% Achievement'], (None,)))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    achievement_ls = LegendSet()
    achievement_ls.name = 'Achievement'
    achievement_ls.add_interval('orange', 0, 25)
    achievement_ls.add_interval('yellow', 25, 40)
    achievement_ls.add_interval('light-green', 40, 60)
    achievement_ls.add_interval('green', 60, None)
    achievement_ls.mappings[num_path_elements+3] = True
    achievement_ls.mappings[num_path_elements+6] = True
    legend_sets.append(achievement_ls)
    rejection_ls = LegendSet()
    rejection_ls.name = 'Sample Rejection'
    rejection_ls.add_interval('orange', 4, None)
    rejection_ls.mappings[num_path_elements+4] = True
    legend_sets.append(rejection_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="viral_load_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="viral_load_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/vl_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def gbv_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'GEND_GBV TARGET: GBV Care',
        'GEND_GBV TARGET: GBV Care Physical and/or Emotional Violence',
        'GEND_GBV TARGET: GBV Care Sexual Violence (Post-Rape Care)',
    )
    targets_short_names = (
        'TARGET: GBV care',
        'TARGET: Physical and/or emotional violence',
        'TARGET: Sexual violence',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))
    data_element_metas += list(product(targets_short_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    targets_care_female_de_names = (
        'GEND_GBV TARGET: GBV Care',
    )
    targets_care_female_short_names = (
        'TARGET: GBV care - Female',
    )
    de_targets_care_female_meta = list(product(targets_care_female_short_names, ('Female',)))
    data_element_metas += de_targets_care_female_meta

    qs_targets_care_female = DataValue.objects.what(*targets_care_female_de_names)
    qs_targets_care_female = qs_targets_care_female.annotate(de_name=Value(targets_care_female_short_names[0], output_field=CharField()))
    qs_targets_care_female = qs_targets_care_female.filter(category_combo__categories__name='Female')
    qs_targets_care_female = qs_targets_care_female.annotate(cat_combo=Value('Female', output_field=CharField()))
    if filter_district:
        qs_targets_care_female = qs_targets_care_female.where(filter_district)
    qs_targets_care_female = qs_targets_care_female.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets_care_female = qs_targets_care_female.when(filter_period[:4])
    qs_targets_care_female = qs_targets_care_female.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets_care_female = qs_targets_care_female.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets_care_female = list(val_targets_care_female)

    gen_raster = grabbag.rasterize(ou_list, de_targets_care_female_meta, val_targets_care_female, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets_care_female2 = list(gen_raster)

    targets_care_male_de_names = (
        'GEND_GBV TARGET: GBV Care',
    )
    targets_care_male_short_names = (
        'TARGET: GBV care - Male',
    )
    de_targets_care_male_meta = list(product(targets_care_male_short_names, ('Male',)))
    data_element_metas += de_targets_care_male_meta

    qs_targets_care_male = DataValue.objects.what(*targets_care_male_de_names)
    qs_targets_care_male = qs_targets_care_male.annotate(de_name=Value(targets_care_male_short_names[0], output_field=CharField()))
    qs_targets_care_male = qs_targets_care_male.filter(category_combo__categories__name='Male')
    qs_targets_care_male = qs_targets_care_male.annotate(cat_combo=Value('Male', output_field=CharField()))
    if filter_district:
        qs_targets_care_male = qs_targets_care_male.where(filter_district)
    qs_targets_care_male = qs_targets_care_male.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets_care_male = qs_targets_care_male.when(filter_period[:4])
    qs_targets_care_male = qs_targets_care_male.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets_care_male = qs_targets_care_male.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets_care_male = list(val_targets_care_male)

    gen_raster = grabbag.rasterize(ou_list, de_targets_care_male_meta, val_targets_care_male, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets_care_male2 = list(gen_raster)

    targets_pep_de_names = (
        'GEND_GBV_PEP TARGET: GBV PEP default',
    )
    targets_pep_short_names = (
        'TARGET: Provided with PEP',
    )
    de_targets_pep_meta = list(product(targets_pep_short_names, (None,)))
    data_element_metas += de_targets_pep_meta

    qs_targets_pep = DataValue.objects.what(*targets_pep_de_names)
    qs_targets_pep = qs_targets_pep.annotate(de_name=Value(targets_pep_short_names[0], output_field=CharField()))
    qs_targets_pep = qs_targets_pep.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets_pep = qs_targets_pep.where(filter_district)
    qs_targets_pep = qs_targets_pep.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets_pep = qs_targets_pep.when(filter_period[:4])
    qs_targets_pep = qs_targets_pep.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets_pep = qs_targets_pep.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets_pep = list(val_targets_pep)

    gen_raster = grabbag.rasterize(ou_list, de_targets_pep_meta, val_targets_pep, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets_pep2 = list(gen_raster)

    sexual_violence_female_de_names = (
        '105-1.3 OPD Abortions Due To Gender Based Violence (GBV)',
        '105-1.3 OPD Sexually Transmitted Infection Due To SGBV',
    )
    sexual_violence_female_short_names = (
        'Sexual violence (post-rape care) - Female',
    )
    de_sexual_violence_female_meta = list(product(sexual_violence_female_short_names, ('Female',)))
    data_element_metas += de_sexual_violence_female_meta

    qs_sexual_violence_female = DataValue.objects.what(*sexual_violence_female_de_names)
    qs_sexual_violence_female = qs_sexual_violence_female.annotate(de_name=Value(sexual_violence_female_short_names[0], output_field=CharField()))
    qs_sexual_violence_female = qs_sexual_violence_female.filter(category_combo__categories__name='Female')
    qs_sexual_violence_female = qs_sexual_violence_female.annotate(cat_combo=Value('Female', output_field=CharField()))
    if filter_district:
        qs_sexual_violence_female = qs_sexual_violence_female.where(filter_district)
    qs_sexual_violence_female = qs_sexual_violence_female.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_sexual_violence_female = qs_sexual_violence_female.when(filter_period)
    qs_sexual_violence_female = qs_sexual_violence_female.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_sexual_violence_female = qs_sexual_violence_female.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sexual_violence_female = list(val_sexual_violence_female)

    gen_raster = grabbag.rasterize(ou_list, de_sexual_violence_female_meta, val_sexual_violence_female, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_sexual_violence_female2 = list(gen_raster)

    sexual_violence_male_de_names = (
        '105-1.3 OPD Abortions Due To Gender Based Violence (GBV)',
        '105-1.3 OPD Sexually Transmitted Infection Due To SGBV',
    )
    sexual_violence_male_short_names = (
        'Sexual violence (post-rape care) - Male',
    )
    de_sexual_violence_male_meta = list(product(sexual_violence_male_short_names, ('Male',)))
    data_element_metas += de_sexual_violence_male_meta

    qs_sexual_violence_male = DataValue.objects.what(*sexual_violence_male_de_names)
    qs_sexual_violence_male = qs_sexual_violence_male.annotate(de_name=Value(sexual_violence_male_short_names[0], output_field=CharField()))
    qs_sexual_violence_male = qs_sexual_violence_male.filter(category_combo__categories__name='Male')
    qs_sexual_violence_male = qs_sexual_violence_male.annotate(cat_combo=Value('Male', output_field=CharField()))
    if filter_district:
        qs_sexual_violence_male = qs_sexual_violence_male.where(filter_district)
    qs_sexual_violence_male = qs_sexual_violence_male.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_sexual_violence_male = qs_sexual_violence_male.when(filter_period)
    qs_sexual_violence_male = qs_sexual_violence_male.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_sexual_violence_male = qs_sexual_violence_male.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sexual_violence_male = list(val_sexual_violence_male)

    gen_raster = grabbag.rasterize(ou_list, de_sexual_violence_male_meta, val_sexual_violence_male, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_sexual_violence_male2 = list(gen_raster)

    sexual_violence_de_names = (
        '105-1.3 OPD Abortions Due To Gender Based Violence (GBV)',
        '105-1.3 OPD Sexually Transmitted Infection Due To SGBV',
    )
    sexual_violence_short_names = (
        'Sexual violence (post-rape care) - TOTAL',
    )
    de_sexual_violence_meta = list(product(sexual_violence_short_names, (None,)))
    data_element_metas += de_sexual_violence_meta

    qs_sexual_violence = DataValue.objects.what(*sexual_violence_de_names)
    qs_sexual_violence = qs_sexual_violence.annotate(de_name=Value(sexual_violence_short_names[0], output_field=CharField()))
    qs_sexual_violence = qs_sexual_violence.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_sexual_violence = qs_sexual_violence.where(filter_district)
    qs_sexual_violence = qs_sexual_violence.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_sexual_violence = qs_sexual_violence.when(filter_period)
    qs_sexual_violence = qs_sexual_violence.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_sexual_violence = qs_sexual_violence.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_sexual_violence = list(val_sexual_violence)

    gen_raster = grabbag.rasterize(ou_list, de_sexual_violence_meta, val_sexual_violence, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_sexual_violence2 = list(gen_raster)

    gbv_care_de_names = (
        '105-1.3 OPD Sexually Transmitted Infection Due To SGBV',
    )
    gbv_care_short_names = (
        'Receiving post-GBV clinical care',
    )
    de_gbv_care_meta = list(product(gbv_care_short_names, (None,)))
    data_element_metas += de_gbv_care_meta

    qs_gbv_care = DataValue.objects.what(*gbv_care_de_names)
    qs_gbv_care = qs_gbv_care.annotate(de_name=Value(gbv_care_short_names[0], output_field=CharField()))
    qs_gbv_care = qs_gbv_care.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_gbv_care = qs_gbv_care.where(filter_district)
    qs_gbv_care = qs_gbv_care.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_gbv_care = qs_gbv_care.when(filter_period)
    qs_gbv_care = qs_gbv_care.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_gbv_care = qs_gbv_care.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_gbv_care = list(val_gbv_care)

    gen_raster = grabbag.rasterize(ou_list, de_gbv_care_meta, val_gbv_care, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_gbv_care2 = list(gen_raster)

    pep_de_names = (
        '106a PEP Q2-Number provided with PEP following - Rape/Sexual Assault or Defilement',
    )
    pep_short_names = (
        'Provided with PEP',
    )
    de_pep_meta = list(product(pep_short_names, (None,)))
    data_element_metas += de_pep_meta

    qs_pep = DataValue.objects.what(*pep_de_names)
    qs_pep = qs_pep.annotate(de_name=Value(pep_short_names[0], output_field=CharField()))
    qs_pep = qs_pep.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_pep = qs_pep.where(filter_district)
    qs_pep = qs_pep.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_pep = qs_pep.when(filter_period)
    qs_pep = qs_pep.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_pep = qs_pep.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_pep = list(val_pep)

    gen_raster = grabbag.rasterize(ou_list, de_pep_meta, val_pep, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_pep2 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_targets_care_female2, val_targets_care_male2, val_targets_pep2, val_sexual_violence_female2, val_sexual_violence_male2, val_sexual_violence2, val_gbv_care2, val_pep2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (target_gbv, target_physical, target_sexual, target_gbv_f, target_gbv_m, target_pep, sexual_f, sexual_m, sexual, gbv_care, pep, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(sexual['numeric_sum'], target_physical['numeric_sum']) and target_physical['numeric_sum']:
            perf_sexual = 100 * sexual['numeric_sum'] / target_physical['numeric_sum']
        else:
            perf_sexual = None
        perf_sexual_val = {
            'de_name': 'Perf% Sexual violence (post-rape care)',
            'cat_combo': None,
            'numeric_sum': perf_sexual,
        }
        perf_sexual_val.update(_group_ou_dict)
        calculated_vals.append(perf_sexual_val)

        if all_not_none(sexual_f['numeric_sum'], target_gbv_f['numeric_sum']) and target_gbv_f['numeric_sum']:
            perf_sexual_f = 100 * sexual_f['numeric_sum'] / target_gbv_f['numeric_sum']
        else:
            perf_sexual_f = None
        perf_sexual_f_val = {
            'de_name': 'Perf% Sexual violence',
            'cat_combo': 'Female',
            'numeric_sum': perf_sexual_f,
        }
        perf_sexual_f_val.update(_group_ou_dict)
        calculated_vals.append(perf_sexual_f_val)

        if all_not_none(sexual_m['numeric_sum'], target_gbv_m['numeric_sum']) and target_gbv_m['numeric_sum']:
            perf_sexual_m = 100 * sexual_m['numeric_sum'] / target_gbv_m['numeric_sum']
        else:
            perf_sexual_m = None
        perf_sexual_m_val = {
            'de_name': 'Perf% Sexual violence',
            'cat_combo': 'Male',
            'numeric_sum': perf_sexual_m,
        }
        perf_sexual_m_val.update(_group_ou_dict)
        calculated_vals.append(perf_sexual_m_val)

        if all_not_none(gbv_care['numeric_sum'], target_gbv['numeric_sum']) and target_gbv['numeric_sum']:
            perf_gbv_care = 100 * gbv_care['numeric_sum'] / target_gbv['numeric_sum']
        else:
            perf_gbv_care = None
        perf_gbv_care_val = {
            'de_name': 'Perf% Receiving post-GBV clinical care',
            'cat_combo': None,
            'numeric_sum': perf_gbv_care,
        }
        perf_gbv_care_val.update(_group_ou_dict)
        calculated_vals.append(perf_gbv_care_val)

        if all_not_none(pep['numeric_sum'], target_pep['numeric_sum']) and target_pep['numeric_sum']:
            perf_pep = 100 * pep['numeric_sum'] / target_pep['numeric_sum']
        else:
            perf_pep = None
        perf_pep_val = {
            'de_name': 'Perf% PEP',
            'cat_combo': None,
            'numeric_sum': perf_pep,
        }
        perf_pep_val.update(_group_ou_dict)
        calculated_vals.append(perf_pep_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Perf% Sexual violence (post-rape care)'], (None,)))
    data_element_metas += list(product(['Perf% Sexual violence'], ('Female',)))
    data_element_metas += list(product(['Perf% Sexual violence'], ('Male',)))
    data_element_metas += list(product(['Perf% Receiving post-GBV clinical care'], (None,)))
    data_element_metas += list(product(['Perf% PEP'], (None,)))
    


    num_path_elements = len(ou_headers)
    legend_sets = list()
    gbv_ls = LegendSet()
    gbv_ls.name = 'GBV Cases'
    gbv_ls.add_interval('orange', 0, 25)
    gbv_ls.add_interval('yellow', 25, 40)
    gbv_ls.add_interval('light-green', 40, 60)
    gbv_ls.add_interval('green', 60, None)
    gbv_ls.mappings[num_path_elements+11] = True
    gbv_ls.mappings[num_path_elements+12] = True
    gbv_ls.mappings[num_path_elements+13] = True
    gbv_ls.mappings[num_path_elements+14] = True
    gbv_ls.mappings[num_path_elements+15] = True
    legend_sets.append(gbv_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="gbv_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="gbv_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/gbv_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def sc_mos_by_site(request, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    month_years = zip([((this_day.month-i-1)%12)+1 for i in range(5*12)], ([this_day.year] * this_day.month) + sorted([this_day.year-i for i in range(1, 5)]*12, reverse=True))
    PREV_5YR_MONTHS = ['{0}-{1:02}'.format(y, m) for m, y in month_years]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_MONTHS:
        filter_period=request.GET['period']
    else:
        filter_period = '{0}-{1:02}'.format(this_year, this_day.month)

    period_desc = filter_period #dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=3).annotate(district=F('parent__parent__name'), subcounty=F('parent__name'), facility=F('name'))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by('district', 'subcounty', 'facility')
    ou_list = list(qs_ou.values_list('district', 'subcounty', 'facility'))
    ou_headers = ['District', 'Subcounty', 'Facility']

    def val_with_subcat_fun(row, col):
        district, subcounty, facility = row
        de_name, subcategory = col
        return { 'district': district, 'subcounty': subcounty, 'facility': facility, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }

    supply_names = [
        '105-6  Zidovudine /Lamivudine/Nevirapine (AZT/3TC/NVP)',
        '105-6 (RHZE) blister strip 150/75/400/275 mg',
        '105-6 Abacavir/Lamivudine (ABC/3TC) 60mg/30mg (Paediatric)',
        '105-6 Amoxicillin dispersible 125mg tablet (For children)',
        '105-6 Artemether/ Lumefantrine 100/20mg tablet',
        '105-6 Bendrofulazide (Aprinox) 5mg',
        '105-6 Blood 450 ml',
        '105-6 CD4 reagent Specify',
        '105-6 Captopril 25mg tablet',
        '105-6 Cardiac Aspirin 75/80 mg',
        '105-6 Ceftriaxone 1g Injection',
        '105-6 Chlorhexidine 20%',
        '105-6 Co-tromoxazole 480mg tablet',
        '105-6 Cotrimoxazole 960mg tablet',
        '105-6 Determine HIV Screening test, tests',
        '105-6 Efavirenz (EFV) 600mg',
        '105-6 Glibenclamide 5mg tablet',
        '105-6 Insulin short-acting',
        '105-6 Mama Kit',
        '105-6 Measles Vaccine',
        '105-6 Metformin 500mg',
        '105-6 Misoprostol 200mcg Tablet',
        '105-6 Nevirapine (NVP) 200mg',
        '105-6 Nevirapine (NVP) 50mg',
        '105-6 Nifedipine tablets 20mg tablet',
        '105-6 ORS Sachets with zinc tablet',
        '105-6 Oxytocin Injection',
        '105-6 Propranolol 40mg tablet',
        '105-6 RH blister strip 150/75 mg',
        '105-6 Ready to use Therapeutic feeds (RUTF)',
        '105-6 Stat-pack HIV Confirmatory rapid tests, tests',
        '105-6 Sulfadoxine / Pyrimethamine tablet',
        '105-6 Tenofovir/Lamivudine (TDF/3TC) 300mg/300mg',
        '105-6 Tenofovir/Lamivudine/Efavirenz (TDF/3TC/EFV) 300mg/300mg/',
        '105-6 Therapeutic milk F100 (100Kcal/100ml)',
        '105-6 Therapeutic milk F75 (75Kcal/100ml)',
        '105-6 Unigold HIV RDT Tie-breaker test, tests',
        '105-6 ZN reagent for AFB',
        '105-6 Zidovudine/Lamivudine (AZT/3TC) 300mg/150m'
    ]
    stock_de_names = ((s+' Days out of stock', s+' Quantity Utilized', s+' Stock at Hand') for s in supply_names)
    stock_de_names = list(chain.from_iterable(stock_de_names)) # flatten the list of tuples of strings into a list of strings
    de_stock_meta = list(product(stock_de_names, (None,)))

    qs_stock = DataValue.objects.what(*stock_de_names).filter(month=filter_period)
    if filter_district:
        qs_stock = qs_stock.where(filter_district)
    qs_stock = qs_stock.annotate(cat_combo=Value(None, output_field=CharField()))
    # qs_stock = qs_stock.annotate(cat_combo=F('category_combo__name'))

    qs_stock = qs_stock.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_stock = qs_stock.annotate(period=F('quarter'))
    qs_stock = qs_stock.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_stock = qs_stock.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_stock = list(val_stock)

    gen_raster = grabbag.rasterize(ou_list, de_stock_meta, val_stock, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_stock2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_stock2), key=lambda x: (x['district'], x['subcounty'], x['facility'])), key=lambda x: (x['district'], x['subcounty'], x['facility']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    calc_names = OrderedDict()
    # perform calculations
    for _group in grouped_vals:
        (district_subcounty_facility, (*stock_vals,)) = _group
        days_out, utilized, on_hand, *other_vals = stock_vals
        
        calculated_vals = list()
        AVG_MONTH_DAYS = 30 # days in month

        for days_out, utilized, on_hand in grouper(stock_vals, 3):
            if all_not_none(utilized['numeric_sum'], days_out['numeric_sum']) and days_out['numeric_sum'] < AVG_MONTH_DAYS:
                avg_consumption = utilized['numeric_sum'] * (AVG_MONTH_DAYS / (AVG_MONTH_DAYS - days_out['numeric_sum']))
            else:
                avg_consumption = None
            # calc_name = on_hand['de_name'].replace('Stock at Hand', 'Adjusted Consumption')
            # calc_names[calc_name] = True
            # avg_consumption_val = {
            #     'district': district_subcounty_facility[0],
            #     'subcounty': district_subcounty_facility[1],
            #     'facility': district_subcounty_facility[2],
            #     'de_name': calc_name,
            #     'cat_combo': None,
            #     'numeric_sum': avg_consumption,
            # }
            # calculated_vals.append(avg_consumption_val)

            if all_not_none(on_hand['numeric_sum']) and avg_consumption and on_hand['numeric_sum'] > 0:
                months_of_stock = on_hand['numeric_sum']/(avg_consumption)
            else:
                months_of_stock = None
                if on_hand['numeric_sum']:
                    months_of_stock = -on_hand['numeric_sum']
            calc_name = on_hand['de_name'].replace('Stock at Hand', 'Months of Stock')
            calc_names[calc_name] = True
            sc_mos_val = {
                'district': district_subcounty_facility[0],
                'subcounty': district_subcounty_facility[1],
                'facility': district_subcounty_facility[2],
                'de_name': calc_name,
                'cat_combo': None,
                'numeric_sum': months_of_stock,
            }
            calculated_vals.append(sc_mos_val)

        if True:
            _group[1] = calculated_vals
        else:
            _group[1].extend(calculated_vals)

    data_element_names = list()
    if False:
        data_element_names += de_stock_meta
    
    data_element_names.extend([(c, None) for c in calc_names])

    mos_base_index = len(ou_headers)
    if False:
        mos_base_index += len(de_stock_meta)
    legend_sets = list()
    sc_mos_ls = LegendSet()
    sc_mos_ls.name = 'Months of Stock (MOS)'
    sc_mos_ls.add_interval('red', 0, 2)
    sc_mos_ls.add_interval('green', 2, 4)
    sc_mos_ls.add_interval('yellow', 4, None)
    for i in range(len(supply_names)):
        if False:
            sc_mos_ls.mappings[mos_base_index+(i*2)] = True
        else:
            sc_mos_ls.mappings[mos_base_index+(i)] = True
    legend_sets.append(sc_mos_ls)
    sc_soh_ls = LegendSet()
    sc_soh_ls.name = 'Stock on Hand (SOH): invalid MOS'
    sc_soh_ls.add_interval('light-green', None, 0)
    for i in range(len(supply_names)):
        if False:
            sc_soh_ls.mappings[mos_base_index+(i*2)] = True
        else:
            sc_soh_ls.mappings[mos_base_index+(i)] = True
    legend_sets.append(sc_soh_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_names))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sc_mos_sites_scorecard.csv"'

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_names)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="sc_mos_sites_scorecard.xlsx"'

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_names,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_MONTHS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        #TODO: this doesn't work if you have more than one LegendSet mapped to the exact same columns
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/sc_mos_sites.html', context)

@login_required
def art_new_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    target_all_de_names = (
        'TX_NEW (N, DSD) TARGET: New on ART default',
    )
    target_all_short_names = (
        'TARGET: New on ART',
    )
    de_target_all_meta = list(product(target_all_de_names, (None,)))
    data_element_metas += list(product(target_all_short_names, (None,)))

    qs_target_all = DataValue.objects.what(*target_all_de_names)
    qs_target_all = qs_target_all.annotate(cat_combo=Value(None, output_field=CharField()))
    # qs_target_all = qs_target_all.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_target_all = qs_target_all.where(filter_district)
    qs_target_all = qs_target_all.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # target_all are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_target_all = qs_target_all.when(filter_period[:4])
    qs_target_all = qs_target_all.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_target_all = qs_target_all.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_target_all = list(val_target_all)

    gen_raster = grabbag.rasterize(ou_list, de_target_all_meta, val_target_all, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_target_all2 = list(gen_raster)

    subcategory_names = ('(15+, Female)', '(15+, Male)', '(<15, Female)', '(<15, Male)')
    subcategory_names2 = ('(<1, Female)', '(<1, Male)', '(1-9, Female)', '(1-9, Male)', '(10-14, Female)', '(10-14, Male)', '(15+, Female)', '(15+, Male)')
    cc_lt_15 = ['<2 Years', '2 - < 5 Years (HIV Care)', '5 - 14 Years']
    cc_ge_15 = ['15 Years and above']
    
    targets_de_names = (
        'TX_NEW (N, Aggregated Age/Sex) TARGET: HIV Prevention Program',
    )
    targets_short_names = (
        'TARGET: New on ART',
    )
    de_targets_meta = list(product(targets_de_names, subcategory_names))
    data_element_metas += list(product(targets_short_names, subcategory_names))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=F('category_combo__name'))
    qs_targets = qs_targets.filter(category_combo__name__in=subcategory_names)
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    art_new_de_names = (
        '106a ART No. of new clients started on ART at this facility during the quarter',
    )
    art_new_short_names = (
        'New on ART',
    )
    de_art_new_meta = list(product(art_new_short_names, (None,)))
    data_element_metas += de_art_new_meta

    qs_art_new = DataValue.objects.what(*art_new_de_names)
    qs_art_new = qs_art_new.annotate(de_name=Value(art_new_short_names[0], output_field=CharField()))
    qs_art_new = qs_art_new.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_art_new = qs_art_new.where(filter_district)
    qs_art_new = qs_art_new.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_new = qs_art_new.when(filter_period)
    qs_art_new = qs_art_new.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_new = qs_art_new.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_new = list(val_art_new)

    gen_raster = grabbag.rasterize(ou_list, de_art_new_meta, val_art_new, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_new2 = list(gen_raster)
    
    art_new_gt_15_de_names = (
        '106a ART No. of new clients started on ART at this facility during the quarter',
    )
    art_new_gt_15_short_names = (
        'New on ART',
    )
    de_art_new_gt_15_meta = list(product(art_new_gt_15_short_names, subcategory_names[:2]))
    data_element_metas += de_art_new_gt_15_meta

    qs_art_new_gt_15 = DataValue.objects.what(*art_new_gt_15_de_names)
    qs_art_new_gt_15 = qs_art_new_gt_15.annotate(de_name=Value(art_new_gt_15_short_names[0], output_field=CharField()))
    qs_art_new_gt_15 = qs_art_new_gt_15.filter(Q(category_combo__categories__name='15 Years and above'))
    qs_art_new_gt_15 = qs_art_new_gt_15.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            # When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            # When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    if filter_district:
        qs_art_new_gt_15 = qs_art_new_gt_15.where(filter_district)
    qs_art_new_gt_15 = qs_art_new_gt_15.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_new_gt_15 = qs_art_new_gt_15.when(filter_period)
    qs_art_new_gt_15 = qs_art_new_gt_15.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_new_gt_15 = qs_art_new_gt_15.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_new_gt_15 = list(val_art_new_gt_15)

    gen_raster = grabbag.rasterize(ou_list, de_art_new_gt_15_meta, val_art_new_gt_15, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_new_gt_152 = list(gen_raster)
    
    art_new_lt_15_de_names = (
        '106a ART No. of new clients started on ART at this facility during the quarter',
    )
    art_new_lt_15_short_names = (
        'New on ART',
    )
    de_art_new_lt_15_meta = list(product(art_new_lt_15_short_names, subcategory_names[2:]))
    data_element_metas += de_art_new_lt_15_meta

    qs_art_new_lt_15 = DataValue.objects.what(*art_new_lt_15_de_names)
    qs_art_new_lt_15 = qs_art_new_lt_15.annotate(de_name=Value(art_new_lt_15_short_names[0], output_field=CharField()))
    qs_art_new_lt_15 = qs_art_new_lt_15.filter(Q(category_combo__categories__name='<2 Years')|Q(category_combo__categories__name='2 - < 5 Years (HIV Care)')|Q(category_combo__categories__name='5 - 14 Years'))
    qs_art_new_lt_15 = qs_art_new_lt_15.annotate(
        cat_combo=Case(
            # When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            # When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    if filter_district:
        qs_art_new_lt_15 = qs_art_new_lt_15.where(filter_district)
    qs_art_new_lt_15 = qs_art_new_lt_15.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_new_lt_15 = qs_art_new_lt_15.when(filter_period)
    qs_art_new_lt_15 = qs_art_new_lt_15.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_new_lt_15 = qs_art_new_lt_15.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_new_lt_15 = list(val_art_new_lt_15)

    gen_raster = grabbag.rasterize(ou_list, de_art_new_lt_15_meta, val_art_new_lt_15, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_new_lt_152 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_target_all2, val_targets2, val_art_new2, val_art_new_gt_152, val_art_new_lt_152), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (target_all, target_over15_f, target_over15_m, target_under15_f, target_under15_m, art_new, art_new_over15_f, art_new_over15_m, art_new_under15_f, art_new_under15_m, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(art_new['numeric_sum'], target_all['numeric_sum']) and target_all['numeric_sum']:
            perf_art_new = 100 * art_new['numeric_sum'] / target_all['numeric_sum']
        else:
            perf_art_new = None
        perf_art_new_val = {
            'de_name': 'Perf% New on ART',
            'cat_combo': None,
            'numeric_sum': perf_art_new,
        }
        perf_art_new_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_new_val)

        if all_not_none(art_new_over15_f['numeric_sum'], target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            perf_art_new_over15_f = 100 * art_new_over15_f['numeric_sum'] / target_over15_f['numeric_sum']
        else:
            perf_art_new_over15_f = None
        perf_art_new_over15_f_val = {
            'de_name': 'Perf% New on ART',
            'cat_combo': '(15+, Female)',
            'numeric_sum': perf_art_new_over15_f,
        }
        perf_art_new_over15_f_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_new_over15_f_val)

        if all_not_none(art_new_over15_m['numeric_sum'], target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            perf_art_new_over15_m = 100 * art_new_over15_m['numeric_sum'] / target_over15_m['numeric_sum']
        else:
            perf_art_new_over15_m = None
        perf_art_new_over15_m_val = {
            'de_name': 'Perf% New on ART',
            'cat_combo': '(15+, Male)',
            'numeric_sum': perf_art_new_over15_m,
        }
        perf_art_new_over15_m_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_new_over15_m_val)

        if all_not_none(art_new_under15_f['numeric_sum'], target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            perf_art_new_under15_f = 100 * art_new_under15_f['numeric_sum'] / target_under15_f['numeric_sum']
        else:
            perf_art_new_under15_f = None
        perf_art_new_under15_f_val = {
            'de_name': 'Perf% New on ART',
            'cat_combo': '(<15, Female)',
            'numeric_sum': perf_art_new_under15_f,
        }
        perf_art_new_under15_f_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_new_under15_f_val)

        if all_not_none(art_new_under15_m['numeric_sum'], target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            perf_art_new_under15_m = 100 * art_new_under15_m['numeric_sum'] / target_under15_m['numeric_sum']
        else:
            perf_art_new_under15_m = None
        perf_art_new_under15_m_val = {
            'de_name': 'Perf% New on ART',
            'cat_combo': '(<15, Male)',
            'numeric_sum': perf_art_new_under15_m,
        }
        perf_art_new_under15_m_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_new_under15_m_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Perf% New on ART'], (None,)))
    data_element_metas += list(product(['Perf% New on ART'], ('(15+, Female)',)))
    data_element_metas += list(product(['Perf% New on ART'], ('(15+, Male)',)))
    data_element_metas += list(product(['Perf% New on ART'], ('(<15, Female)',)))
    data_element_metas += list(product(['Perf% New on ART'], ('(<15, Male)',)))


    num_path_elements = len(ou_headers)
    legend_sets = list()
    art_new_ls = LegendSet()
    art_new_ls.name = 'New on ART'
    art_new_ls.add_interval('orange', 0, 25)
    art_new_ls.add_interval('yellow', 25, 40)
    art_new_ls.add_interval('light-green', 40, 60)
    art_new_ls.add_interval('green', 60, None)
    art_new_ls.mappings[num_path_elements+10] = True
    art_new_ls.mappings[num_path_elements+11] = True
    art_new_ls.mappings[num_path_elements+12] = True
    art_new_ls.mappings[num_path_elements+13] = True
    art_new_ls.mappings[num_path_elements+14] = True
    legend_sets.append(art_new_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="art_new_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="art_new_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/art_new_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def art_active_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    target_all_de_names = (
        'TX_CURR (N, DSD) TARGET: Receiving ART default',
    )
    target_all_short_names = (
        'TARGET: Active on ART',
    )
    de_target_all_meta = list(product(target_all_de_names, (None,)))
    data_element_metas += list(product(target_all_short_names, (None,)))

    qs_target_all = DataValue.objects.what(*target_all_de_names)
    qs_target_all = qs_target_all.annotate(cat_combo=Value(None, output_field=CharField()))
    # qs_target_all = qs_target_all.annotate(cat_combo=F('category_combo__name'))
    if filter_district:
        qs_target_all = qs_target_all.where(filter_district)
    qs_target_all = qs_target_all.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, but this is a cumulative target, so filter by year component of period and *DO NOT* divide result by 4 to get quarter
    qs_target_all = qs_target_all.when(filter_period[:4])
    qs_target_all = qs_target_all.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_target_all = qs_target_all.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_target_all = list(val_target_all)

    gen_raster = grabbag.rasterize(ou_list, de_target_all_meta, val_target_all, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_target_all2 = list(gen_raster)

    subcategory_names = ('(15+, Female)', '(15+, Male)', '(<15, Female)', '(<15, Male)')
    subcategory_names2 = ('(<1, Female)', '(<1, Male)', '(1-9, Female)', '(1-9, Male)', '(10-14, Female)', '(10-14, Male)', '(15+, Female)', '(15+, Male)')
    cc_lt_15 = ['<2 Years', '2 - < 5 Years (HIV Care)', '5 - 14 Years']
    cc_ge_15 = ['15 Years and above']
    
    targets_de_names = (
        'TX_CURR (N, Aggregated Age/Sex) TARGET: Receiving ART',
    )
    targets_short_names = (
        'TARGET: Active on ART',
    )
    de_targets_meta = list(product(targets_de_names, subcategory_names))
    data_element_metas += list(product(targets_short_names, subcategory_names))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=F('category_combo__name'))
    qs_targets = qs_targets.filter(category_combo__name__in=subcategory_names)
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    # targets are annual, but this is a cumulative target, so filter by year component of period and *DO NOT* divide result by 4 to get quarter
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    art_active_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    art_active_short_names = (
        'Active on ART',
    )
    de_art_active_meta = list(product(art_active_short_names, (None,)))
    data_element_metas += de_art_active_meta

    qs_art_active = DataValue.objects.what(*art_active_de_names)
    qs_art_active = qs_art_active.annotate(de_name=Value(art_active_short_names[0], output_field=CharField()))
    qs_art_active = qs_art_active.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_art_active = qs_art_active.where(filter_district)
    qs_art_active = qs_art_active.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_active = qs_art_active.when(filter_period)
    qs_art_active = qs_art_active.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_active = qs_art_active.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_active = list(val_art_active)

    gen_raster = grabbag.rasterize(ou_list, de_art_active_meta, val_art_active, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_active2 = list(gen_raster)
    
    art_active_gt_15_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    art_active_gt_15_short_names = (
        'Active on ART',
    )
    de_art_active_gt_15_meta = list(product(art_active_gt_15_short_names, subcategory_names[:2]))
    data_element_metas += de_art_active_gt_15_meta

    qs_art_active_gt_15 = DataValue.objects.what(*art_active_gt_15_de_names)
    qs_art_active_gt_15 = qs_art_active_gt_15.annotate(de_name=Value(art_active_gt_15_short_names[0], output_field=CharField()))
    qs_art_active_gt_15 = qs_art_active_gt_15.filter(Q(category_combo__categories__name='15 Years and above'))
    qs_art_active_gt_15 = qs_art_active_gt_15.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            # When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            # When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    if filter_district:
        qs_art_active_gt_15 = qs_art_active_gt_15.where(filter_district)
    qs_art_active_gt_15 = qs_art_active_gt_15.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_active_gt_15 = qs_art_active_gt_15.when(filter_period)
    qs_art_active_gt_15 = qs_art_active_gt_15.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_active_gt_15 = qs_art_active_gt_15.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_active_gt_15 = list(val_art_active_gt_15)

    gen_raster = grabbag.rasterize(ou_list, de_art_active_gt_15_meta, val_art_active_gt_15, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_active_gt_152 = list(gen_raster)
    
    art_active_lt_15_de_names = (
        '106a ART No. active on ART on 1st line ARV regimen',
        '106a ART No. active on ART on 2nd line ARV regimen',
        '106a ART No. active on ART on 3rd line or higher ARV regimen',
    )
    art_active_lt_15_short_names = (
        'Active on ART',
    )
    de_art_active_lt_15_meta = list(product(art_active_lt_15_short_names, subcategory_names[2:]))
    data_element_metas += de_art_active_lt_15_meta

    qs_art_active_lt_15 = DataValue.objects.what(*art_active_lt_15_de_names)
    qs_art_active_lt_15 = qs_art_active_lt_15.annotate(de_name=Value(art_active_lt_15_short_names[0], output_field=CharField()))
    qs_art_active_lt_15 = qs_art_active_lt_15.filter(Q(category_combo__categories__name='<2 Years')|Q(category_combo__categories__name='2 - < 5 Years (HIV Care)')|Q(category_combo__categories__name='5 - 14 Years'))
    qs_art_active_lt_15 = qs_art_active_lt_15.annotate(
        cat_combo=Case(
            # When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            # When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    if filter_district:
        qs_art_active_lt_15 = qs_art_active_lt_15.where(filter_district)
    qs_art_active_lt_15 = qs_art_active_lt_15.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_art_active_lt_15 = qs_art_active_lt_15.when(filter_period)
    qs_art_active_lt_15 = qs_art_active_lt_15.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_art_active_lt_15 = qs_art_active_lt_15.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_art_active_lt_15 = list(val_art_active_lt_15)

    gen_raster = grabbag.rasterize(ou_list, de_art_active_lt_15_meta, val_art_active_lt_15, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_art_active_lt_152 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_target_all2, val_targets2, val_art_active2, val_art_active_gt_152, val_art_active_lt_152), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (target_all, target_over15_f, target_over15_m, target_under15_f, target_under15_m, art_active, art_active_over15_f, art_active_over15_m, art_active_under15_f, art_active_under15_m, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(art_active['numeric_sum'], target_all['numeric_sum']) and target_all['numeric_sum']:
            perf_art_active = 100 * art_active['numeric_sum'] / target_all['numeric_sum']
        else:
            perf_art_active = None
        perf_art_active_val = {
            'de_name': 'Perf% Active on ART',
            'cat_combo': None,
            'numeric_sum': perf_art_active,
        }
        perf_art_active_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_active_val)

        if all_not_none(art_active_over15_f['numeric_sum'], target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            perf_art_active_over15_f = 100 * art_active_over15_f['numeric_sum'] / target_over15_f['numeric_sum']
        else:
            perf_art_active_over15_f = None
        perf_art_active_over15_f_val = {
            'de_name': 'Perf% Active on ART',
            'cat_combo': '(15+, Female)',
            'numeric_sum': perf_art_active_over15_f,
        }
        perf_art_active_over15_f_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_active_over15_f_val)

        if all_not_none(art_active_over15_m['numeric_sum'], target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            perf_art_active_over15_m = 100 * art_active_over15_m['numeric_sum'] / target_over15_m['numeric_sum']
        else:
            perf_art_active_over15_m = None
        perf_art_active_over15_m_val = {
            'de_name': 'Perf% Active on ART',
            'cat_combo': '(15+, Male)',
            'numeric_sum': perf_art_active_over15_m,
        }
        perf_art_active_over15_m_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_active_over15_m_val)

        if all_not_none(art_active_under15_f['numeric_sum'], target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            perf_art_active_under15_f = 100 * art_active_under15_f['numeric_sum'] / target_under15_f['numeric_sum']
        else:
            perf_art_active_under15_f = None
        perf_art_active_under15_f_val = {
            'de_name': 'Perf% Active on ART',
            'cat_combo': '(<15, Female)',
            'numeric_sum': perf_art_active_under15_f,
        }
        perf_art_active_under15_f_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_active_under15_f_val)

        if all_not_none(art_active_under15_m['numeric_sum'], target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            perf_art_active_under15_m = 100 * art_active_under15_m['numeric_sum'] / target_under15_m['numeric_sum']
        else:
            perf_art_active_under15_m = None
        perf_art_active_under15_m_val = {
            'de_name': 'Perf% Active on ART',
            'cat_combo': '(<15, Male)',
            'numeric_sum': perf_art_active_under15_m,
        }
        perf_art_active_under15_m_val.update(_group_ou_dict)
        calculated_vals.append(perf_art_active_under15_m_val)

        _group[1].extend(calculated_vals)

    data_element_metas += list(product(['Perf% Active on ART'], (None,)))
    data_element_metas += list(product(['Perf% Active on ART'], ('(15+, Female)',)))
    data_element_metas += list(product(['Perf% Active on ART'], ('(15+, Male)',)))
    data_element_metas += list(product(['Perf% Active on ART'], ('(<15, Female)',)))
    data_element_metas += list(product(['Perf% Active on ART'], ('(<15, Male)',)))


    num_path_elements = len(ou_headers)
    legend_sets = list()
    art_active_ls = LegendSet()
    art_active_ls.name = 'Active on ART'
    art_active_ls.add_interval('orange', 0, 25)
    art_active_ls.add_interval('yellow', 25, 40)
    art_active_ls.add_interval('light-green', 40, 60)
    art_active_ls.add_interval('green', 60, None)
    art_active_ls.mappings[num_path_elements+10] = True
    art_active_ls.mappings[num_path_elements+11] = True
    art_active_ls.mappings[num_path_elements+12] = True
    art_active_ls.mappings[num_path_elements+13] = True
    art_active_ls.mappings[num_path_elements+14] = True
    legend_sets.append(art_active_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="art_active_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="art_active_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/art_active_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def mnch_preg_birth_scorecard(request, org_unit_level=2, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }
    # annotations for data collected at subcounty level
    SUBCOUNTY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(2, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'Catchment Population',
    )
    targets_short_names = (
        'Catchment Population',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))
    data_element_metas += list(product(targets_short_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**SUBCOUNTY_LEVEL_ANNOTATIONS)
    # population estimates are annual, so filter by year component of period
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    anc_de_names = (
        '105-2.1 A17:HIV+ Pregnant Women already on ART before 1st ANC (ART-K)',
        '105-2.1 A1:ANC 1st Visit for women',
        '105-2.1 A1:ANC 1st Visit for women (No. in 1st Trimester)',
        '105-2.1 A2:ANC 4th Visit for women',
        '105-2.1 A6:First dose IPT (IPT1)',
        '105-2.1 A7:Second dose IPT (IPT2)',
        '105-2.1 HIV+ Pregnant Women initiated on ART for EMTCT (ART)',
        '105-2.2 HIV+ women initiating ART in maternity',
        '105-2.2 OPD Maternal deaths',
        '105-2.2a Deliveries in unit',
        '105-2.3 HIV+ women initiating ART in PNC',
        '105-2.3 Vitamin A supplementation given to mothers',
        '108-3 MSP Caesarian Sections',
    )
    anc_short_names = (
        # empty, no shortnames needed
    )
    de_anc_meta = list(product(anc_de_names, (None,)))
    data_element_metas += de_anc_meta

    qs_anc = DataValue.objects.what(*anc_de_names)
    qs_anc = qs_anc.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_anc = qs_anc.where(filter_district)
    qs_anc = qs_anc.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_anc = qs_anc.when(filter_period)
    qs_anc = qs_anc.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_anc = qs_anc.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_anc = list(val_anc)

    gen_raster = grabbag.rasterize(ou_list, de_anc_meta, val_anc, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_anc2 = list(gen_raster)

    anc_adolescent_de_names = (
        '105-2.1 A1:ANC 1st Visit for women',
    )
    anc_adolescent_short_names = (
        # empty, no shortnames needed
    )
    de_anc_adolescent_meta = list(product(anc_adolescent_de_names, (None,)))
    data_element_metas += de_anc_adolescent_meta

    qs_anc_adolescent = DataValue.objects.what(*anc_adolescent_de_names)
    qs_anc_adolescent = qs_anc_adolescent.filter(category_combo__categories__name='10-19 Years')
    qs_anc_adolescent = qs_anc_adolescent.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_anc_adolescent = qs_anc_adolescent.where(filter_district)
    qs_anc_adolescent = qs_anc_adolescent.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_anc_adolescent = qs_anc_adolescent.when(filter_period)
    qs_anc_adolescent = qs_anc_adolescent.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_anc_adolescent = qs_anc_adolescent.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_anc_adolescent = list(val_anc_adolescent)

    gen_raster = grabbag.rasterize(ou_list, de_anc_adolescent_meta, val_anc_adolescent, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_anc_adolescent2 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_anc2, val_anc_adolescent2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (catchment_pop, anc_already_art, anc1, anc1_1st_trimester, anc4, ipt1, ipt2, anc_started_art, maternity_started_art, maternal_deaths, deliveries, pnc_started_art, vit_a_maternity, caesarian, anc1_adolescent, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(catchment_pop['numeric_sum']):
            expected_pregnant = (catchment_pop['numeric_sum'] * Decimal(0.05))/4 # split by quarter
        else:
            expected_pregnant = None
        expected_pregnant_val = {
            'de_name': 'Expected Pregnancies',
            'cat_combo': None,
            'numeric_sum': expected_pregnant,
        }
        expected_pregnant_val.update(_group_ou_dict)
        calculated_vals.append(expected_pregnant_val)

        if all_not_none(catchment_pop['numeric_sum']):
            adolescent_pop = catchment_pop['numeric_sum'] * Decimal(0.128)
        else:
            adolescent_pop = None
        adolescent_pop_val = {
            'de_name': 'Adolescent Population',
            'cat_combo': None,
            'numeric_sum': adolescent_pop,
        }
        adolescent_pop_val.update(_group_ou_dict)
        calculated_vals.append(adolescent_pop_val)

        if all_not_none(catchment_pop['numeric_sum']):
            expected_pregnant_hiv = (catchment_pop['numeric_sum'] * Decimal(0.0485) * Decimal(0.058))/4 # split by quarter
        else:
            expected_pregnant_hiv = None
        expected_pregnant_hiv_val = {
            'de_name': 'All expected pregnancies in a catchment population multiplied by HIV prevalence',
            'cat_combo': None,
            'numeric_sum': expected_pregnant_hiv,
        }
        expected_pregnant_hiv_val.update(_group_ou_dict)
        calculated_vals.append(expected_pregnant_hiv_val)

        if all_not_none(catchment_pop['numeric_sum']):
            expected_deliver = (catchment_pop['numeric_sum'] * Decimal(0.0485))/4 # split by quarter
        else:
            expected_deliver = None
        expected_deliver_val = {
            'de_name': 'Expected Deliveries',
            'cat_combo': None,
            'numeric_sum': expected_deliver,
        }
        expected_deliver_val.update(_group_ou_dict)
        calculated_vals.append(expected_deliver_val)

        if all_not_none(anc1['numeric_sum'], expected_pregnant) and expected_pregnant:
            anc1_percent = 100 * anc1['numeric_sum'] / expected_pregnant
        else:
            anc1_percent = None
        anc1_percent_val = {
            'de_name': '% ANC1 Attendance coverage---Target=90%',
            'cat_combo': None,
            'numeric_sum': anc1_percent,
        }
        anc1_percent_val.update(_group_ou_dict)
        calculated_vals.append(anc1_percent_val)

        if all_not_none(anc1_1st_trimester['numeric_sum'], expected_pregnant) and expected_pregnant:
            anc1_1st_trimester_percent = 100 * anc1_1st_trimester['numeric_sum'] / expected_pregnant
        else:
            anc1_1st_trimester_percent = None
        anc1_1st_trimester_percent_val = {
            'de_name': '% of pregnant women attending 1st ANC visit within the 1st trimester---Target=45%',
            'cat_combo': None,
            'numeric_sum': anc1_1st_trimester_percent,
        }
        anc1_1st_trimester_percent_val.update(_group_ou_dict)
        calculated_vals.append(anc1_1st_trimester_percent_val)

        if all_not_none(anc1_adolescent['numeric_sum'], adolescent_pop) and adolescent_pop:
            anc1_adolescent_percent = 100 * anc1_adolescent['numeric_sum'] / adolescent_pop
        else:
            anc1_adolescent_percent = None
        anc1_adolescent_percent_val = {
            'de_name': 'Adolescent  pregnancy rate (10 -19 years of age)  Target--<5%',
            'cat_combo': None,
            'numeric_sum': anc1_adolescent_percent,
        }
        anc1_adolescent_percent_val.update(_group_ou_dict)
        calculated_vals.append(anc1_adolescent_percent_val)

        if all_not_none(anc4['numeric_sum'], expected_pregnant) and expected_pregnant:
            anc4_percent = 100 * anc4['numeric_sum'] / expected_pregnant
        else:
            anc4_percent = None
        anc4_percent_val = {
            'de_name': '% ANC4 Attendance coverage---Target=60%',
            'cat_combo': None,
            'numeric_sum': anc4_percent,
        }
        anc4_percent_val.update(_group_ou_dict)
        calculated_vals.append(anc4_percent_val)

        if all_not_none(ipt1['numeric_sum'], expected_pregnant) and expected_pregnant:
            ipt1_percent = 100 * ipt1['numeric_sum'] / expected_pregnant
        else:
            ipt1_percent = None
        ipt1_percent_val = {
            'de_name': 'IPT1 Coverage--Target=90%',
            'cat_combo': None,
            'numeric_sum': ipt1_percent,
        }
        ipt1_percent_val.update(_group_ou_dict)
        calculated_vals.append(ipt1_percent_val)

        if all_not_none(ipt2['numeric_sum'], expected_pregnant) and expected_pregnant:
            ipt2_percent = 100 * ipt2['numeric_sum'] / expected_pregnant
        else:
            ipt2_percent = None
        ipt2_percent_val = {
            'de_name': 'IPT2 Coverage--Target=90%',
            'cat_combo': None,
            'numeric_sum': ipt2_percent,
        }
        ipt2_percent_val.update(_group_ou_dict)
        calculated_vals.append(ipt2_percent_val)

        if all_not_none(vit_a_maternity['numeric_sum'], deliveries['numeric_sum']) and deliveries['numeric_sum']:
            vit_a_maternity_percent = 100 * vit_a_maternity['numeric_sum'] / deliveries['numeric_sum']
        else:
            vit_a_maternity_percent = None
        vit_a_maternity_percent_val = {
            'de_name': 'VITA Supplementation for mothers--Target =90%',
            'cat_combo': None,
            'numeric_sum': vit_a_maternity_percent,
        }
        vit_a_maternity_percent_val.update(_group_ou_dict)
        calculated_vals.append(vit_a_maternity_percent_val)

        if all_not_none(maternal_deaths['numeric_sum'], deliveries['numeric_sum']) and deliveries['numeric_sum']:
            maternal_deaths_percent = 100 * maternal_deaths['numeric_sum'] / deliveries['numeric_sum']
        else:
            maternal_deaths_percent = None
        maternal_deaths_percent_val = {
            'de_name': 'Maternal mortality',
            'cat_combo': None,
            'numeric_sum': maternal_deaths_percent,
        }
        maternal_deaths_percent_val.update(_group_ou_dict)
        calculated_vals.append(maternal_deaths_percent_val)

        if all_not_none(expected_pregnant_hiv) and expected_pregnant_hiv:
            emtct_art_percent = 100 * sum_zero(anc_already_art['numeric_sum'], anc_started_art['numeric_sum'], maternity_started_art['numeric_sum'], pnc_started_art['numeric_sum']) / expected_pregnant_hiv
        else:
            emtct_art_percent = None
        emtct_art_percent_val = {
            'de_name': '% of eMTCT eligible women on ART----95%',
            'cat_combo': None,
            'numeric_sum': emtct_art_percent,
        }
        emtct_art_percent_val.update(_group_ou_dict)
        calculated_vals.append(emtct_art_percent_val)

        if all_not_none(deliveries['numeric_sum'], expected_deliver) and expected_deliver:
            deliveries_in_unit_percent = 100 * deliveries['numeric_sum'] / expected_deliver
        else:
            deliveries_in_unit_percent = None
        deliveries_in_unit_percent_val = {
            'de_name': '% of institutional deliveries  Target=60%',
            'cat_combo': None,
            'numeric_sum': deliveries_in_unit_percent,
        }
        deliveries_in_unit_percent_val.update(_group_ou_dict)
        calculated_vals.append(deliveries_in_unit_percent_val)

        if all_not_none(caesarian['numeric_sum'], expected_deliver) and expected_deliver:
            caesarean_percent = 100 * caesarian['numeric_sum'] / expected_deliver
        else:
            caesarean_percent = None
        caesarean_percent_val = {
            'de_name': 'Caesarean section rate (10%-15%)',
            'cat_combo': None,
            'numeric_sum': caesarean_percent,
        }
        caesarean_percent_val.update(_group_ou_dict)
        calculated_vals.append(caesarean_percent_val)

        _group[1] = calculated_vals # override source values

    data_element_metas = list() # override source values
    data_element_metas += list(product(['Expected Pregnancies (5 % of population)'], (None,)))
    data_element_metas += list(product(['Adolescent Population (12.8 % of population)'], (None,)))
    data_element_metas += list(product(['All expected pregnancies in a catchment population multiplied by HIV prevalence'], (None,)))
    data_element_metas += list(product(['Expected Deliveries (4.8 % of population)'], (None,)))
    data_element_metas += list(product(['% ANC1 Attendance coverage---Target=90%'], (None,)))
    data_element_metas += list(product(['% of pregnant women attending 1st ANC visit within the 1st trimester---Target=45%'], (None,)))
    data_element_metas += list(product(['Adolescent  pregnancy rate (10 -19 years of age)  Target--<5%'], (None,)))
    data_element_metas += list(product(['% ANC4 Attendance coverage---Target=60%'], (None,)))
    data_element_metas += list(product(['IPT1 Coverage--Target=90%'], (None,)))
    data_element_metas += list(product(['IPT2 Coverage--Target=90%'], (None,)))
    data_element_metas += list(product(['VITA Supplementation for mothers--Target =90%'], (None,)))
    data_element_metas += list(product(['Maternal mortality'], (None,)))
    data_element_metas += list(product(['% of eMTCT eligible women on ART----95%'], (None,)))
    data_element_metas += list(product(['% of institutional deliveries  Target=60%'], (None,)))
    data_element_metas += list(product(['Caesarean section rate (10%-15%)'], (None,)))


    num_path_elements = len(ou_headers)
    legend_sets = list()
    anc1_emtct_ls = LegendSet()
    anc1_emtct_ls.name = 'ANC1 & eMTCT'
    anc1_emtct_ls.add_interval('red', 0, 80)
    anc1_emtct_ls.add_interval('yellow', 80, 95)
    anc1_emtct_ls.add_interval('green', 95, None)
    anc1_emtct_ls.mappings[num_path_elements+4] = True
    anc1_emtct_ls.mappings[num_path_elements+12] = True
    legend_sets.append(anc1_emtct_ls)
    anc1_1st_tri_ls = LegendSet()
    anc1_1st_tri_ls.name = 'ANC1 (1st Trimester)'
    anc1_1st_tri_ls.add_interval('red', 0, 35)
    anc1_1st_tri_ls.add_interval('yellow', 35, 45)
    anc1_1st_tri_ls.add_interval('green', 45, None)
    anc1_1st_tri_ls.mappings[num_path_elements+5] = True
    legend_sets.append(anc1_1st_tri_ls)
    anc1_adolescent_ls = LegendSet()
    anc1_adolescent_ls.name = 'Adolescent Pregnancies'
    anc1_adolescent_ls.add_interval('green', 0, 2)
    anc1_adolescent_ls.add_interval('yellow', 2, 5)
    anc1_adolescent_ls.add_interval('red', 5, None)
    anc1_adolescent_ls.mappings[num_path_elements+6] = True
    legend_sets.append(anc1_adolescent_ls)
    anc4_in_unit_ls = LegendSet()
    anc4_in_unit_ls.name = 'ANC4 & Deliveries in Unit'
    anc4_in_unit_ls.add_interval('red', 0, 45)
    anc4_in_unit_ls.add_interval('yellow', 45, 60)
    anc4_in_unit_ls.add_interval('green', 60, None)
    anc4_in_unit_ls.mappings[num_path_elements+7] = True
    anc4_in_unit_ls.mappings[num_path_elements+13] = True
    legend_sets.append(anc4_in_unit_ls)
    ipt1_ipt2_vita_ls = LegendSet()
    ipt1_ipt2_vita_ls.name = 'IPT1, IPT2 & Vit. A Supplementation'
    ipt1_ipt2_vita_ls.add_interval('red', 0, 80)
    ipt1_ipt2_vita_ls.add_interval('yellow', 80, 90)
    ipt1_ipt2_vita_ls.add_interval('green', 90, None)
    ipt1_ipt2_vita_ls.mappings[num_path_elements+8] = True
    ipt1_ipt2_vita_ls.mappings[num_path_elements+9] = True
    ipt1_ipt2_vita_ls.mappings[num_path_elements+10] = True
    legend_sets.append(ipt1_ipt2_vita_ls)
    maternal_mortality_ls = LegendSet()
    maternal_mortality_ls.name = 'Maternal Mortality'
    maternal_mortality_ls.add_interval('green', 0, 2)
    maternal_mortality_ls.add_interval('yellow', 2, 5)
    maternal_mortality_ls.add_interval('red', 5, None)
    maternal_mortality_ls.mappings[num_path_elements+11] = True
    legend_sets.append(maternal_mortality_ls)
    caesarian_ls = LegendSet()
    caesarian_ls.name = 'Caesarean section rate'
    caesarian_ls.add_interval('red', 0, 6)
    caesarian_ls.add_interval('yellow', 6, 10)
    caesarian_ls.add_interval('red', 15, None)
    caesarian_ls.mappings[num_path_elements+14] = True
    legend_sets.append(caesarian_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mnch_preg_birth_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="mnch_preg_birth_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/mnch_preg_birth_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def mnch_pnc_child_scorecard(request, org_unit_level=2, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at facility level
    FACILITY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(3, prefix='org_unit__').items() if k in OU_PATH_FIELDS }
    # annotations for data collected at subcounty level
    SUBCOUNTY_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(2, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    targets_de_names = (
        'Catchment Population',
    )
    targets_short_names = (
        'Catchment Population',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))
    data_element_metas += list(product(targets_short_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_targets = qs_targets.where(filter_district)
    qs_targets = qs_targets.annotate(**SUBCOUNTY_LEVEL_ANNOTATIONS)
    # population estimates are annual, so filter by year component of period
    qs_targets = qs_targets.when(filter_period[:4])
    qs_targets = qs_targets.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_targets2 = list(gen_raster)

    maternity_de_names = (
        '105-1.3 OPD Neonatal  Sepsis (0-7days)',
        '105-2.2 Birth Asyphyxia',
        '105-2.2 No. of mothers who initiated breastfeeding within the 1st hour after delivery (Total)',
        '105-2.2a Deliveries in unit',
        '105-2.2b Deliveries in unit(Fresh Still births)',
        '105-2.2c Deliveries in unit(Macerated still births)',
        '105-2.2d Deliveries in unit(Live Births)',
        '105-2.3 Postnatal Attendances 6 Days',
        '105-2.3 Postnatal Attendances 6 Hours',
    )
    maternity_short_names = (
        # empty, no shortnames needed
    )
    de_maternity_meta = list(product(maternity_de_names, (None,)))
    data_element_metas += de_maternity_meta

    qs_maternity = DataValue.objects.what(*maternity_de_names)
    qs_maternity = qs_maternity.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_maternity = qs_maternity.where(filter_district)
    qs_maternity = qs_maternity.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_maternity = qs_maternity.when(filter_period)
    qs_maternity = qs_maternity.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_maternity = qs_maternity.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_maternity = list(val_maternity)

    gen_raster = grabbag.rasterize(ou_list, de_maternity_meta, val_maternity, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_maternity2 = list(gen_raster)

    vaccine_under_1_de_names = (
        '105-2.11 BCG',
        '105-2.11 DPT-HepB+Hib 3',
        '105-2.11 Polio 3',
    )
    vaccine_under_1_short_names = (
        # empty, no shortnames needed
    )
    de_vaccine_under_1_meta = list(product(vaccine_under_1_de_names, (None,)))
    data_element_metas += de_vaccine_under_1_meta

    qs_vaccine_under_1 = DataValue.objects.what(*vaccine_under_1_de_names)
    qs_vaccine_under_1 = qs_vaccine_under_1.filter(category_combo__categories__name='Under 1')
    qs_vaccine_under_1 = qs_vaccine_under_1.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_vaccine_under_1 = qs_vaccine_under_1.where(filter_district)
    qs_vaccine_under_1 = qs_vaccine_under_1.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_vaccine_under_1 = qs_vaccine_under_1.when(filter_period)
    qs_vaccine_under_1 = qs_vaccine_under_1.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_vaccine_under_1 = qs_vaccine_under_1.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_vaccine_under_1 = list(val_vaccine_under_1)

    gen_raster = grabbag.rasterize(ou_list, de_vaccine_under_1_meta, val_vaccine_under_1, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_vaccine_under_12 = list(gen_raster)

    under_five_categs = ('0-28 Days', '29 Days-4 Years')

    under_5_de_names = (
        '105-1.1 OPD New Attendance',
        '105-1.3 OPD Diarrhoea-Acute',
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
        '105-1.3 OPD Pneumonia',
    )
    under_5_short_names = (
        # empty, no shortnames needed
    )
    de_under_5_meta = list(product(under_5_de_names, (None,)))
    data_element_metas += de_under_5_meta

    qs_under_5 = DataValue.objects.what(*under_5_de_names)
    qs_under_5 = qs_under_5.filter(category_combo__categories__name__in=under_five_categs)
    qs_under_5 = qs_under_5.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_under_5 = qs_under_5.where(filter_district)
    qs_under_5 = qs_under_5.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_under_5 = qs_under_5.when(filter_period)
    qs_under_5 = qs_under_5.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_under_5 = qs_under_5.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_under_5 = list(val_under_5)

    gen_raster = grabbag.rasterize(ou_list, de_under_5_meta, val_under_5, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_under_52 = list(gen_raster)

    other_de_names = (
        '105-2.11 PCV 3',
        '105-2.8 Dewormed 2nd Dose in the Year',
        '105-2.8 Vit A Suplement 2nd Dose in theYear',
    )
    other_short_names = (
        # empty, no shortnames needed
    )
    de_other_meta = list(product(other_de_names, (None,)))
    data_element_metas += de_other_meta

    qs_other = DataValue.objects.what(*other_de_names)
    qs_other = qs_other.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_other = qs_other.where(filter_district)
    qs_other = qs_other.annotate(**FACILITY_LEVEL_ANNOTATIONS)
    qs_other = qs_other.when(filter_period)
    qs_other = qs_other.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_other = qs_other.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_other = list(val_other)

    gen_raster = grabbag.rasterize(ou_list, de_other_meta, val_other, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_other2 = list(gen_raster)


    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_maternity2, val_vaccine_under_12, val_under_52, val_other2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))


    # perform calculations
    for _group in grouped_vals:
        (_group_ou_path, (catchment_pop, neonatal_sepsis, asyphyxia, breastfeeding, deliveries, still_fresh, still_macerated, live_births, pnc_6_days, pnc_6_hours, bcg_under1, dpt3_under1, polio3_under1, new_attend_under5, acute_diarr_under5, malaria_under5, malaria_conf_under5, pneum_under5, pcv, deworm, vitamin_a, *other_vals)) = _group
        _group_ou_dict = dict(zip(OU_PATH_FIELDS, _group_ou_path))
        
        calculated_vals = list()

        if all_not_none(catchment_pop['numeric_sum']):
            expected_deliver = (catchment_pop['numeric_sum'] * Decimal(0.0485))/4 # split by quarter
        else:
            expected_deliver = None
        expected_deliver_val = {
            'de_name': 'Expected Deliveries',
            'cat_combo': None,
            'numeric_sum': expected_deliver,
        }
        expected_deliver_val.update(_group_ou_dict)
        calculated_vals.append(expected_deliver_val)

        if all_not_none(catchment_pop['numeric_sum']):
            expected_under_1_pop = catchment_pop['numeric_sum'] * Decimal(0.043)
        else:
            expected_under_1_pop = None
        expected_under_1_pop_val = {
            'de_name': 'Number of children below one year in a given population',
            'cat_combo': None,
            'numeric_sum': expected_under_1_pop,
        }
        expected_under_1_pop_val.update(_group_ou_dict)
        calculated_vals.append(expected_under_1_pop_val)

        if all_not_none(catchment_pop['numeric_sum']):
            expected_under_5_malaria = (catchment_pop['numeric_sum'] * Decimal(0.177))/4 # split by quarter
        else:
            expected_under_5_malaria = None
        expected_under_5_malaria_val = {
            'de_name': 'Expected under-five with positive test for malaria',
            'cat_combo': None,
            'numeric_sum': expected_under_5_malaria,
        }
        expected_under_5_malaria_val.update(_group_ou_dict)
        calculated_vals.append(expected_under_5_malaria_val)

        if all_not_none(pnc_6_days['numeric_sum'], pnc_6_hours['numeric_sum'], expected_deliver) and expected_deliver:
            pnc_6_days_percent = 100 * (pnc_6_days['numeric_sum']+pnc_6_hours['numeric_sum']) / expected_deliver
        else:
            pnc_6_days_percent = None
        pnc_6_days_percent_val = {
            'de_name': '% of Mothers receiving PNC checks within 6 days----Target=60%',
            'cat_combo': None,
            'numeric_sum': pnc_6_days_percent,
        }
        pnc_6_days_percent_val.update(_group_ou_dict)
        calculated_vals.append(pnc_6_days_percent_val)

        if all_not_none(breastfeeding['numeric_sum'], deliveries['numeric_sum']) and deliveries:
            breastfeeding_percent = 100 * breastfeeding['numeric_sum'] / deliveries['numeric_sum']
        else:
            breastfeeding_percent = None
        breastfeeding_percent_val = {
            'de_name': '% of  mothers initiating breastfeeding within 1 hour after birth--Target=90%',
            'cat_combo': None,
            'numeric_sum': breastfeeding_percent,
        }
        breastfeeding_percent_val.update(_group_ou_dict)
        calculated_vals.append(breastfeeding_percent_val)

        if all_not_none(asyphyxia['numeric_sum'], live_births['numeric_sum']) and live_births['numeric_sum']:
            asyphyxia_percent = 100 * asyphyxia['numeric_sum'] / live_births['numeric_sum']
        else:
            asyphyxia_percent = None
        asyphyxia_percent_val = {
            'de_name': '% of babies with Birth Asphyxia ---<1.1',
            'cat_combo': None,
            'numeric_sum': asyphyxia_percent,
        }
        asyphyxia_percent_val.update(_group_ou_dict)
        calculated_vals.append(asyphyxia_percent_val)

        expected_live = sum_zero(expected_deliver) - sum_zero(still_fresh['numeric_sum'], still_macerated['numeric_sum'])
        if all_not_none(neonatal_sepsis['numeric_sum'], expected_live) and expected_live:
            sepsis_percent = 100 * neonatal_sepsis['numeric_sum'] / expected_live
        else:
            sepsis_percent = None
        sepsis_percent_val = {
            'de_name': '% of neonates (aged 0 -28 days) presenting to health facilities with sepsis/infections <1.1',
            'cat_combo': None,
            'numeric_sum': sepsis_percent,
        }
        sepsis_percent_val.update(_group_ou_dict)
        calculated_vals.append(sepsis_percent_val)

        if all_not_none(dpt3_under1['numeric_sum'], expected_under_1_pop) and expected_under_1_pop:
            dpt3_percent = 100 * dpt3_under1['numeric_sum'] / expected_under_1_pop
        else:
            dpt3_percent = None
        dpt3_percent_val = {
            'de_name': 'DPT 3 coverage--Target=97%',
            'cat_combo': None,
            'numeric_sum': dpt3_percent,
        }
        dpt3_percent_val.update(_group_ou_dict)
        calculated_vals.append(dpt3_percent_val)

        if all_not_none(bcg_under1['numeric_sum'], expected_under_1_pop) and expected_under_1_pop:
            bcg_percent = 100 * bcg_under1['numeric_sum'] / expected_under_1_pop
        else:
            bcg_percent = None
        bcg_percent_val = {
            'de_name': 'BCGCoverage---Target=97%',
            'cat_combo': None,
            'numeric_sum': bcg_percent,
        }
        bcg_percent_val.update(_group_ou_dict)
        calculated_vals.append(bcg_percent_val)

        if all_not_none(polio3_under1['numeric_sum'], expected_under_1_pop) and expected_under_1_pop:
            polio3_percent = 100 * polio3_under1['numeric_sum'] / expected_under_1_pop
        else:
            polio3_percent = None
        polio3_percent_val = {
            'de_name': 'Polio3 Coverage---97%',
            'cat_combo': None,
            'numeric_sum': polio3_percent,
        }
        polio3_percent_val.update(_group_ou_dict)
        calculated_vals.append(polio3_percent_val)

        if all_not_none(malaria_conf_under5['numeric_sum']) and malaria_under5['numeric_sum']:
            malaria_conf_lab_under5_percent = 100 * malaria_conf_under5['numeric_sum'] / malaria_under5['numeric_sum']
        else:
            malaria_conf_lab_under5_percent = None
        malaria_conf_lab_under5_percent_val = {
            'de_name': '% of children U5 diagnosed with malaria who have laboratory confirmation.-----90%',
            'cat_combo': None,
            'numeric_sum': malaria_conf_lab_under5_percent,
        }
        malaria_conf_lab_under5_percent_val.update(_group_ou_dict)
        calculated_vals.append(malaria_conf_lab_under5_percent_val)

        if all_not_none(malaria_conf_under5['numeric_sum']) and expected_under_5_malaria:
            malaria_conf_under5_percent = 100 * malaria_conf_under5['numeric_sum'] / expected_under_5_malaria
        else:
            malaria_conf_under5_percent = None
        malaria_conf_under5_percent_val = {
            'de_name': '% 0f children under five with confirmed malaria---Target<20%',
            'cat_combo': None,
            'numeric_sum': malaria_conf_under5_percent,
        }
        malaria_conf_under5_percent_val.update(_group_ou_dict)
        calculated_vals.append(malaria_conf_under5_percent_val)

        if all_not_none(acute_diarr_under5['numeric_sum']) and new_attend_under5['numeric_sum']:
            acute_diarr_under5_percent = 100 * acute_diarr_under5['numeric_sum'] / new_attend_under5['numeric_sum']
        else:
            acute_diarr_under5_percent = None
        acute_diarr_under5_percent_val = {
            'de_name': '% under 5 treated with diarrhorea---Target=<20%',
            'cat_combo': None,
            'numeric_sum': acute_diarr_under5_percent,
        }
        acute_diarr_under5_percent_val.update(_group_ou_dict)
        calculated_vals.append(acute_diarr_under5_percent_val)

        if all_not_none(pneum_under5['numeric_sum']) and new_attend_under5['numeric_sum']:
            pneum_under5_percent = 100 * pneum_under5['numeric_sum'] / new_attend_under5['numeric_sum']
        else:
            pneum_under5_percent = None
        pneum_under5_percent_val = {
            'de_name': '% under 5 treated with pneumonia----Target=<20%',
            'cat_combo': None,
            'numeric_sum': pneum_under5_percent,
        }
        pneum_under5_percent_val.update(_group_ou_dict)
        calculated_vals.append(pneum_under5_percent_val)

        if all_not_none(vitamin_a['numeric_sum']) and expected_under_1_pop:
            vitamin_a_percent = 100 * vitamin_a['numeric_sum'] / expected_under_1_pop
        else:
            vitamin_a_percent = None
        vitamin_a_percent_val = {
            'de_name': ' Vit A Suplement 2nd Dose COVERAGE  in theYear---Target=97%',
            'cat_combo': None,
            'numeric_sum': vitamin_a_percent,
        }
        vitamin_a_percent_val.update(_group_ou_dict)
        calculated_vals.append(vitamin_a_percent_val)

        if all_not_none(deworm['numeric_sum']) and expected_under_1_pop:
            deworm_percent = 100 * deworm['numeric_sum'] / expected_under_1_pop
        else:
            deworm_percent = None
        deworm_percent_val = {
            'de_name': '105-2.8 Dewormed 2nd Dose COVERAGE in the Year----Target=97%',
            'cat_combo': None,
            'numeric_sum': deworm_percent,
        }
        deworm_percent_val.update(_group_ou_dict)
        calculated_vals.append(deworm_percent_val)

        if all_not_none(pcv['numeric_sum']) and expected_under_1_pop:
            pcv_percent = 100 * pcv['numeric_sum'] / expected_under_1_pop
        else:
            pcv_percent = None
        pcv_percent_val = {
            'de_name': 'PCV3 Coverage----Target=97%',
            'cat_combo': None,
            'numeric_sum': pcv_percent,
        }
        pcv_percent_val.update(_group_ou_dict)
        calculated_vals.append(pcv_percent_val)

        # _group[1].extend(calculated_vals)
        _group[1] = calculated_vals # override source values

    data_element_metas = list() # override source values
    data_element_metas += list(product(['Expected Deliveries (4.8 % of population)'], (None,)))
    data_element_metas += list(product(['Number of children below one year in a given population (4.3 % of population)'], (None,)))
    data_element_metas += list(product(['Expected under-five with positive test for malaria (17.7 % of population)'], (None,)))
    data_element_metas += list(product(['% of Mothers receiving PNC checks within 6 days----Target=60%'], (None,)))
    data_element_metas += list(product(['% of  mothers initiating breastfeeding within 1 hour after birth--Target=90%'], (None,)))
    data_element_metas += list(product(['% of babies with Birth Asphyxia ---<1.1'], (None,)))
    data_element_metas += list(product(['% of neonates (aged 0 -28 days) presenting to health facilities with sepsis/infections <1.1'], (None,)))
    data_element_metas += list(product(['DPT 3 coverage--Target=97% '], (None,)))
    data_element_metas += list(product(['BCGCoverage---Target=97%'], (None,)))
    data_element_metas += list(product(['Polio3 Coverage---97%'], (None,)))
    data_element_metas += list(product(['% of children U5 diagnosed with malaria who have laboratory confirmation.-----90%'], (None,)))
    data_element_metas += list(product(['% 0f children under five with confirmed malaria---Target<20%'], (None,)))
    data_element_metas += list(product(['% under 5 treated with diarrhorea---Target=<20%'], (None,)))
    data_element_metas += list(product(['% under 5 treated with pneumonia----Target=<20%'], (None,)))
    data_element_metas += list(product([' Vit A Suplement 2nd Dose COVERAGE  in theYear---Target=97%'], (None,)))
    data_element_metas += list(product(['105-2.8 Dewormed 2nd Dose COVERAGE in the Year----Target=97%'], (None,)))
    data_element_metas += list(product(['PCV3 Coverage----Target=97%'], (None,)))


    num_path_elements = len(ou_headers)
    legend_sets = list()
    pnc_6days_ls = LegendSet()
    pnc_6days_ls.name = 'PNC check within 6 days'
    pnc_6days_ls.add_interval('red', 0, 45)
    pnc_6days_ls.add_interval('yellow', 45, 60)
    pnc_6days_ls.add_interval('green', 60, None)
    pnc_6days_ls.mappings[num_path_elements+3] = True
    legend_sets.append(pnc_6days_ls)
    breast_vaccination_ls = LegendSet()
    breast_vaccination_ls.name = 'Breastfeeding, DPT3, BCG and  Polio3'
    breast_vaccination_ls.add_interval('red', 0, 80)
    breast_vaccination_ls.add_interval('yellow', 80, 95)
    breast_vaccination_ls.add_interval('green', 95, None)
    breast_vaccination_ls.mappings[num_path_elements+4] = True
    breast_vaccination_ls.mappings[num_path_elements+7] = True
    breast_vaccination_ls.mappings[num_path_elements+8] = True
    breast_vaccination_ls.mappings[num_path_elements+9] = True
    legend_sets.append(breast_vaccination_ls)
    asphyxia_sepsis_ls = LegendSet()
    asphyxia_sepsis_ls.name = 'Birth Asphyxia and Neonatal Sepsis'
    asphyxia_sepsis_ls.add_interval('green', 0, 1.1)
    asphyxia_sepsis_ls.add_interval('yellow', 1.1, 1.4)
    asphyxia_sepsis_ls.add_interval('red', 1.4, None)
    asphyxia_sepsis_ls.mappings[num_path_elements+5] = True
    asphyxia_sepsis_ls.mappings[num_path_elements+6] = True
    legend_sets.append(asphyxia_sepsis_ls)
    mal_conf_treat_ls = LegendSet()
    mal_conf_treat_ls.name = 'Malaria Treatment with Lab Confirmation'
    mal_conf_treat_ls.add_interval('red', 0, 70)
    mal_conf_treat_ls.add_interval('yellow', 70, 90)
    mal_conf_treat_ls.add_interval('green', 90, None)
    mal_conf_treat_ls.mappings[num_path_elements+10] = True
    legend_sets.append(mal_conf_treat_ls)
    mal_conf_diarr_pneum_ls = LegendSet()
    mal_conf_diarr_pneum_ls.name = 'Under 5: Confirmed Malaria, Diarrhorea and Pneumonia'
    mal_conf_diarr_pneum_ls.add_interval('green', 0, 20)
    mal_conf_diarr_pneum_ls.add_interval('yellow', 20, 30)
    mal_conf_diarr_pneum_ls.add_interval('red', 30, None)
    mal_conf_diarr_pneum_ls.mappings[num_path_elements+11] = True
    mal_conf_diarr_pneum_ls.mappings[num_path_elements+12] = True
    mal_conf_diarr_pneum_ls.mappings[num_path_elements+13] = True
    legend_sets.append(mal_conf_diarr_pneum_ls)
    vita_deworm_pcv_ls = LegendSet()
    vita_deworm_pcv_ls.name = 'Vit. A, Deworming and PCV3'
    vita_deworm_pcv_ls.add_interval('red', 0, 80)
    vita_deworm_pcv_ls.add_interval('yellow', 80, 97)
    vita_deworm_pcv_ls.add_interval('green', 97, None)
    vita_deworm_pcv_ls.mappings[num_path_elements+14] = True
    vita_deworm_pcv_ls.mappings[num_path_elements+15] = True
    vita_deworm_pcv_ls.mappings[num_path_elements+16] = True
    legend_sets.append(vita_deworm_pcv_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mnch_pnc_child_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="mnch_pnc_child_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response


    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/mnch_pnc_child_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)

@login_required
def lqas_scorecard(request, org_unit_level=3, output_format='HTML'):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YRS = ['%d' % (y,) for y in range(this_year, this_year-6, -1)]
    DISTRICT_LIST = list(OrgUnit.objects.filter(level=1).order_by('name').values_list('name', flat=True))
    OU_PATH_FIELDS = OrgUnit.level_fields(org_unit_level)[1:] # skip the topmost/country level
    # annotations for data collected at district level
    DISTRICT_LEVEL_ANNOTATIONS = { k:v for k,v in OrgUnit.level_annotations(1, prefix='org_unit__').items() if k in OU_PATH_FIELDS }

    if 'period' in request.GET and request.GET['period'] in PREV_5YRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d' % (this_year,)

    period_desc = filter_period

    if 'district' in request.GET and request.GET['district'] in DISTRICT_LIST:
        filter_district = OrgUnit.objects.get(name=request.GET['district'])
    else:
        filter_district = None

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=org_unit_level).annotate(**OrgUnit.level_annotations(org_unit_level))
    if filter_district:
        qs_ou = qs_ou.filter(Q(lft__gte=filter_district.lft) & Q(rght__lte=filter_district.rght))
    qs_ou = qs_ou.order_by(*OU_PATH_FIELDS)

    ou_list = list(qs_ou.values_list(*OU_PATH_FIELDS))
    ou_headers = OrgUnit.level_names(org_unit_level)[1:] # skip the topmost/country level

    def orgunit_vs_de_catcombo_default(row, col):
        val_dict = dict(zip(OU_PATH_FIELDS, row))
        de_name, subcategory = col
        val_dict.update({ 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None })
        return val_dict

    data_element_metas = list()

    lqas_de_names = (
        '%  of children 0-59 months who slept under a ITN the night preceding the survey',
        '%  of individuals who know at least two signs and symptoms of TB',
        '%  of individuals who know how HIV transmission occur from an infected mother to child',
        '%  of individuals who know two key actions that reduce HIV transmission from an infected mother to her child',
        '%  of mothers of children 0-11 months who attended ANC at least 4 times during last pregnancy',
        '%  of mothers of children 0-23 months who received two or more doses of IPT2 during their last pregnancy ',
        '%  of the male youth 15-24yrs who are circumcised',
        '%  of youth 15-24 years who perceive low or no risk of getting HIV/AIDS infection',
        '%  of youth who have had sexual intercourse before the age of 15 years',
        '% of Households with at least one ITN',
        '% of children age 36-59 months who are developmentally on track in literacy-numeracy, physical, social-emotional, and learning domains, and the early child deve',
        '% of children aged 0-59 months who had a fever in the last two weeks and were tested for malaria ',
        '% of individuals who had sex with a non-marital or non-cohabiting sexual partner in the last 12 months',
        '% of individuals who had sex with more than one sexual partner in the last 12 months',
        '% of individuals who know how TB is transmitted',
        '% of individuals who know that TB is curable disease',
        '% of individuals who know the risk of not completing TB treatment',
        '% of individuals who were counselled and received an HIV test in last 12 months and know their results',
        '% of mothers of children 0-11 months who were assisted by a trained health worker during delivery',
        '% of mothers of children 0-59 months who know two or more ways to prevent malaria',
        '% of mothers of children under five years who know two or more signs and  symptoms of malaria ',
        '% of women and men age 15 years and above with comprehensive knowledge of HIV',
        '% of women and men aged 15-49 who experienced sexual violence in the last 12 months',
        '% of women in the reproductive age group 15-49 who known at least 3 methods of family planning and have used the method ',
    )
    de_lqas_meta = list(product(lqas_de_names, (None,)))
    data_element_metas += list(product(lqas_de_names, (None,)))

    qs_lqas = DataValue.objects.what(*lqas_de_names)
    qs_lqas = qs_lqas.annotate(cat_combo=Value(None, output_field=CharField()))
    if filter_district:
        qs_lqas = qs_lqas.where(filter_district)
    qs_lqas = qs_lqas.annotate(**DISTRICT_LEVEL_ANNOTATIONS)
    qs_lqas = qs_lqas.when(filter_period)
    qs_lqas = qs_lqas.order_by(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period')
    val_lqas = qs_lqas.values(*OU_PATH_FIELDS, 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_lqas_meta, val_lqas, ou_path_from_dict, lambda x: (x['de_name'], x['cat_combo']), orgunit_vs_de_catcombo_default)
    val_lqas2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_lqas2), key=ou_path_from_dict), key=ou_path_from_dict)
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    num_path_elements = len(ou_headers)
    legend_sets = list()
    bfk_ls = LegendSet()
    bfk_ls.name = 'LQAS - BFK'
    bfk_ls.add_interval('red', 0, 40)
    bfk_ls.add_interval('yellow', 40, 60)
    bfk_ls.add_interval('green', 60, None)
    bfk_ls.mappings[num_path_elements+1] = True
    bfk_ls.mappings[num_path_elements+12] = True
    bfk_ls.mappings[num_path_elements+17] = True
    legend_sets.append(bfk_ls)
    cdi_ls = LegendSet()
    cdi_ls.name = 'LQAS - CDI'
    cdi_ls.add_interval('red', 0, 45)
    cdi_ls.add_interval('yellow', 45, 65)
    cdi_ls.add_interval('green', 65, None)
    cdi_ls.mappings[num_path_elements+2] = True
    cdi_ls.mappings[num_path_elements+3] = True
    cdi_ls.mappings[num_path_elements+6] = True
    legend_sets.append(cdi_ls)
    e_ls = LegendSet()
    e_ls.name = 'LQAS - E'
    e_ls.add_interval('red', 0, 5.9)
    e_ls.add_interval('yellow', 5.9, 17.7)
    e_ls.add_interval('green', 17.7, None)
    e_ls.mappings[num_path_elements+13] = True
    legend_sets.append(e_ls)
    gs_ls = LegendSet()
    gs_ls.name = 'LQAS - GS'
    gs_ls.add_interval('red', 0, 18.3)
    gs_ls.add_interval('yellow', 18.3, 55)
    gs_ls.add_interval('green', 55, None)
    gs_ls.mappings[num_path_elements+4] = True
    gs_ls.mappings[num_path_elements+7] = True
    legend_sets.append(gs_ls)
    h_ls = LegendSet()
    h_ls.name = 'LQAS - H'
    h_ls.add_interval('red', 0, 3.3)
    h_ls.add_interval('yellow', 3.3, 10)
    h_ls.add_interval('green', 10, None)
    h_ls.mappings[num_path_elements+8] = True
    legend_sets.append(h_ls)
    jrt_ls = LegendSet()
    jrt_ls.name = 'LQAS - JRT'
    jrt_ls.add_interval('red', 0, 26.7)
    jrt_ls.add_interval('yellow', 26.7, 80)
    jrt_ls.add_interval('green', 80, None)
    jrt_ls.mappings[num_path_elements+9] = True
    jrt_ls.mappings[num_path_elements+15] = True
    jrt_ls.mappings[num_path_elements+18] = True
    legend_sets.append(jrt_ls)
    l_ls = LegendSet()
    l_ls.name = 'LQAS - L'
    l_ls.add_interval('red', 0, 27.3)
    l_ls.add_interval('yellow', 27.3, 82)
    l_ls.add_interval('green', 82, None)
    l_ls.mappings[num_path_elements+14] = True
    legend_sets.append(l_ls)
    m_ls = LegendSet()
    m_ls.name = 'LQAS - M'
    m_ls.add_interval('red', 0, 31)
    m_ls.add_interval('yellow', 31, 93)
    m_ls.add_interval('green', 93, None)
    m_ls.mappings[num_path_elements+16] = True
    legend_sets.append(m_ls)
    n_ls = LegendSet()
    n_ls.name = 'LQAS - N'
    n_ls.add_interval('red', 0, 10)
    n_ls.add_interval('yellow', 10, 30)
    n_ls.add_interval('green', 30, None)
    n_ls.mappings[num_path_elements+5] = True
    legend_sets.append(n_ls)
    o_ls = LegendSet()
    o_ls.name = 'LQAS - O'
    o_ls.add_interval('red', 0, 30)
    o_ls.add_interval('yellow', 30, 90)
    o_ls.add_interval('green', 90, None)
    o_ls.mappings[num_path_elements+0] = True
    legend_sets.append(o_ls)
    p_ls = LegendSet()
    p_ls.name = 'LQAS - P'
    p_ls.add_interval('red', 0, 8.3)
    p_ls.add_interval('yellow', 8.3, 25)
    p_ls.add_interval('green', 25, None)
    p_ls.mappings[num_path_elements+19] = True
    legend_sets.append(p_ls)
    q_ls = LegendSet()
    q_ls.name = 'LQAS - Q'
    q_ls.add_interval('red', 0, 50)
    q_ls.add_interval('yellow', 50, 75)
    q_ls.add_interval('green', 75, None)
    q_ls.mappings[num_path_elements+20] = True
    legend_sets.append(q_ls)


    def grouped_data_generator(grouped_data):
        for group_ou_path, group_values in grouped_data:
            yield (*group_ou_path, *tuple(map(lambda val: val['numeric_sum'], group_values)))

    if output_format == 'CSV':
        import csv
        value_rows = list()
        value_rows.append((*ou_headers, *data_element_metas))
        for row in grouped_data_generator(grouped_vals):
            value_rows.append(row)

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="lqas_{0}_scorecard.csv"'.format(OrgUnit.get_level_field(org_unit_level))

        writer = csv.writer(response, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerows(value_rows)

        return response

    if output_format == 'EXCEL':
        wb = openpyxl.workbook.Workbook()
        ws = wb.active # workbooks are created with at least one worksheet
        ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = chain(ou_headers, data_element_metas)
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            ou_path, g_val_list = g
            for col_idx, ou in enumerate(ou_path, start=1):
                ws.cell(row=i, column=col_idx, value=ou)
            for j, g_val in enumerate(g_val_list, start=len(ou_path)+1):
                ws.cell(row=i, column=j, value=g_val['numeric_sum'])

        for ls in legend_sets:
            # apply conditional formatting from LegendSets
            for rule in ls.openpyxl_rules():
                for cell_range in ls.excel_ranges():
                    ws.conditional_formatting.add(cell_range, rule)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="lqas_{0}_scorecard.xlsx"'.format(OrgUnit.get_level_field(org_unit_level))

        return response

    context = {
        'grouped_data': grouped_vals,
        'ou_headers': ou_headers,
        'data_element_names': data_element_metas,
        'legend_sets': legend_sets,
        'period_desc': period_desc,
        'period_list': PREV_5YRS,
        'district_list': DISTRICT_LIST,
        'excel_url': make_excel_url(request.path),
        'csv_url': make_csv_url(request.path),
        'legend_set_mappings': { tuple([i-len(ou_headers) for i in ls.mappings]):ls.canonical_name() for ls in legend_sets },
    }

    return render(request, 'cannula/lqas_{0}.html'.format(OrgUnit.get_level_field(org_unit_level)), context)


#reports logic
def indexreport(request):
    context = {
        'validation_rules': ValidationRule.objects.all().values_list('id', 'name')
    }
    return render(request, 'cannula/index_reports.html', context)

def reports_sites_2016_to_2018(request):
    context = {
    'id' : 1,
    } 
    return render(request, 'cannula/performance_summary_oct_2016–sep_2017.html', context)

def reports_sites_2017_to_2018(request):
    context = {
    'id' : 1,
    } 
    return render(request, 'cannula/performance_summary_oct_2017–sep_2018.html', context)

def downloadreport(request, path):
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/pdf")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    
    
    raise Http404

def lqas_by_site(request, output_format='HTML'):
 
    alldis  =   []
    year2015 = []
    year2016 = []
    year2017 = []

    PERIOD_LIST = list(lqas_dataset.objects.all().order_by('period').values_list('period', flat=True).distinct())
    PERIOD_LIST.append("None")
    DISTRICT_LIST = list(lqas_dataset.objects.all().order_by('district').values_list('district', flat=True).distinct())

    #Jacob filter logic and load all content code
    if 'district' in request.GET and 'period' in request.GET:
        filter_district = request.GET.get('district')
        filter_period   = request.GET.get('period')

        if request.GET.get('district') == "" and request.GET.get('period') != "None":
            query_results = lqas_dataset.objects.filter(period='%s' %(request.GET.get('period')))
        elif request.GET.get('district') == "" and request.GET.get('period') == "None":
            filter_district = None
            filter_period   = None
            query_results = lqas_dataset.objects.all()
            alldis=[avg('ca0'),avg('ca1'),avg('ca2'),avg('ca3'),avg('ca4'),avg('ca5'),avg('ca6'),avg('ca7'),avg('ca9'),avg('ca10'),avg('ca11'),avg('ca12'),avg('ca13'),avg('ca14'),avg('ca15'),avg('ca16'),avg('ca17'),avg('ca18'),avg('ca19'),avg('ca20'),avg('ca21'),avg('ca22'),avg('ca23'),avg('ca24')] 
            year2015=[avgforyear('ca0','2015'),avgforyear('ca1','2015'),avgforyear('ca2','2015'),avgforyear('ca3','2015'),avgforyear('ca4','2015'),avgforyear('ca5','2015'),avgforyear('ca6','2015'),avgforyear('ca7','2015'),avgforyear('ca9','2015'),avgforyear('ca10','2015'),avgforyear('ca11','2015'),avgforyear('ca12','2015'),avgforyear('ca13','2015'),avgforyear('ca14','2015'),avgforyear('ca15','2015'),avgforyear('ca16','2015'),avgforyear('ca17','2015'),avgforyear('ca18','2015'),avgforyear('ca19','2015'),avgforyear('ca20','2015'),avgforyear('ca21','2015'),avgforyear('ca22','2015'),avgforyear('ca23','2015'),avgforyear('ca24','2015')] 
            year2016=[avgforyear('ca0','2016'),avgforyear('ca1','2016'),avgforyear('ca2','2016'),avgforyear('ca3','2016'),avgforyear('ca4','2016'),avgforyear('ca5','2016'),avgforyear('ca6','2016'),avgforyear('ca7','2016'),avgforyear('ca9','2016'),avgforyear('ca10','2016'),avgforyear('ca11','2016'),avgforyear('ca12','2016'),avgforyear('ca13','2016'),avgforyear('ca14','2016'),avgforyear('ca15','2016'),avgforyear('ca16','2016'),avgforyear('ca17','2016'),avgforyear('ca18','2016'),avgforyear('ca19','2016'),avgforyear('ca20','2016'),avgforyear('ca21','2016'),avgforyear('ca22','2016'),avgforyear('ca23','2016'),avgforyear('ca24','2016')] 
            year2017=[avgforyear('ca0','2017'),avgforyear('ca1','2017'),avgforyear('ca2','2017'),avgforyear('ca3','2017'),avgforyear('ca4','2017'),avgforyear('ca5','2017'),avgforyear('ca6','2017'),avgforyear('ca7','2017'),avgforyear('ca9','2017'),avgforyear('ca10','2017'),avgforyear('ca11','2017'),avgforyear('ca12','2017'),avgforyear('ca13','2017'),avgforyear('ca14','2017'),avgforyear('ca15','2017'),avgforyear('ca16','2017'),avgforyear('ca17','2017'),avgforyear('ca18','2017'),avgforyear('ca19','2017'),avgforyear('ca20','2017'),avgforyear('ca21','2017'),avgforyear('ca22','2017'),avgforyear('ca23','2017'),avgforyear('ca24','2017')] 

        else:
            query_results = lqas_dataset.objects.filter(district='%s' %(request.GET.get('district')), period='%s' %(request.GET.get('period')))        
    else:
        filter_district = None
        filter_period   = None
        query_results = lqas_dataset.objects.all()
        alldis=[avg('ca0'),avg('ca1'),avg('ca2'),avg('ca3'),avg('ca4'),avg('ca5'),avg('ca6'),avg('ca7'),avg('ca9'),avg('ca10'),avg('ca11'),avg('ca12'),avg('ca13'),avg('ca14'),avg('ca15'),avg('ca16'),avg('ca17'),avg('ca18'),avg('ca19'),avg('ca20'),avg('ca21'),avg('ca22'),avg('ca23'),avg('ca24')] 
        year2015=[avgforyear('ca0','2015'),avgforyear('ca1','2015'),avgforyear('ca2','2015'),avgforyear('ca3','2015'),avgforyear('ca4','2015'),avgforyear('ca5','2015'),avgforyear('ca6','2015'),avgforyear('ca7','2015'),avgforyear('ca9','2015'),avgforyear('ca10','2015'),avgforyear('ca11','2015'),avgforyear('ca12','2015'),avgforyear('ca13','2015'),avgforyear('ca14','2015'),avgforyear('ca15','2015'),avgforyear('ca16','2015'),avgforyear('ca17','2015'),avgforyear('ca18','2015'),avgforyear('ca19','2015'),avgforyear('ca20','2015'),avgforyear('ca21','2015'),avgforyear('ca22','2015'),avgforyear('ca23','2015'),avgforyear('ca24','2015')] 
        year2016=[avgforyear('ca0','2016'),avgforyear('ca1','2016'),avgforyear('ca2','2016'),avgforyear('ca3','2016'),avgforyear('ca4','2016'),avgforyear('ca5','2016'),avgforyear('ca6','2016'),avgforyear('ca7','2016'),avgforyear('ca9','2016'),avgforyear('ca10','2016'),avgforyear('ca11','2016'),avgforyear('ca12','2016'),avgforyear('ca13','2016'),avgforyear('ca14','2016'),avgforyear('ca15','2016'),avgforyear('ca16','2016'),avgforyear('ca17','2016'),avgforyear('ca18','2016'),avgforyear('ca19','2016'),avgforyear('ca20','2016'),avgforyear('ca21','2016'),avgforyear('ca22','2016'),avgforyear('ca23','2016'),avgforyear('ca24','2016')] 
        year2017=[avgforyear('ca0','2017'),avgforyear('ca1','2017'),avgforyear('ca2','2017'),avgforyear('ca3','2017'),avgforyear('ca4','2017'),avgforyear('ca5','2017'),avgforyear('ca6','2017'),avgforyear('ca7','2017'),avgforyear('ca9','2017'),avgforyear('ca10','2017'),avgforyear('ca11','2017'),avgforyear('ca12','2017'),avgforyear('ca13','2017'),avgforyear('ca14','2017'),avgforyear('ca15','2017'),avgforyear('ca16','2017'),avgforyear('ca17','2017'),avgforyear('ca18','2017'),avgforyear('ca19','2017'),avgforyear('ca20','2017'),avgforyear('ca21','2017'),avgforyear('ca22','2017'),avgforyear('ca23','2017'),avgforyear('ca24','2017')] 
    query_results_targets = lqas_target.objects.all()
    
    #this is not needed as we can export the table direct
    #if output_format == 'EXCEL':
     #   wb = openpyxl.workbook.Workbook()
     #   ws = wb.active # workbooks are created with at least one worksheet
     #   ws.title = 'Sheet1' # unfortunately it is named "Sheet" not "Sheet1"
     #   ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
     #   ws.page_setup.paperSize = ws.PAPERSIZE_A4
     #   worksheet = wb.active
        
        #column hearders
     #   col_headers=['Period',
	#	'District',
	#	'% of individuals who were counselled and received an HIV test in last 12 months and know their results',
	#	'% of individuals who know how HIV transmission occur from an infected mother to child',
	#	'% of individuals who know two key actions that reduce HIV transmission from an infected mother to her child',
	#	'% of individuals who had sex with more than one sexual partner in the last 12 months',
	#	'% of individuals who had sex with a non-marital or non-cohabiting sexual partner in the last 12 months',
	#	'% of youth 15-24 years who perceive low or no risk of getting HIV/AIDS infection',
	#	'% of youth who have had sexual intercourse before the age of 15 years',
	#	'% of the male youth 15-24yrs who are circumcised',
	#	'% of individuals who know that TB is curable disease',
	#	'% of individuals who know at least two signs and symptoms of TB',
	#	'% of individuals who know how TB is transmitted',
	#	'% of individuals who know the risk of not completing TB treatment',
	#	'% of mothers of children 0-23 months who received two or more doses of IPT2 during their last pregnancy ',
	#	'% of children 0-59 months who slept under a ITN the night preceding the survey',
	#	'% of mothers of children 0-59 months who know two or more ways to prevent malaria',
	#	'% of mothers of children under five years who know two or more signs and  symptoms of malaria ',
	#	'% of Households with at least one ITN',
	#	'% of mothers of children 0-11 months who attended ANC at least 4 times during last pregnancy',
	#	'% of mothers of children 0-11 months who were assisted by a trained health worker during delivery',
	#	'% of women and men age 15 years and above with comprehensive knowledge of HIV',
	#	'% of women in the reproductive age group 15-49 who known at least 3 methods of family planning and have used the method', 
	#	'% of children aged 0-59 months who had a fever in the last two weeks and were tested for malaria ',
	#	'% of children age 36-59 months who are developmentally on track in literacy-numeracy, physical, social-emotional, and learning domains, and the early child development index score (developmentally on track in at least three of these four domains)',
	#	'% of women and men aged 15-49 who experienced sexual violence in the last 12 months']

    #    ws.append(col_headers)

    #   for item in query_results:
    #      tempArray=[item.period,item.district,item.ca0,item.ca1,item.ca2,item.ca3,item.ca4,item.ca5,item.ca6,item.ca7,item.ca9,item.ca10,item.ca11,item.ca12,item.ca13,item.ca14,item.ca15,item.ca16,item.ca17,item.ca18,item.ca19,item.ca20,item.ca21,item.ca22,item.ca23,item.ca24]
    #     ws.append(tempArray)
        
    #    response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    #    response['Content-Disposition'] = 'attachment; filename="community_lqas_scorecard.xlsx"'

    #    return response
        
    
    context = {
    'lqasdatasets' : query_results,
    'targets' : query_results_targets, 
    'districtavg' : alldis,
    'period_desc': filter_period,
    'period_list': PERIOD_LIST,
    'district_list': DISTRICT_LIST,
    'year2015': year2015,
    'year2016': year2016,
    'year2017': year2017,
    #'excel_url': make_excel_url(request.path),
    #'csv_url': make_csv_url(request.path),
    } 
    return render(request, 'cannula/lqas_sites.html', context)

def avg(name):
    from django.db import connection
    cursor = connection.cursor()
    #cursor.execute('SELECT AVG ('+name+') FROM cannula_lqas_dataset;')
    #cursor.execute('SELECT to_char(AVG ('+name+'),\'99999999999999999D99\') FROM cannula_lqas_dataset;')
    cursor.execute('SELECT AVG ('+name+') FROM cannula_lqas_dataset;')
    return int(cursor.fetchone()[0])

def avgforyear(name,year):
    from django.db import connection
    cursor = connection.cursor()
    #cursor.execute('SELECT AVG ('+name+') FROM cannula_lqas_dataset;')
    #cursor.execute('SELECT to_char(AVG ('+name+'),\'99999999999999999D99\') FROM cannula_lqas_dataset where period=\''+year+'\';')
    cursor.execute('SELECT AVG ('+name+') FROM cannula_lqas_dataset where period=\''+year+'\';')
    return int(cursor.fetchone()[0])


